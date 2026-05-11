#!/usr/bin/env python3
"""Generate a multi-tab XLSX: one tab per week, day-by-day grid of every cadence's touch.

Tab layout per week:
  Row 1: title + week range
  Row 2: column headers — Account | DM/Phone | Trigger | Score | Mon date | Tue date | Wed date | Thu date | Fri date
  Rows 3+: one row per active cadence, with the scheduled touch action in each day column.
"""
import json, datetime as dt, base64, sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT = Path("/Users/ricknini/Documents/ADP PEO")
state = json.load(open(ROOT / "_sales_os_state.json"))
TODAY = dt.date(2026, 5, 11)
N_WEEKS = 2  # current week + next, kept small for Drive upload

# ───── Styles ─────
title_font = Font(bold=True, size=14, color="1F4E78")
h_font = Font(bold=True, size=11, color="FFFFFF")
h_fill = PatternFill("solid", fgColor="1F4E78")
warm_fill = PatternFill("solid", fgColor="FFD966")     # 🔥 warm followup
email_fill = PatternFill("solid", fgColor="DAE3F3")    # E
drop_fill = PatternFill("solid", fgColor="FFE699")     # D
ldap_fill = PatternFill("solid", fgColor="E2EFD9")     # LinkedIn
call_fill = PatternFill("solid", fgColor="F4B084")     # P
brk_fill = PatternFill("solid", fgColor="D9D9D9")      # Breakup
focus_fill = PatternFill("solid", fgColor="C6EFCE")    # in-focus highlight
done_font = Font(italic=True, color="909090", strike=True)

# Action label per channel
CH_LABEL = {
    "email":    "📧 EMAIL",
    "linkedin": "💼 LinkedIn",
    "drop":     "🚗 DROP",
    "call":     "📞 CALL",
    "breakup":  "✉️ BREAKUP",
}
CH_FILL = {
    "email": email_fill, "linkedin": ldap_fill, "drop": drop_fill,
    "call": call_fill, "breakup": brk_fill,
}

def week_dates(week_start):
    """Return list of (date, day_name) for Mon-Fri starting from week_start."""
    return [(week_start + dt.timedelta(days=i),
             ["Mon","Tue","Wed","Thu","Fri"][i]) for i in range(5)]

def find_touches_in_week(cadence, mon, fri):
    """Return {date_iso → touch} for any touches in [mon, fri]."""
    out = {}
    for t in cadence.get("touches", []):
        sf = t.get("scheduled_for","")
        if not sf: continue
        try:
            d = dt.date.fromisoformat(sf)
        except Exception:
            continue
        if mon <= d <= fri:
            out[sf] = t
    return out

wb = Workbook()
wb.remove(wb.active)

active = state["active_cadences"]
# Sort by score descending so highest-priority accounts at top of each weekly grid
active = sorted(active, key=lambda x: -x.get("score", 0))

# Build weekly tabs
this_monday = TODAY - dt.timedelta(days=TODAY.weekday())
for week_idx in range(N_WEEKS):
    wk_mon = this_monday + dt.timedelta(days=7 * week_idx)
    wk_fri = wk_mon + dt.timedelta(days=4)
    tab_name = f"Week_{wk_mon.strftime('%m-%d')}"
    ws = wb.create_sheet(tab_name)

    # Title row
    ws["A1"] = f"WEEK {week_idx+1}: {wk_mon.isoformat()} → {wk_fri.isoformat()}"
    ws["A1"].font = title_font
    ws.merge_cells("A1:I1")
    ws["A2"] = ("📧 Email · 💼 LinkedIn · 🚗 Drop · 📞 Call · ✉️ Breakup    |    "
                "💰=has health · 🌐=multi-state · 🔥=warm followup    |    "
                "Day cell = action that day · grey-strike = already done · blank = no touch this day")
    ws["A2"].font = Font(italic=True, color="606060", size=10)
    ws.merge_cells("A2:I2")
    ws.row_dimensions[2].height = 24
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")

    # Header row
    dates = week_dates(wk_mon)
    headers = ["Account", "DM / Phone", "Trigger / Signals", "Score"] + \
              [f"{nm} {d.strftime('%m/%d')}" for d, nm in dates]
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col_idx, value=h)
        c.font = h_font; c.fill = h_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 28
    ws.freeze_panes = "B5"

    # Row 5+: one row per active cadence that has at least one touch this week
    row = 5
    for c in active:
        in_week = find_touches_in_week(c, wk_mon, wk_fri)
        # Also include accounts with no touches this week IF the cadence is still active and started ≤21d ago
        try:
            start = dt.date.fromisoformat(c["start_date"])
        except Exception:
            start = TODAY
        days_in = (wk_fri - start).days
        active_in_week = -7 <= days_in <= 28   # cadence is "live" this week if it's started within ~30 days
        if not in_week and not active_in_week:
            continue

        # Account name + flags
        flags = []
        if c.get("status") == "warm_followup": flags.append("🔥")
        if c.get("has_health_benefits"): flags.append("💰")
        if c.get("multi_state_likely"): flags.append("🌐")
        if c.get("disqualify_recommendation"): flags.append("❌")
        flag_str = " ".join(flags)
        account_cell = ws.cell(row=row, column=1, value=f"{flag_str} {c['company'][:38]}".strip())
        if c.get("status") == "warm_followup":
            account_cell.fill = warm_fill
        elif c.get("has_health_benefits") or c.get("multi_state_likely"):
            account_cell.fill = focus_fill
        account_cell.font = Font(bold=True, size=11)

        # DM / Phone
        dm_line = c.get("dm_name") or "—"
        phone = c.get("phone","") or "—"
        ws.cell(row=row, column=2, value=f"{dm_line}\n{phone}").alignment = Alignment(wrap_text=True, vertical="top")

        # Trigger
        trig = c.get("primary_trigger","")
        if c.get("primary_trigger") == "displacement":
            trig += f"\n{c.get('incumbent_peo','')[:24]}"
        ws.cell(row=row, column=3, value=trig).alignment = Alignment(wrap_text=True, vertical="top")

        # Score
        score_cell = ws.cell(row=row, column=4, value=round(c.get("score",0),1))
        score_cell.alignment = Alignment(horizontal="center")
        if c.get("score",0) >= 40: score_cell.fill = PatternFill("solid", fgColor="C6EFCE")
        elif c.get("score",0) >= 30: score_cell.fill = PatternFill("solid", fgColor="FFEB9C")

        # Per-day cells
        for col_idx, (d, _) in enumerate(dates, 5):
            t = in_week.get(d.isoformat())
            if not t:
                ws.cell(row=row, column=col_idx, value="")
                continue
            label = CH_LABEL.get(t["channel"], "?")
            day_label = f"D{t['day']}"
            text = f"{label}\n{day_label}"
            notes = (t.get("notes","") or "")[:60]
            if notes: text += f"\n{notes}"
            cell = ws.cell(row=row, column=col_idx, value=text)
            cell.fill = CH_FILL.get(t["channel"], PatternFill("solid", fgColor="FFFFFF"))
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="center")
            if t.get("completed"):
                cell.font = done_font
                # Show outcome
                outcome = t.get("outcome","")
                if outcome:
                    cell.value = f"{label} ({outcome})\n{day_label}"

        # Row height — taller if multiple lines
        ws.row_dimensions[row].height = 48
        row += 1

    # Footer with weekly summary
    row += 1
    ws.cell(row=row, column=1, value=f"WEEKLY TARGETS: 45 touches · 3 meetings · 50 active accounts").font = Font(bold=True, italic=True, color="606060")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)

    # Column widths
    widths = [40, 20, 22, 7, 22, 22, 22, 22, 22]
    for col_letter, w in zip("ABCDEFGHI", widths):
        ws.column_dimensions[col_letter].width = w

# Save XLSX
OUT = ROOT / "Weekly_Cadence_Grid.xlsx"
wb.save(OUT)
print(f"Saved: {OUT}")
print(f"Tabs created: {wb.sheetnames}")
print(f"File size: {OUT.stat().st_size:,} bytes")

# Print base64 to stdout (for upload)
with OUT.open("rb") as f:
    b64 = base64.b64encode(f.read()).decode("ascii")
print(f"BASE64_LEN: {len(b64)}")
# Save b64 to a side file for the upload step
(ROOT / "_weekly_xlsx_b64.txt").write_text(b64)
print("Base64 saved to _weekly_xlsx_b64.txt for upload")
