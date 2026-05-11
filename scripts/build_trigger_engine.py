#!/usr/bin/env python3
"""
TRIGGER ENGINE — converts the enrichment SQLite into a ranked daily call list
plus per-trigger drill-downs and the weekly Operator Dashboard.

Run weekly (or whenever data refreshes). Output:
    Trigger_Engine_Output.xlsx

Usage:
    python3 scripts/build_trigger_engine.py

Methodology (anchored to Eastern_NC_HRO_Operator_Playbook.md §1.2 and §3.5):
    Each company gets a composite TRIGGER_SCORE = sum of signal weights
    × ICP multiplier × NAICS multiplier (0 = hard exclude).
    Top-ranked accounts that have NOT been worked recently surface to
    the Daily_Top_25 tab — that's the call list.
"""
from __future__ import annotations
import sqlite3, datetime as dt, re, os, sys
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = Path("/Users/ricknini/Documents/ADP PEO")
DB = ROOT / "enrichment/db/pipeline.sqlite"
OUT = ROOT / "Trigger_Engine_Output.xlsx"
WEEKLY_PIPELINE = ROOT / "ADP_Weekly_Pipeline_MASTER.xlsx"

TODAY = dt.date.today()
TODAY_ISO = TODAY.isoformat()

# === SIGNAL WEIGHTS (calibrate over time as you measure conversion) ===
# Higher = more likely to convert to first meeting.
WC_RENEWAL_WEIGHTS = [
    (30,  5.0, "WC renewal in next 30 days"),
    (60,  4.0, "WC renewal in 31–60 days"),
    (90,  3.5, "WC renewal in 61–90 days"),
    (120, 2.5, "WC renewal in 91–120 days"),
]
HEALTH_RENEWAL_WEIGHTS = [
    (60,  4.0, "Health (5500) renewal in next 60 days"),
    (120, 3.0, "Health renewal in 61–120 days"),
]
OSHA_RECENT_WEIGHT     = 4.5  # citation/injury in last 90 days
HIRING_VELOCITY_WEIGHT = 3.0  # 3+ open jobs on Indeed
ALE_THRESHOLD_WEIGHT   = 3.5  # 38-48 EE → ACA crossover risk
HAS_5500_WEIGHT        = 2.5  # filed own 5500 = has benefits = has budget
PEO_SCORE_TOP_WEIGHT   = 2.0  # composite signal already computed by prior pipeline

# ICP multipliers
ICP_EE_BAND_BONUS  = 1.5  # 11–50 EE
NEAR_ICP_BAND_BONUS = 1.2 # 51–55 EE
ICP_INDUSTRY_BONUS = 1.3  # manufacturing/engineering/HVAC/biotech/pharma/tech

# Hard exclusions (per Exclusion_Rules in weekly pipeline)
EXCLUDED_NAICS_PREFIXES = ("23", "484", "562")  # construction, trucking, waste
EXCLUDED_NAICS_EXACT    = {"561730"}  # landscape

ICP_NAICS_PREFIXES = {
    # Manufacturing
    "31": "Manufacturing", "32": "Manufacturing", "33": "Manufacturing",
    # Wholesale/distribution
    "42": "Wholesale Trade",
    # Transportation/logistics (excluding 484 trucking)
    "488": "Logistics support",
    "493": "Warehousing",
    # Tech / IT
    "51": "Information / Tech",
    "5415": "Computer systems design",
    # Finance/insurance
    "52": "Finance/Insurance",
    # Real estate (commercial)
    "531": "Real Estate",
    # Professional services (engineering 5413, legal 5411, mgmt 5416, R&D 5417)
    "5413": "Engineering / A&E",
    "5416": "Mgmt Consulting",
    "5417": "Scientific R&D / Biotech",
    "5419": "Other Prof Services",
    # Health care
    "62": "Health Care/Social Assist",
    # Hospitality (limited service)
    "7211": "Hotels (limited svc)",
    "7223": "Special food svc",
}

VERTICAL_ALIASES = {
    # Map NAICS prefix → vertical bucket for downstream segmentation
    "5413": "Engineering",
    "5417": "Biotech / Life Sci",
    "31": "Manufacturing", "32": "Manufacturing", "33": "Manufacturing",
    "51": "Tech",
    "5415": "Tech",
    "238": "Construction (excluded)",
    "484": "Trucking (excluded)",
    "62": "Health Care",
    "52": "Finance/Insurance",
    "5411": "Legal",
}
# HVAC isn't a clean NAICS prefix — common codes 238220, 238110, 238210
HVAC_NAICS = {"238220", "238210", "238110", "238290"}

# ──────────────────────────────────────────────────────────────────
def parse_date(s):
    """Best-effort parse of mixed-format dates from the enrichment table."""
    if not s: return None
    s = str(s).strip()
    # ISO 2026-10-15
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        try: return dt.date(int(m[1]), int(m[2]), int(m[3]))
        except ValueError: return None
    # US 12/31/2025
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})", s)
    if m:
        try: return dt.date(int(m[3]), int(m[1]), int(m[2]))
        except ValueError: return None
    return None

def to_int(v):
    try: return int(float(str(v).replace(",","")))
    except: return None

def vertical_for(naics):
    if not naics: return "Unknown"
    s = str(naics).strip()
    if s in HVAC_NAICS: return "HVAC / Plumbing"
    for k_len in (6, 4, 3, 2):
        prefix = s[:k_len]
        if prefix in VERTICAL_ALIASES: return VERTICAL_ALIASES[prefix]
        if prefix in ICP_NAICS_PREFIXES: return ICP_NAICS_PREFIXES[prefix]
    return f"NAICS {s[:3]}"

def is_excluded_naics(naics):
    if not naics: return False
    s = str(naics).strip()
    if s in EXCLUDED_NAICS_EXACT: return True
    return any(s.startswith(p) for p in EXCLUDED_NAICS_PREFIXES)

def is_icp_industry(naics):
    if not naics: return False
    s = str(naics).strip()
    if s in HVAC_NAICS: return True
    for k_len in (6, 4, 3, 2):
        if s[:k_len] in ICP_NAICS_PREFIXES: return True
    return False

# ──────────────────────────────────────────────────────────────────
def load_worked_accounts():
    """Pull EINs/names from the Worked_Accounts_Registry to exclude already-touched."""
    worked = set()
    try:
        wb = load_workbook(WEEKLY_PIPELINE, read_only=True, data_only=True)
        if "Worked_Accounts_Registry" not in wb.sheetnames:
            return worked
        ws = wb["Worked_Accounts_Registry"]
        hdr = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        ix = {str(h or "").strip().lower(): i for i, h in enumerate(hdr)}
        ein_col = next((ix[k] for k in ("ein","fein","tax_id") if k in ix), None)
        name_col = next((ix[k] for k in ("company","company_name","name","name_normalized") if k in ix), None)
        if name_col is None: return worked
        for row in ws.iter_rows(min_row=2, values_only=True):
            n = row[name_col]
            if n: worked.add(re.sub(r"[^a-z0-9]","", str(n).lower()))
            if ein_col is not None:
                e = row[ein_col]
                if e: worked.add(re.sub(r"\D","",str(e)))
        wb.close()
    except Exception as ex:
        print(f"  [warn] couldn't load Worked_Accounts_Registry: {ex}", file=sys.stderr)
    return worked

# ──────────────────────────────────────────────────────────────────
def main():
    print("Loading SQLite + Worked_Accounts_Registry…")
    db = sqlite3.connect(str(DB))
    db.row_factory = sqlite3.Row
    worked = load_worked_accounts()
    print(f"  {len(worked):,} worked-account keys to exclude")

    # 1. Pull every fresh-territory company
    rows = db.execute("""
        SELECT c.company_id, c.name_display, c.name_normalized, c.city, c.zip, c.county, c.state, c.domain
        FROM companies c
        WHERE c.company_id NOT IN (SELECT company_id FROM drops)
    """).fetchall()
    print(f"  {len(rows):,} non-dropped companies")

    # 2. Bulk-load enrichment fields we care about
    fields_of_interest = (
        "wc_expiration_date","wc_effective_date","wc_carrier","wc_status","wc_churn_status",
        "schedule_a_carriers","carrier_premium_total","health_carrier_name",
        "5500_plan_year","is_5500_filer","ee_count_dol5500","pdf_url_5500",
        "indeed_open_jobs","indeed_has_health","indeed_has_401k","hiring_signal",
        "osha_total_injuries","osha_trir","osha_cliff_score","osha_reporting_year","osha_annual_avg_employees",
        "ee_count","ee_band","naics_inferred","naics","naics_inferred_top3",
        "industry","phone","website",
        "peo_score","peo_tier","peo_rationale",
    )
    placeholders = ",".join("?" for _ in fields_of_interest)
    enrichments = defaultdict(dict)
    rows_e = db.execute(f"""
        SELECT company_id, field, value FROM enrichments
        WHERE field IN ({placeholders})
    """, fields_of_interest).fetchall()
    for r in rows_e:
        cid, fld, val = r["company_id"], r["field"], r["value"]
        # Take latest non-null value (simple approach — refine later if needed)
        if val is None or val == "": continue
        enrichments[cid][fld] = val
    print(f"  {len(enrichments):,} companies have ≥1 trigger-relevant enrichment")

    # 3. Compute trigger score per company
    print("\nComputing trigger scores…")
    scored = []
    for r in rows:
        cid = r["company_id"]
        nm = r["name_display"] or ""
        nm_key = re.sub(r"[^a-z0-9]","", nm.lower())
        if nm_key in worked: continue  # skip already-worked

        e = enrichments.get(cid, {})
        if not e: continue  # no signals at all → skip

        naics = e.get("naics_inferred") or e.get("naics") or ""
        if is_excluded_naics(naics): continue

        triggers = []
        score = 0.0

        # --- WC renewal trigger ---
        wc_exp = parse_date(e.get("wc_expiration_date"))
        if wc_exp:
            days = (wc_exp - TODAY).days
            for cap, w, label in WC_RENEWAL_WEIGHTS:
                if 0 <= days <= cap and not any(t.startswith("WC renewal") for t in [tt[0] for tt in triggers]):
                    score += w
                    triggers.append((label, w, f"renewal {wc_exp} ({days} days), carrier={e.get('wc_carrier','?')}"))
                    break

        # --- Health renewal trigger (5500-driven) ---
        health_carriers = e.get("schedule_a_carriers") or e.get("health_carrier_name")
        plan_year = e.get("5500_plan_year")
        if health_carriers and plan_year:
            # Most NC small-group plans renew 1/1, 4/1, 7/1, or 10/1
            # We don't have per-plan renewal date, but having Schedule A means budget exists
            # Treat as health renewal trigger if 5500 was filed for plan year >= prior calendar year
            try:
                py = int(str(plan_year)[:4])
                if py >= TODAY.year - 1:
                    # Approximate next renewal — assume 1/1 next year as default
                    score += HEALTH_RENEWAL_WEIGHTS[1][1]  # broader window
                    triggers.append(("Health (5500) renewal eligible", HEALTH_RENEWAL_WEIGHTS[1][1],
                                     f"5500 plan year {py}, carriers={str(health_carriers)[:40]}"))
            except (ValueError, TypeError):
                pass

        # --- OSHA recent hit ---
        osha_year = to_int(e.get("osha_reporting_year"))
        osha_inj  = to_int(e.get("osha_total_injuries"))
        if osha_year and osha_year >= TODAY.year - 1 and (osha_inj or 0) >= 1:
            score += OSHA_RECENT_WEIGHT
            triggers.append(("OSHA injury/citation in last 18 mo", OSHA_RECENT_WEIGHT,
                             f"{osha_inj} injuries reported {osha_year}"))

        # --- Hiring velocity ---
        open_jobs = to_int(e.get("indeed_open_jobs"))
        if open_jobs and open_jobs >= 3:
            score += HIRING_VELOCITY_WEIGHT
            triggers.append((f"Hiring velocity ({open_jobs} open Indeed jobs)", HIRING_VELOCITY_WEIGHT,
                             f"{open_jobs} jobs posted"))

        # --- ALE threshold (38-48 EE) ---
        ee = to_int(e.get("ee_count")) or to_int(e.get("ee_count_dol5500")) or to_int(e.get("osha_annual_avg_employees"))
        if ee and 38 <= ee <= 48:
            score += ALE_THRESHOLD_WEIGHT
            triggers.append(("Approaching 50-FTE / ACA threshold", ALE_THRESHOLD_WEIGHT,
                             f"{ee} EE — crossing 50 triggers ACA reporting"))

        # --- Has 5500 filer (budget signal) ---
        if e.get("is_5500_filer") == "y":
            score += HAS_5500_WEIGHT
            triggers.append(("5500 filer (has benefits/budget)", HAS_5500_WEIGHT, "filed own Form 5500"))

        # --- PEO score top quartile ---
        peo_score = e.get("peo_score")
        if peo_score:
            try:
                ps = float(peo_score)
                if ps >= 70:
                    score += PEO_SCORE_TOP_WEIGHT
                    triggers.append((f"High PEO fit score ({ps:.0f})", PEO_SCORE_TOP_WEIGHT,
                                     e.get("peo_rationale","")[:80]))
            except ValueError:
                pass

        if score == 0: continue  # no triggers fired

        # --- Apply multipliers ---
        ee_int = ee
        if ee_int and 11 <= ee_int <= 50:
            score *= ICP_EE_BAND_BONUS
        elif ee_int and 51 <= ee_int <= 55:
            score *= NEAR_ICP_BAND_BONUS

        if is_icp_industry(naics):
            score *= ICP_INDUSTRY_BONUS

        scored.append({
            "company_id": cid,
            "company": nm,
            "city": r["city"] or "",
            "zip": r["zip"] or "",
            "county": r["county"] or "",
            "state": r["state"] or "NC",
            "domain": r["domain"] or e.get("website",""),
            "phone": e.get("phone",""),
            "ee": ee_int,
            "naics": naics,
            "vertical": vertical_for(naics),
            "wc_carrier": e.get("wc_carrier",""),
            "wc_expiration": e.get("wc_expiration_date",""),
            "score": round(score, 2),
            "triggers": triggers,  # list of (label, weight, evidence)
            "trigger_summary": " | ".join(t[0] for t in triggers),
            "top_evidence": triggers[0][2] if triggers else "",
        })

    scored.sort(key=lambda x: -x["score"])
    print(f"  {len(scored):,} accounts with ≥1 trigger fired")

    # 4. Build workbook
    print("\nWriting workbook…")
    wb = Workbook()
    title_font = Font(bold=True, size=14, color="1F4E78")
    hdr_font = Font(bold=True, size=11, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    sub_fill = PatternFill("solid", fgColor="D9E2F3")
    hot_fill = PatternFill("solid", fgColor="FFE699")
    metric_font = Font(bold=True, size=18, color="1F4E78")

    # ===== Tab 1: Daily_Top_25 (THE CALL LIST) =====
    ws = wb.active; ws.title = "Daily_Top_25"
    ws["A1"] = f"DAILY CALL LIST — {TODAY_ISO} — {min(25, len(scored))} hottest trigger-warm accounts"
    ws["A1"].font = title_font; ws.merge_cells("A1:K1")
    ws["A2"] = "Sorted by trigger score. Each row has a defensible opener — anchor every touch to the trigger evidence."
    ws["A2"].font = Font(italic=True, color="606060"); ws.merge_cells("A2:K2")

    hdr = ["#","Score","Company","City","ZIP","Phone","EE","Vertical","Top Trigger","Evidence","Opener Suggestion"]
    for i, h in enumerate(hdr, 1):
        c = ws.cell(row=4, column=i, value=h); c.font = hdr_font; c.fill = hdr_fill
    ws.freeze_panes = "A5"

    def opener_for(rec):
        triggers = rec["triggers"]
        if not triggers: return ""
        primary_label = triggers[0][0]
        if "WC renewal" in primary_label:
            return (f"Pulled your WC coverage from the NC Industrial Commission database — "
                    f"looks like your policy with {rec['wc_carrier'] or '[carrier]'} renews "
                    f"{rec['wc_expiration'] or 'soon'}. Most owners in {rec['vertical']} "
                    f"are seeing 8–15% increases this cycle. Worth a 20-min conversation before you sign?")
        if "Health" in primary_label:
            return (f"Pulled your 5500 — {triggers[0][2]}. Industry-wide small-group renewals "
                    f"this cycle are running 15–30%. The TotalSource master plan typically produces "
                    f"a 5–18% reduction vs. standalone renewal. Worth running the comparison?")
        if "OSHA" in primary_label:
            return (f"Saw the OSHA filing on the public database. The reason I'm reaching out: "
                    f"most owners after a citation realize they don't have in-house bandwidth to "
                    f"actually get and stay compliant. We run an outsourced compliance + safety "
                    f"program with on-site OSHA-30 training. 20 min?")
        if "Hiring velocity" in primary_label:
            return (f"Noticed {triggers[0][2]}. Most owners in {rec['vertical']} tell me hiring "
                    f"is the bottleneck. Inside TotalSource you offer Fortune-500 benefits a "
                    f"competitor can't match — used as a recruiting weapon, not a cost center. "
                    f"Worth a quick comparison?")
        if "50-FTE" in primary_label or "ACA" in primary_label:
            return (f"Noticed you're at {rec['ee']} EE. When companies in {rec['vertical']} cross "
                    f"50 FTE, ACA reporting adds ~100 hrs/yr of admin and creates 1095-C penalty "
                    f"exposure. We handle that transition without you adding headcount. "
                    f"Open to comparing notes?")
        if "5500 filer" in primary_label:
            return (f"You filed a 5500 last year — so I know you have benefits and a budget to "
                    f"manage. Most owners in {rec['vertical']} at {rec['ee'] or '[your size]'} EE "
                    f"don't realize TotalSource consolidates 401(k) admin, health, WC, and HR "
                    f"under one master plan. 15 min?")
        return f"Trigger: {primary_label}. {triggers[0][2]}"

    for i, rec in enumerate(scored[:25], 5):
        rank = i - 4
        vals = [rank, rec["score"], rec["company"][:50], rec["city"], rec["zip"], rec["phone"],
                rec["ee"] or "", rec["vertical"], rec["triggers"][0][0] if rec["triggers"] else "",
                rec["top_evidence"][:70], opener_for(rec)]
        for j, v in enumerate(vals, 1):
            ws.cell(row=i, column=j, value=v)
        if rec["score"] >= 8.0:
            for j in range(1, len(hdr)+1):
                ws.cell(row=i, column=j).fill = hot_fill

    widths = [4, 7, 38, 18, 7, 16, 6, 22, 32, 50, 80]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ===== Tab 2: WC_Renewals_120d =====
    ws2 = wb.create_sheet("WC_Renewals_120d")
    ws2["A1"] = "WC renewals in next 120 days — sorted by days-until-renewal (urgent first)"
    ws2["A1"].font = title_font; ws2.merge_cells("A1:J1")
    h2 = ["Company","City","ZIP","Phone","EE","Vertical","WC Carrier","Renewal","Days Out","Score"]
    for i, h in enumerate(h2, 1):
        c = ws2.cell(row=3, column=i, value=h); c.font = hdr_font; c.fill = hdr_fill
    ws2.freeze_panes = "A4"
    wc_rows = []
    for rec in scored:
        for label, _, _ in rec["triggers"]:
            if label.startswith("WC renewal"):
                wc_exp = parse_date(rec["wc_expiration"])
                days = (wc_exp - TODAY).days if wc_exp else 999
                wc_rows.append((days, rec))
                break
    wc_rows.sort(key=lambda x: (x[0], -x[1]["score"]))
    for r, (days, rec) in enumerate(wc_rows, 4):
        ws2.cell(row=r, column=1, value=rec["company"][:50])
        ws2.cell(row=r, column=2, value=rec["city"])
        ws2.cell(row=r, column=3, value=rec["zip"])
        ws2.cell(row=r, column=4, value=rec["phone"])
        ws2.cell(row=r, column=5, value=rec["ee"])
        ws2.cell(row=r, column=6, value=rec["vertical"])
        ws2.cell(row=r, column=7, value=rec["wc_carrier"])
        ws2.cell(row=r, column=8, value=rec["wc_expiration"])
        ws2.cell(row=r, column=9, value=days)
        ws2.cell(row=r, column=10, value=rec["score"])
    for i, w in enumerate([38,18,7,16,6,22,22,12,8,7], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # ===== Tab 3: Health_Renewals =====
    ws3 = wb.create_sheet("Health_Renewals")
    ws3["A1"] = "5500-driven health renewal targets — Schedule A carriers + plan year ≥ last calendar year"
    ws3["A1"].font = title_font; ws3.merge_cells("A1:I1")
    h3 = ["Company","City","ZIP","Phone","EE","Vertical","Carriers (Sched A)","Plan Year","Score"]
    for i, h in enumerate(h3, 1):
        c = ws3.cell(row=3, column=i, value=h); c.font = hdr_font; c.fill = hdr_fill
    ws3.freeze_panes = "A4"
    health_rows = [rec for rec in scored if any("Health" in t[0] for t in rec["triggers"])]
    health_rows.sort(key=lambda x: -x["score"])
    for r, rec in enumerate(health_rows, 4):
        ws3.cell(row=r, column=1, value=rec["company"][:50])
        ws3.cell(row=r, column=2, value=rec["city"])
        ws3.cell(row=r, column=3, value=rec["zip"])
        ws3.cell(row=r, column=4, value=rec["phone"])
        ws3.cell(row=r, column=5, value=rec["ee"])
        ws3.cell(row=r, column=6, value=rec["vertical"])
        ev = next((t[2] for t in rec["triggers"] if "Health" in t[0]), "")
        ws3.cell(row=r, column=7, value=ev[:60])
        ws3.cell(row=r, column=8, value=ev.split("year ")[-1].split(",")[0] if "year " in ev else "")
        ws3.cell(row=r, column=9, value=rec["score"])
    for i, w in enumerate([38,18,7,16,6,22,40,12,7], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # ===== Tab 4: OSHA_Hot =====
    ws4 = wb.create_sheet("OSHA_Hot")
    ws4["A1"] = "OSHA-recent accounts (injury or citation in last 18 mo) — wounded, looking for help"
    ws4["A1"].font = title_font; ws4.merge_cells("A1:H1")
    h4 = ["Company","City","ZIP","Phone","EE","Vertical","Evidence","Score"]
    for i, h in enumerate(h4, 1):
        c = ws4.cell(row=3, column=i, value=h); c.font = hdr_font; c.fill = hdr_fill
    ws4.freeze_panes = "A4"
    osha_rows = [rec for rec in scored if any("OSHA" in t[0] for t in rec["triggers"])]
    osha_rows.sort(key=lambda x: -x["score"])
    for r, rec in enumerate(osha_rows, 4):
        ws4.cell(row=r, column=1, value=rec["company"][:50])
        ws4.cell(row=r, column=2, value=rec["city"])
        ws4.cell(row=r, column=3, value=rec["zip"])
        ws4.cell(row=r, column=4, value=rec["phone"])
        ws4.cell(row=r, column=5, value=rec["ee"])
        ws4.cell(row=r, column=6, value=rec["vertical"])
        ws4.cell(row=r, column=7, value=next((t[2] for t in rec["triggers"] if "OSHA" in t[0]), ""))
        ws4.cell(row=r, column=8, value=rec["score"])
    for i, w in enumerate([38,18,7,16,6,22,40,7], 1):
        ws4.column_dimensions[get_column_letter(i)].width = w

    # ===== Tab 5: Hiring_Velocity =====
    ws5 = wb.create_sheet("Hiring_Velocity")
    ws5["A1"] = "Active hiring (3+ open jobs on Indeed) — capacity stretched, HR overhead climbing"
    ws5["A1"].font = title_font; ws5.merge_cells("A1:H1")
    h5 = ["Company","City","ZIP","Phone","EE","Vertical","Evidence","Score"]
    for i, h in enumerate(h5, 1):
        c = ws5.cell(row=3, column=i, value=h); c.font = hdr_font; c.fill = hdr_fill
    ws5.freeze_panes = "A4"
    hire_rows = [rec for rec in scored if any("Hiring" in t[0] for t in rec["triggers"])]
    hire_rows.sort(key=lambda x: -x["score"])
    for r, rec in enumerate(hire_rows, 4):
        ws5.cell(row=r, column=1, value=rec["company"][:50])
        ws5.cell(row=r, column=2, value=rec["city"])
        ws5.cell(row=r, column=3, value=rec["zip"])
        ws5.cell(row=r, column=4, value=rec["phone"])
        ws5.cell(row=r, column=5, value=rec["ee"])
        ws5.cell(row=r, column=6, value=rec["vertical"])
        ws5.cell(row=r, column=7, value=next((t[2] for t in rec["triggers"] if "Hiring" in t[0]), ""))
        ws5.cell(row=r, column=8, value=rec["score"])
    for i, w in enumerate([38,18,7,16,6,22,40,7], 1):
        ws5.column_dimensions[get_column_letter(i)].width = w

    # ===== Tab 6: Operator_Dashboard =====
    ws6 = wb.create_sheet("Operator_Dashboard")
    ws6["A1"] = f"OPERATOR DASHBOARD — Week of {TODAY_ISO}"
    ws6["A1"].font = title_font; ws6.merge_cells("A1:F1")
    ws6["A2"] = "Update Friday afternoon. Single number to optimize: STAGE-PROGRESSIONS-PER-WEEK (target 8–12)."
    ws6["A2"].font = Font(italic=True, color="606060"); ws6.merge_cells("A2:F2")

    ws6["A4"] = "WEEKLY LEADING INDICATORS (inputs)"; ws6["A4"].font = hdr_font; ws6["A4"].fill = hdr_fill
    ws6.merge_cells("A4:E4")
    li = [("Total touches", 50), ("Cold touches", 25), ("Warm touches (referral/CPA/broker)", 15),
          ("Field drops", 10), ("New triggered accounts identified", 5),
          ("★ Stage progressions (≥1 stage advance)", "8–12"),
          ("Discovery conversations", 3), ("Business Profiles collected", 1)]
    ws6["A5"] = "Metric"; ws6["B5"] = "Target"; ws6["C5"] = "This Week"; ws6["D5"] = "Last Week"; ws6["E5"] = "Trend"
    for c in "ABCDE": ws6[f"{c}5"].font = Font(bold=True); ws6[f"{c}5"].fill = sub_fill
    for i, (m, t) in enumerate(li, 6):
        ws6.cell(row=i, column=1, value=m)
        ws6.cell(row=i, column=2, value=t)
        if "Stage progressions" in m:
            ws6.cell(row=i, column=1).font = Font(bold=True, color="C00000")

    row = 6 + len(li) + 2
    ws6.cell(row=row, column=1, value="WEEKLY OUTCOME METRICS (outputs)").font = hdr_font
    ws6.cell(row=row, column=1).fill = hdr_fill
    ws6.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 1
    om = [("First meetings booked", 3), ("First meetings held (vs cancelled)", 2.5),
          ("Proposals/quotes delivered", 0.5), ("Closes signed", "—")]
    ws6.cell(row=row, column=1, value="Metric").font = Font(bold=True); ws6.cell(row=row, column=1).fill = sub_fill
    ws6.cell(row=row, column=2, value="Target").font = Font(bold=True); ws6.cell(row=row, column=2).fill = sub_fill
    ws6.cell(row=row, column=3, value="This Week").font = Font(bold=True); ws6.cell(row=row, column=3).fill = sub_fill
    row += 1
    for m, t in om:
        ws6.cell(row=row, column=1, value=m); ws6.cell(row=row, column=2, value=t); row += 1

    row += 2
    ws6.cell(row=row, column=1, value="MONTHLY FUNNEL DIAGNOSTICS").font = hdr_font
    ws6.cell(row=row, column=1).fill = hdr_fill
    ws6.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 1
    fr = [("Touches → meetings booked", "~6%", "<3% = message broken"),
          ("Meetings booked → meetings held", "65–75%", "<50% = confirmation broken"),
          ("Meetings held → discovery completed", "80%+", "<60% = qualification too loose"),
          ("Discovery → proposal", "30–50%", "<20% = discovery shallow"),
          ("Proposal → close", "25–40%", "<15% = MEDDPICC gaps")]
    ws6.cell(row=row, column=1, value="Ratio").font = Font(bold=True); ws6.cell(row=row, column=1).fill = sub_fill
    ws6.cell(row=row, column=2, value="Target").font = Font(bold=True); ws6.cell(row=row, column=2).fill = sub_fill
    ws6.cell(row=row, column=3, value="Diagnosis if below").font = Font(bold=True); ws6.cell(row=row, column=3).fill = sub_fill
    row += 1
    for r_, t_, d_ in fr:
        ws6.cell(row=row, column=1, value=r_); ws6.cell(row=row, column=2, value=t_); ws6.cell(row=row, column=3, value=d_); row += 1

    row += 2
    ws6.cell(row=row, column=1, value="THE SINGLE NUMBER TO OPTIMIZE").font = title_font
    ws6.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 1
    ws6.cell(row=row, column=1, value="Stage progressions per week").font = Font(bold=True, size=12)
    ws6.cell(row=row, column=2, value="Target").font = Font(bold=True, size=12)
    ws6.cell(row=row, column=3, value="This Week").font = Font(bold=True, size=12)
    row += 1
    ws6.cell(row=row, column=1, value="(accounts that advanced ≥1 stage)").font = Font(italic=True, color="606060")
    ws6.cell(row=row, column=2, value="8–12").font = metric_font
    ws6.cell(row=row, column=3, value="—").font = metric_font

    for col, w in zip("ABCDE", [42, 12, 14, 14, 14]):
        ws6.column_dimensions[col].width = w

    # ===== Tab 7: README =====
    ws7 = wb.create_sheet("README")
    ws7["A1"] = "Trigger Engine — README"; ws7["A1"].font = title_font
    readme = [
        "",
        f"Generated: {TODAY_ISO}",
        f"Source DB: enrichment/db/pipeline.sqlite",
        "",
        "REFRESH:",
        "  python3 scripts/build_trigger_engine.py",
        "  (overwrites Trigger_Engine_Output.xlsx in this folder)",
        "",
        "WHAT THIS DOES:",
        "  Reads every trigger-relevant enrichment field, computes a composite score per",
        "  account, applies ICP multipliers (EE band, industry), excludes already-worked",
        "  accounts (from Worked_Accounts_Registry in ADP_Weekly_Pipeline_MASTER.xlsx) and",
        "  excluded NAICS (23* construction, 484* trucking, 562* waste, 561730 landscape).",
        "",
        "TABS:",
        "  • Daily_Top_25 — TODAY's call list. Each row has a defensible opener.",
        "  • WC_Renewals_120d — every territory account with WC renewal in next 120 days.",
        "  • Health_Renewals — Schedule A health-renewal targets.",
        "  • OSHA_Hot — accounts with OSHA injury/citation in last 18 mo.",
        "  • Hiring_Velocity — accounts with 3+ open Indeed jobs.",
        "  • Operator_Dashboard — your weekly leading-indicator tracker.",
        "",
        "HOW TO USE (Monday morning workflow):",
        "  1. Open Daily_Top_25 — that's your 25-account call list for the day.",
        "  2. Each row's 'Opener Suggestion' is the trigger-anchored opening line.",
        "  3. After each touch: log it in Worked_Accounts_Registry (ADP_Weekly_Pipeline_MASTER).",
        "  4. Friday: update the Operator_Dashboard tab. Track stage-progressions/wk.",
        "",
        "TUNING THE SCORE:",
        "  Edit the SIGNAL WEIGHTS section at the top of build_trigger_engine.py.",
        "  After 4 weeks of data, recalibrate by computing actual conversion-to-meeting",
        "  per trigger type and re-weighting up/down.",
        "",
        "WHAT THIS DOES NOT YET DO (data gaps to close):",
        "  • LinkedIn 'Office Mom departure' detection (needs LI Sales Nav pull)",
        "  • NC SoS recent member-change scraping (acquisition signal)",
        "  • Building-permit ingestion (new-location signal)",
        "  • CPA/broker channel partner tracking (separate workbook)",
        "",
    ]
    for i, line in enumerate(readme, 2):
        ws7.cell(row=i, column=1, value=line)
    ws7.column_dimensions["A"].width = 100

    wb.save(OUT)
    print(f"\nSaved: {OUT}")
    print(f"\n=== TODAY'S TOP 5 PREVIEW ===")
    for i, rec in enumerate(scored[:5], 1):
        print(f"  {i}. {rec['company'][:42]:42} | {rec['city'][:12]:12} | EE={rec['ee'] or '?':>4} | "
              f"score={rec['score']:>5.1f} | {rec['triggers'][0][0][:40]}")
    print(f"\nTotal scored accounts: {len(scored):,}")
    print(f"Hot accounts (score ≥ 8.0): {sum(1 for r in scored if r['score'] >= 8.0):,}")
    print(f"Warm accounts (score 5.0–7.9): {sum(1 for r in scored if 5 <= r['score'] < 8):,}")

if __name__ == "__main__":
    main()
