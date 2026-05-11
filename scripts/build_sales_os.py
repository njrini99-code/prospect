#!/usr/bin/env python3
"""
SALES OPERATING SYSTEM — daily orchestrator for Nick's ADP TotalSource territory.

Reads:
  - enrichment/db/pipeline.sqlite (trigger data)
  - Trigger_Engine_Output.xlsx (ranked universe — must run build_trigger_engine.py first)
  - ADP_Weekly_Pipeline_MASTER.xlsx → Worked_Accounts_Registry (excludes already-worked)
  - _sales_os_state.json (active cadences + history; auto-creates)
  - _charleston_territory_zips.csv NOT used (this is Nick's NC territory; uses zip_centroids in DB)

Writes:
  - Sales_OS_MASTER.xlsx (8 tabs — re-generated each run)
  - _sales_os_state.json (persists active cadence state)

Runs:
  python3 scripts/build_sales_os.py             # daily — generates today's queue
  python3 scripts/build_sales_os.py --init      # day 1 — seeds 50 accounts into cadence
  python3 scripts/build_sales_os.py --advance N # advance state by N days (testing)

The system architecture:
  4-touch give-give-give-ask cadence per account:
    Day 0 → email (give: industry brief)
    Day 4 → LinkedIn DM (give: their specific data)
    Day 9 → drop-by (give: printed analysis)
    Day 14 → phone call (ask: 20-min meeting)

  Cadences are stagger-started by territory route day:
    Mon → Wake/Triangle cluster
    Tue → Pitt/Greenville cluster
    Wed → Northern (Nash/Edgecombe/Wilson/Halifax)
    Thu → Cumberland/Sandhills (alt-week Coastal Carteret/Craven)
    Fri → no new cadences (CRM hygiene + planning + giving)

  10 new cadences per active day × 4 active days = 40 new accounts/wk.
  Each cadence runs 14 days → ~80 active cadences in steady state.
  Use the 50-account portfolio rule: keep top 50 by score in active state at any time.
"""
from __future__ import annotations
import sqlite3, json, datetime as dt, re, sys, os, math, functools
from pathlib import Path
from collections import defaultdict, Counter
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path("/Users/ricknini/Documents/ADP PEO")
DB = ROOT / "enrichment/db/pipeline.sqlite"
TRIGGER_OUT = ROOT / "Trigger_Engine_Output.xlsx"
WEEKLY_PIPELINE = ROOT / "ADP_Weekly_Pipeline_MASTER.xlsx"
STATE_FILE = ROOT / "_sales_os_state.json"
OUT = ROOT / "Sales_OS_MASTER.xlsx"
COMPETING_PEO_CSV = ROOT / "MASTER/uploads_to_drive/ALL_COMPETING_PEO_ACCOUNTS.csv"

# Allow date override for simulation/backfill: --simulate-date YYYY-MM-DD
def _today_with_override():
    for i, a in enumerate(sys.argv):
        if a == "--simulate-date" and i + 1 < len(sys.argv):
            try:
                return dt.date.fromisoformat(sys.argv[i + 1])
            except Exception:
                pass
        if a.startswith("--simulate-date="):
            try:
                return dt.date.fromisoformat(a.split("=", 1)[1])
            except Exception:
                pass
    return dt.date.today()

TODAY = _today_with_override()
TODAY_ISO = TODAY.isoformat()
WEEKDAY = TODAY.weekday()  # 0=Mon, 4=Fri

# ─── ROUTE / DAY CONSTANTS ──────────────────────────────────────────
ROUTE_BY_WEEKDAY = {
    0: ("Wake/Triangle (Wake, Durham, Orange, Johnston)", {"27", "28"}),  # zip prefixes are loose; refined per-zip below
    1: ("Pitt/Greenville crescent + Lenoir", set()),
    2: ("Northern (Nash, Edgecombe, Wilson, Halifax)", set()),
    3: ("Cumberland/Sandhills (alt: Coastal Carteret/Craven)", set()),
    4: ("Friday — CRM + planning + giving (no new cadences)", set()),
}
ROUTE_NAMES = {0: "wake", 1: "pitt", 2: "northern", 3: "cumberland", 4: "friday"}

# County-based clustering (more precise than zip prefix)
WAKE_COUNTIES = {"Wake", "Durham", "Orange", "Johnston", "Chatham", "Granville", "Person", "Vance", "Warren", "Franklin"}
PITT_COUNTIES = {"Pitt", "Greene", "Lenoir", "Wayne", "Wilson"}
NORTHERN_COUNTIES = {"Nash", "Edgecombe", "Halifax", "Northampton", "Hertford", "Bertie", "Martin"}
CUMBERLAND_COUNTIES = {"Cumberland", "Sampson", "Bladen", "Robeson", "Hoke", "Moore", "Lee", "Harnett", "Scotland", "Richmond"}
COASTAL_COUNTIES = {"Carteret", "Craven", "Pamlico", "Beaufort", "Hyde", "Tyrrell", "Washington", "Dare", "Currituck", "Camden", "Pasquotank", "Perquimans", "Chowan", "Gates", "Onslow", "Pender", "New Hanover", "Brunswick", "Columbus", "Duplin", "Jones"}

def county_to_route_day(county: str) -> int:
    """Returns weekday (0-4) when this county should be visited."""
    if not county: return -1
    c = county.strip().title()
    if c in WAKE_COUNTIES: return 0
    if c in PITT_COUNTIES: return 1
    if c in NORTHERN_COUNTIES: return 2
    if c in CUMBERLAND_COUNTIES:
        # Alt-week: even week = Cumberland, odd week = Coastal
        return 3
    if c in COASTAL_COUNTIES:
        return 3  # share the slot, alternate weekly
    return -1

# ─── ICP EXCLUSIONS (per CLAUDE.md + 2026-05 focus refinement) ──────────────────
# HVAC (23822) and Electric (23821) are EXPLICITLY IN-ICP per Nick's 2026-05-10 guidance.
# Construction proper (236, 237) and most other specialty trades stay excluded.
EXCLUDE_NAICS_PREFIXES = (
    "236",          # building construction
    "237",          # heavy/civil construction
    "23811",        # poured concrete foundation
    "23812",        # structural steel
    "23813",        # framing carpentry
    "23814",        # masonry
    "23815",        # glass & glazing
    "23816",        # roofing
    "23817",        # siding
    "23818",        # other structure exterior
    "23819",        # other building exterior
    "23829",        # other building equipment (NOT 23821 elec or 23822 HVAC)
    "23831",        # drywall
    "23832",        # painting
    "23833",        # flooring
    "23834",        # tile
    "23835",        # finish carpentry
    "23839",        # other building finishing
    "23891",        # site prep
    "23899",        # other specialty trade
    "484",          # truck transportation
    "562",          # waste/remediation
    "711",          # arts/entertainment
    "712",          # museums/zoos
    "722",          # food service
    "812",          # personal/laundry services
    "445",          # food/beverage retail
    "446",          # health/personal care retail
    "448",          # clothing retail
)
EXCLUDE_NAICS_EXACT = {"561730"}  # landscaping
EXCLUDE_NAME_REGEX = re.compile(
    # Construction proper (NOT HVAC, NOT Electric — those are in-ICP)
    r"\b(construction|contractor|contracting|builders?|"
    # Trucking / freight / heavy haul
    r"trucking|freight|hauling|truck\s*stop|truck\s*leasing|dry\s*bulk|"
    # Construction trades (NOT HVAC, NOT Electric)
    r"excavat\w*|paving|roof(?:ing|ers?)|plumb\w+|"
    r"drywall|concrete|mason\w*|carpent\w*|paint(?:ing|ers)|janitorial|"
    r"landscap\w*|lawn\s+care|"
    # Food / hospitality / arts / retail-service
    r"catering|cuisine|restaurant\w*|cafe\w*|bistro|pizz\w+|food\s+truck|"
    r"brewery|brewing\s+co|winery|distillery|"
    r"theater|theatre|cinema|gallery|salon|barber|tanning|nail\s+spa|"
    r"jewel(?:ers|ry)|wine\s+(?:shop|store|cellar)|"
    # Industry-edge cases (audit 2026-05-10)
    r"insurance\s+agency|wealth\s+management|"
    # Public sector + healthcare institutions (out of focus per 2026-05)
    r"public\s+school|county\s+(?:public\s+)?school|"
    r"city\s+of|town\s+of|department\s+of|university|community\s+college|"
    r"\bhospital\b|medical\s+center|health\s+system|"
    # Trade associations (members, not employers in our sense)
    r"\w+\s+association\b|trade\s+association|chamber\s+of\s+commerce|"
    r"roundtable|"
    # Retail / personal services (extended 2026-05-10 audit)
    r"\bnails?\b|\boptometry\b|\bvision\b|eye\s+care|eyecare|optical|"
    r"olive\s+oil|\bgourmet\b|specialty\s+food|"
    r"auto\s+group|auto\s+sales|auto\s+dealer|car\s+dealer|"
    r"collection\s+service|debt\s+collect|"
    r"learning\s+center|tutoring|"
    r"behavioral\s+(?:services|health)|psychiatric|"
    r"\bpharmacy\b|optometr\w*|\bdds\b|dental(?:\s+(?:office|clinic|practice))?"
    r")\b",
    re.I,
)
# UNC/Duke mail-drop ZIPs (per memory: biotechs at these ZIPs usually don't operate in NC)
UNC_DUKE_MAILDROPS = {"27599", "27710", "27705", "27708", "27704"}
RTP_PROPER_ZIP = "27709"
EXCLUDE_ZIPS = UNC_DUKE_MAILDROPS | {RTP_PROPER_ZIP}

# ─── ICP FOCUS BOOST (per Nick 2026-05-10) ──────────────────────────────────────
# Focus verticals get +50% score; off-focus verticals get -40%.
ICP_FOCUS_VERTICAL_PATTERNS = (
    "engineering", "manufactur", "tech", "saas", "software", "it ",
    "consulting", "mgmt cons", "professional services", "prof services",
    "hvac", "electric",
)
ICP_OFFOCUS_VERTICAL_PATTERNS = (
    "health care", "healthcare", "biotech", "life sci", "pharma",
    "retail", "hotel", "hospitality", "real estate", "finance", "insurance",
    "wholesale trade", "naics 7", "naics 8", "naics 4",
)
ICP_FOCUS_NAICS_PREFIXES = (
    "31", "32", "33",   # manufacturing
    "23821",            # electrical contractors
    "23822",            # HVAC / plumbing-heating-AC
    "5413",             # engineering services
    "5415", "5112", "5182",  # IT / software / hosting
    "5414", "5416", "5417", "5418", "5419",  # other professional services
)

# HARD-DROP name patterns — these accounts get dropped at load time, not just penalized.
# Added 2026-05-10 audit: catches Sylvan, Quality Mart, Pinehurst Coins, J.R. Furniture,
# THE FIDELITY BANK, holdings shells, and similar clearly-off-focus rows that slipped through.
OFF_FOCUS_NAME_PATTERNS = (
    # Education / tutoring
    "sylvan", "kumon", "mathnasium", "tutoring", "learning center",
    # Retail / convenience
    " mart", "quality mart", "convenience store", "gas station", "truck stop",
    "furniture", "appliance", "mattress", "carpet",
    "coins", "pawn", "hobby", "thrift", "antique",
    "grocery", "supermarket", "produce ", " soap ", "soap bar", "soap co",
    "museum", "aquarium",
    # Banking / finance (out of focus per 2026-05)
    " bank", "credit union", "savings & loan", "mortgage",
    " finance ", "finance inc", "financial services", "financial group",
    # Logistics / freight (narrowed out per 2026-05 — Nick's focus excludes logistics)
    "logistics", "freight", "fulfillment center",
    # Furniture / upholstery
    "upholstery", "furnishings",
    # Auto retail / service (not auto mfg)
    "auto group", "auto sales", "auto dealer", "car dealer",
    "automotive service", "tire ", "muffler", "transmission",
    # Holdings / shells (generic, hard to underwrite)
    "holdings, llc", " holdings llc",
    # Healthcare retail / specialty practices
    "behavioral services", "behavioral health", "psychiatric",
    "therapy", "developmental",
    "optometry", "optical", "vision care",
    "pharmacy", "drug store",
    # Food retail / craft beverage
    "olive oil", "gourmet", "specialty food",
    "wine ", "winery", "vineyard", "distillery", "spirits",
    # Personal services
    "salon", " spa ", "boutique", "tanning",
    "dry clean", "laundry", "cleaners",
    # Misc retail clues
    "marketplace", "naturally", " organic ", "pet ",
    # Religious / non-profit
    "ministry", " church", "fellowship",
)

def is_off_focus_by_name(name: str) -> bool:
    if not name: return False
    n = name.lower()
    return any(p in n for p in OFF_FOCUS_NAME_PATTERNS)
# Public/national-brand strings (no word-boundary tricks needed — exact substring match)
EXCLUDE_NAME_SUBSTR = (
    "LENOVO", "PEP BOYS", "ABC SUPPLY", "CATALENT", "AISIN", "AMAZON", "WALMART",
    "TARGET CORP", "HOME DEPOT", "LOWE'S", "FEDEX", "UPS INC", "GOODWILL",
    "MCDONALD", "STARBUCKS", "DOLLAR GENERAL", "DOLLAR TREE",
    # National beverage chains (Fortune 500)
    "COCA-COLA", "COCA COLA", "PEPSI", "DR PEPPER", "ANHEUSER",
    # NC state institutions
    "CHERRY HOSPITAL", "DOROTHEA DIX", "BROUGHTON HOSPITAL",
)
# ICP EE band: 11–55 (Near-ICP cap)
ICP_EE_MIN = 11
ICP_EE_MAX = 55
EXCLUDE_VERTICAL_PREFIXES = (
    "NAICS 23", "NAICS 484", "NAICS 562", "NAICS 711", "NAICS 712",
    "NAICS 722", "NAICS 812",
)
EXCLUDE_VERTICAL_EXACT = {"NAICS 561730"}
EXCLUDE_DROP_REASONS = {
    "IN_REGISTRY","IN_CRM","IS_ADP_CLIENT","IS_COMPETING_PEO","IS_COMPETING_PEO_ORG",
    "LATE_ACQUISITION","ACQUIRED","DEAD","DOA_UNVERIFIABLE","HQ_OUT_OF_STATE",
    "OUT_OF_TERRITORY","VIRTUAL_MAILBOX","RETAIL_NOISE","EXCLUSION_RULE",
    "ENTERPRISE_PARENT","ABOVE_ICP_EE",
}

def _norm_key(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", (s or "").lower())

def is_naics_excluded(naics: str) -> bool:
    if not naics: return False
    n = str(naics).strip()
    if n in EXCLUDE_NAICS_EXACT: return True
    return any(n.startswith(p) for p in EXCLUDE_NAICS_PREFIXES)

@functools.cache
def load_excluded_keys():
    """Set of normalized company-name keys to never seed: pipeline drops + Worked_Accounts_Registry."""
    keys = set()
    if DB.exists():
        try:
            db = sqlite3.connect(str(DB))
            placeholders = ",".join(["?"] * len(EXCLUDE_DROP_REASONS))
            for (n,) in db.execute(
                f"SELECT c.name_display FROM drops d "
                f"JOIN companies c ON c.company_id = d.company_id "
                f"WHERE d.reason_code IN ({placeholders})",
                tuple(EXCLUDE_DROP_REASONS),
            ).fetchall():
                k = _norm_key(n)
                if k: keys.add(k)
            db.close()
        except Exception as e:
            print(f"  [warn] could not read pipeline drops: {e}", file=sys.stderr)
    if WEEKLY_PIPELINE.exists():
        try:
            wb = load_workbook(WEEKLY_PIPELINE, read_only=True, data_only=True)
            if "Worked_Accounts_Registry" in wb.sheetnames:
                ws = wb["Worked_Accounts_Registry"]
                rows = list(ws.iter_rows(values_only=True))
                if rows:
                    hdr = rows[0]
                    col = {h: i for i, h in enumerate(hdr) if h}
                    name_col = (
                        col.get("Company_Normalized")
                        or col.get("Company")
                        or col.get("Company Name")
                        or 0
                    )
                    for r in rows[1:]:
                        if not r: continue
                        n = r[name_col] if isinstance(name_col, int) and name_col < len(r) else None
                        k = _norm_key(str(n) if n else "")
                        if k: keys.add(k)
            wb.close()
        except Exception as e:
            print(f"  [warn] could not read Worked_Accounts_Registry: {e}", file=sys.stderr)
    return keys

@functools.cache
def load_enrichment_maps():
    """company-key → {phone, ee, dm_name, dm_email, owner_title, website} from enrichments table.
    Used to backfill holes in displacement CSV + trigger output before seeding."""
    out = {}  # key → dict
    if not DB.exists(): return out
    fields_we_want = {
        "phone", "ee_count", "primary_contact_name", "primary_contact_title",
        "owner_or_ceo", "owner_title", "contact_email", "website",
    }
    try:
        db = sqlite3.connect(str(DB))
        placeholders = ",".join(["?"] * len(fields_we_want))
        q = (f"SELECT c.name_display, e.field, e.value "
             f"FROM enrichments e JOIN companies c ON c.company_id = e.company_id "
             f"WHERE e.field IN ({placeholders})")
        for n_disp, fld, val in db.execute(q, tuple(fields_we_want)).fetchall():
            k = _norm_key(n_disp)
            if not k or not val: continue
            rec = out.setdefault(k, {})
            if fld not in rec:
                rec[fld] = str(val).strip()
        db.close()
    except Exception as e:
        print(f"  [warn] could not load enrichment maps: {e}", file=sys.stderr)
    return out


def enrich_record(rec, enrich_maps):
    """Fill holes in a single account record from enrichments. In-place."""
    e = enrich_maps.get(rec["key"])
    if not e: return
    if not rec.get("phone") and e.get("phone"):
        rec["phone"] = re.sub(r"\D", "", e["phone"])[:11]
    if not rec.get("ee") and e.get("ee_count"):
        try: rec["ee"] = int(float(e["ee_count"]))
        except (ValueError, TypeError): pass
    if not rec.get("dm_name"):
        rec["dm_name"] = e.get("primary_contact_name") or e.get("owner_or_ceo") or ""
    if not rec.get("dm_email") and e.get("contact_email"):
        rec["dm_email"] = e["contact_email"]
    if not rec.get("website") and e.get("website"):
        rec["website"] = e["website"]


@functools.cache
def load_peo_signal_maps():
    """Per Nick 2026-05-11 ICP reframe: health > WC, multi-state + growth + compliance lead.
    Returns {company_key → {has_5500, has_health_carriers, growth_signal, federal_contractor, multi_state, carrier_count, on_competing_peo}}.
    Used by score_overlay() to rescore the bench."""
    out = {}
    if not DB.exists(): return out
    db = sqlite3.connect(str(DB))
    # Step 1: company_id → name_key
    cid_to_key = {}
    for cid, name in db.execute("SELECT company_id, name_display FROM companies").fetchall():
        if name:
            cid_to_key[cid] = _norm_key(name)
    # Step 2: Build signal maps in single passes
    signals_we_want = {
        "has_5500", "has_health", "growth_signal", "federal_contractor",
        "defense_contractors_uei", "defense_contractors_cage",
        "carrier_count", "on_competing_peo", "5500_plan_year",
        "biz_event_acquisition", "sba_certs_uei", "sosnc_status",
    }
    placeholders = ",".join(["?"] * len(signals_we_want))
    for cid, fld, val in db.execute(
        f"SELECT company_id, field, value FROM enrichments WHERE field IN ({placeholders})",
        tuple(signals_we_want),
    ).fetchall():
        k = cid_to_key.get(cid)
        if not k or not val: continue
        rec = out.setdefault(k, {})
        rec[fld] = str(val).strip()
    # Step 3: companies that show up in carriers table → has_health_carriers + carrier_count
    for cid, n in db.execute("SELECT company_id, COUNT(*) FROM carriers GROUP BY company_id").fetchall():
        k = cid_to_key.get(cid)
        if k:
            out.setdefault(k, {})["has_health_carriers_via_table"] = "y"
            out[k]["health_carrier_count"] = n
    db.close()
    return out


def score_overlay(rec, peo_signals):
    """Per Nick 2026-05-11: rebalance loaded score toward the actual PEO buy signals.
    Adds boost for: has_5500, multi-state, growth, federal contractor, carrier consolidation.
    Subtracts penalty for: pure WC renewal (demote vs health-driven triggers).
    Returns the boost amount (added to rec['score'])."""
    sig = peo_signals.get(rec["key"], {})
    boost = 0.0

    # ── PRIMARY PEO BUY SIGNALS (Nick 2026-05-11 reframe) ──
    # Has 5500 + Schedule A carriers = proves benefits budget + category understanding
    if sig.get("has_5500") == "y" or sig.get("has_health") == "y" or sig.get("has_health_carriers_via_table") == "y":
        boost += 6.0
        rec["has_health_benefits"] = True

    # Multi-state operations — per-state compliance is THE PEO value
    multi_state_signals = (
        sig.get("federal_contractor") == "y" or
        sig.get("defense_contractors_uei") or
        sig.get("sba_certs_uei") or
        sig.get("defense_contractors_cage")
    )
    if multi_state_signals:
        boost += 5.5
        rec["multi_state_likely"] = True

    # Growth signal → talent retention pressure
    growth_val = (sig.get("growth_signal") or "").upper()
    if growth_val in ("RAPID GROWTH", "STRONG GROWTH"):
        boost += 4.5
        rec["growth_tier"] = "strong"
    elif growth_val in ("MODERATE GROWTH",):
        boost += 3.0
        rec["growth_tier"] = "moderate"

    # Federal contractor → SCA/FAR compliance burden (overlap with multi-state but distinct opener)
    if sig.get("federal_contractor") == "y" and not rec.get("federal_contractor_boosted"):
        boost += 1.0  # small bump on top of multi-state
        rec["federal_contractor"] = True

    # Carrier fragmentation (3+ health carriers on 5500) — admin pain
    try:
        cc = int(sig.get("carrier_count") or sig.get("health_carrier_count") or 0)
    except (ValueError, TypeError):
        cc = 0
    if cc >= 3:
        boost += 4.0
        rec["multi_carrier_consolidation"] = cc

    # Recent acquisition event
    if sig.get("biz_event_acquisition") == "y":
        boost += 3.5
        rec["acquired_recently"] = True

    # ── SECONDARY: WC renewal — DEMOTE per Nick's reframe ──
    if rec.get("primary_trigger") == "wc_renewal":
        boost -= 2.5  # net effect: WC renewal accounts drop unless they ALSO have health/multi-state/growth

    return boost


@functools.cache
def load_wc_carrier_map():
    """company-key → wc_carrier (from enrichments). Backfills the trigger engine's null carrier column."""
    m = {}
    if not DB.exists(): return m
    try:
        db = sqlite3.connect(str(DB))
        for n_disp, val in db.execute(
            "SELECT c.name_display, e.value "
            "FROM enrichments e JOIN companies c ON c.company_id = e.company_id "
            "WHERE e.field = 'wc_carrier'"
        ).fetchall():
            k = _norm_key(n_disp)
            if k and k not in m and val:
                m[k] = str(val).strip()
        db.close()
    except Exception as e:
        print(f"  [warn] could not load WC carrier map: {e}", file=sys.stderr)
    return m


@functools.cache
def load_naics_map():
    """company-key → naics_inferred (latest). Cached — DB scan is the slowest call (~10s)."""
    m = {}
    if not DB.exists(): return m
    try:
        db = sqlite3.connect(str(DB))
        for n_disp, val in db.execute(
            "SELECT c.name_display, e.value "
            "FROM enrichments e JOIN companies c ON c.company_id = e.company_id "
            "WHERE e.field = 'naics_inferred'"
        ).fetchall():
            k = _norm_key(n_disp)
            if k and k not in m and val:
                m[k] = str(val).strip()
        db.close()
    except Exception as e:
        print(f"  [warn] could not load NAICS map: {e}", file=sys.stderr)
    return m

# ─── CADENCE TEMPLATES — 4-touch / 21-day per Nick 2026-05-11 ───────────────────
# Touch-2 (drop) day is COUNTY-DEPENDENT and computed at seed time. This base list
# uses the "average drop day" of D3 — the actual day is overridden per-cadence based
# on route_day (0=Wake flex/Tue, 1=Tue Pitt, 2=Wed Northern, 3=Thu Cumberland).
CADENCE = [
    # day_offset, channel, action_label, time_block
    (0,  "email",    "Email touch-1: 1-pager trigger brief, NO ASK",               "10:00-11:30"),
    (3,  "drop",     "Field drop touch-2: printed analysis to gatekeeper",         "field 10:00-14:30"),
    (8,  "linkedin", "LinkedIn touch-3: connect + brief DM referencing prior",     "10:00-12:00"),
    (15, "call",     "Phone touch-4: ASK for 15-min meeting",                      "10:00-12:00"),
    (22, "breakup",  "Break-up email touch-5 (optional): 'closing the loop'",      "09:00-10:00 Fri"),
]
CADENCE_DAYS = {c[0] for c in CADENCE}

# Per-route-day drop offset (cadence Touch-2). Routes drops to field days only.
# rd 0 (Wake) is flexible — defaults to Tue, can be Wed/Thu if Tue is full.
ROUTE_DAY_DROP_OFFSET = {
    0: 1,   # Wake → Tue of seed week
    1: 1,   # Pitt → Tue
    2: 2,   # Northern → Wed
    3: 3,   # Cumberland/Coastal → Thu
}

def build_touch_schedule(start_date, route_day):
    """Returns list of touch dicts with scheduled_for dates.
    Touch-2 (drop) uses route-day-aware offset; others use static CADENCE day_offset."""
    touches = []
    for d, ch, _label, _tb in CADENCE:
        # Override drop day per route_day cluster
        if ch == "drop":
            offset = ROUTE_DAY_DROP_OFFSET.get(route_day, 3)
        else:
            offset = d
        touches.append({
            "day": d,
            "channel": ch,
            "scheduled_for": (start_date + dt.timedelta(days=offset)).isoformat(),
            "completed": False,
            "outcome": "",
            "notes": "",
            "broker_captured": "",
        })
    return touches

OUTREACH_OPENERS = {
    "wc_renewal": {
        "email_subject": "Quick read — eastern NC {VERTICAL} WC renewal benchmark",
        "email_body": (
            "Hey {FIRST_NAME},\n\n"
            "{WC_LEAD_LINE}\n\n"
            "Rather than pitch you anything, I put together a 1-page benchmark of how {VERTICAL} "
            "mod rates in eastern NC are tracking this cycle. Most owners are seeing 8-15% "
            "increases on top of any mod movement. Attached.\n\n"
            "No reply needed — just figured you'd find it useful before the renewal letter shows up.\n\n"
            "— Nick\nADP TotalSource | [PHONE] | [CALENDLY]"
        ),
        "linkedin": (
            "Hey {FIRST_NAME} — sent you an email last week with the {VERTICAL} mod-rate "
            "benchmark. {WC_LINKEDIN_REF} Happy to send the comparable analysis for your size "
            "band — {N} {VERTICAL} employers in {COUNTY} that we have visibility on. No ask. "
            "Want me to send it over?"
        ),
        "drop_script": (
            "30-second drop. Walk in, smile, address gatekeeper:\n"
            "\"Hi — I'm Nick with ADP. I'm not selling anything today. I have a one-pager I "
            "put together for {OWNER}'s WC renewal cycle. Could you make sure she/he gets it? "
            "My card's on top — no pressure to call, just figured it'd be useful before the "
            "renewal letter comes.\"\n"
            "Capture: gatekeeper name, owner schedule, current HR/payroll system."
        ),
        "call_script": (
            "Hey {FIRST_NAME}, Nick with ADP. Sent you the {VERTICAL} mod-rate benchmark by "
            "email two weeks ago, dropped a printed copy off last week. Your renewal with "
            "{CARRIER} is about {DAYS_TO_RENEWAL} days out — most owners I talk to in your "
            "size band are starting to compare options around now. Worth a 20-minute call "
            "before you sign?"
        ),
    },
    "health_renewal": {
        "email_subject": "Industry-wide small-group renewals at 15-30% — quick benchmark",
        "email_body": (
            "Hey {FIRST_NAME},\n\n"
            "Pulled your most recent Form 5500 — looks like your medical plan is on a "
            "{PLAN_YEAR} cycle.\n\n"
            "Heads-up: small-group renewals across NC are coming in at 15-30% this cycle "
            "(worst we've seen in a decade). The way TotalSource works, your employees roll "
            "into ADP's master plan — same UHC/Aetna/BCBSNC network options, but underwritten "
            "as part of a 750,000+ life pool. Typical first-year reduction vs. standalone "
            "renewal: 5-18%.\n\n"
            "Attached: 1-page benchmark of how {COUNTY} {VERTICAL} employers are running "
            "this cycle. No ask.\n\n— Nick"
        ),
        "linkedin": (
            "Hey {FIRST_NAME} — sent the small-group renewal benchmark last week. Wanted to "
            "add: I have visibility into the actual master-plan numbers for your {VERTICAL}/"
            "{EE}-EE band. If you want me to run a side-by-side against your current plan "
            "once you get your renewal letter, I can have the comparison back to you in 48 "
            "hours. No commitment. Worth keeping in your back pocket?"
        ),
        "drop_script": (
            "30-second drop. \"Hi — Nick with ADP. Not selling anything today. Have a "
            "small-group renewal benchmark for {OWNER}. Could you pass it along? Card on top.\""
        ),
        "call_script": (
            "{FIRST_NAME}, Nick again. The reason this renewal cycle matters more than usual: "
            "the carriers shopping the small-group market are quietly thinning networks AND "
            "raising rates simultaneously. The master-plan structure protects against both. "
            "15-minute call before you commit to anything?"
        ),
    },
    "osha_recent": {
        "email_subject": "Saw the {DATE} OSHA filing — not a pitch",
        "email_body": (
            "{FIRST_NAME},\n\n"
            "I saw the OSHA filing on the public database from {DATE}. Hope your team member "
            "is doing OK.\n\n"
            "I won't pretend that's not painful. The reason I'm reaching out: most owners "
            "after a citation realize they don't have in-house bandwidth to actually get and "
            "stay compliant. We run an outsourced compliance + safety program for {VERTICAL} "
            "employers in eastern NC — on-site OSHA-30 training, documented protocols, monthly "
            "safety walkthroughs. Point is to reduce repeat-citation exposure, not to sell "
            "software.\n\nAttached: 1-page summary.\n\n— Nick"
        ),
        "linkedin": (
            "{FIRST_NAME} — followed up on email with the program summary. I work with three "
            "{VERTICAL} employers in {COUNTY} who came to us after a citation; happy to put "
            "you in touch with one of them if a peer reference would help. Just say the word."
        ),
        "drop_script": (
            "30-second drop. Be quiet, be respectful. \"Hi — Nick with ADP. I have a "
            "safety-program one-pager for {OWNER}. No pitch — just leaving it.\""
        ),
        "call_script": (
            "{FIRST_NAME}, Nick — third touch. I won't keep pestering. The 15-min call I'm "
            "offering is how to insulate from a *second* citation, not what happened in the "
            "first. If that's the conversation you want, my Calendly is in the email. If not, "
            "no hard feelings."
        ),
    },
    "hiring_velocity": {
        "email_subject": "Noticed {N_JOBS} open roles — comp benchmark for {VERTICAL} {COUNTY}",
        "email_body": (
            "{FIRST_NAME},\n\n"
            "Saw {N_JOBS} open roles on Indeed — congrats on the growth. The hiring market "
            "for {VERTICAL} talent in eastern NC is brutal this year; I put together a 1-page "
            "benchmark of what the comp/benefits packages look like at the F500 anchors in "
            "your space.\n\n"
            "Most sub-50-employee companies in {VERTICAL} tell me they can't compete on "
            "benefits, so they end up overpaying on base. Inside TotalSource, you get a "
            "benefits package that matches the F500 anchors — used as a recruiting weapon, "
            "not a cost center. Net cost is usually flat.\n\n— Nick"
        ),
        "linkedin": (
            "Hey {FIRST_NAME} — saw you're hiring. Most clients in {VERTICAL} tell me "
            "candidates take a counter-offer when they see the benefits gap vs. the F500s "
            "next door. The fix is usually a benefits package that punches above the company's "
            "size, not raising base. 15-min call?"
        ),
        "drop_script": (
            "30-second drop. \"Hi — Nick with ADP. Have a comp-benefits benchmark for "
            "{OWNER}. {N_JOBS} open roles → figured this would be useful. Card on top.\""
        ),
        "call_script": (
            "{FIRST_NAME}, Nick. Quick one — your roles are still open per Indeed. If a "
            "benefits package that matches the F500 anchors could close that gap, that's a "
            "15-minute conversation. If not, ignore me."
        ),
    },
    "ale_threshold": {
        "email_subject": "Heads-up before you cross 50 FTE — ACA reporting timeline",
        "email_body": (
            "{FIRST_NAME},\n\n"
            "You came across my radar at {EE} employees. Quick heads-up:\n\n"
            "When companies cross 50 full-time-equivalents, three things kick in:\n"
            "1. Applicable Large Employer status (ACA).\n"
            "2. 1094-C / 1095-C reporting (~100 hrs/yr of admin).\n"
            "3. Employer-mandate penalty exposure if 1095-Cs are wrong "
            "($2,970-$4,460 per FTE per year).\n\n"
            "Most owners don't realize this until the first 1095-C deadline lands and their "
            "CPA quotes them $8K-$15K to fix it.\n\n"
            "Inside TotalSource, the ACA reporting is handled at the master level. Attached: "
            "1-page ACA-readiness checklist.\n\n— Nick"
        ),
        "linkedin": (
            "{FIRST_NAME} — sent the ACA checklist last week. The $2,970-per-FTE penalty math "
            "is what gets owners' attention. Worth a 15-min call before you cross the threshold?"
        ),
        "drop_script": (
            "30-second drop. \"Hi — Nick with ADP. Have an ACA-readiness checklist for "
            "{OWNER}. At {EE} EE you're approaching the threshold. Card on top.\""
        ),
        "call_script": (
            "{FIRST_NAME}, Nick. Last touch. The ACA timeline is real and it gets expensive "
            "fast if 1095-Cs aren't right. 15-minute call to walk through your specific exposure?"
        ),
    },
    "default": {
        "email_subject": "Quick benchmark for {COMPANY} — no ask attached",
        "email_body": (
            "Hey {FIRST_NAME},\n\nNoticed {COMPANY} on the {VERTICAL} list for {COUNTY}. "
            "Put together a quick benchmark of what eastern NC {VERTICAL} employers in your "
            "size band are seeing on HR/comp/benefits this year. Attached.\n\nNo reply needed.\n\n— Nick"
        ),
        "linkedin": (
            "Hey {FIRST_NAME} — sent a {VERTICAL} benchmark to your inbox last week. Worth "
            "comparing notes for 15 min?"
        ),
        "drop_script": (
            "30-second drop. \"Hi — Nick with ADP. Have a benchmark for {OWNER}. Card on top.\""
        ),
        "call_script": (
            "{FIRST_NAME}, Nick. Followed up via email and dropped by the office. Worth a "
            "15-minute call if any of it landed?"
        ),
    },
}


# ─── VERTICAL HOOKS (per playbook §1.6 — overlay one sentence into emails) ──────
VERTICAL_HOOKS = {
    "HVAC": "Eastern NC HVAC mods are tracking 12–18% over last cycle on EMR alone — Cintas-grade audit programs are the lever.",
    "Engineering": "PE/SE retention is the eastern NC engineering bottleneck this cycle; the master-plan benefits package is what's actually moving 3-yr retention numbers.",
    "Technology": "When sub-50-EE tech competes with a F500 anchor (Lenovo, IBM, Cisco), it's never base — it's the equity-versus-benefits gap candidates compare during offer week.",
    "Manufacturing": "WC mod movement on NCCI 3724/3041/3066 class codes is what's killing margins this cycle — and what TotalSource's safety program directly addresses.",
    "Biotech / Life Sci": "FDA-inspection readiness on the HR/employment side (I-9, training records, harassment compliance) is now the #1 deal-killer in funding diligence — TotalSource handles all of it.",
    "Pharma": "DEA / cGMP compliance + the new HHS/FDA hiring scrutiny on background checks for handlers — TotalSource is the only PEO with a pharma compliance specialty desk.",
    "Health Care": "DOL is auditing healthcare 5500s this cycle for fiduciary breach (especially small practices on broker-built plans) — master-plan structure removes the fiduciary load.",
    "Mgmt Consulting": "Multi-state worker exposure + 1099-vs-W-2 reclassification (NCDOL is enforcing aggressively) is the silent compliance bomb on consulting firms — handled at master level inside TotalSource.",
    "Wholesale Trade": "Wholesale trade WC mods + the freight/logistics bleed-over into your class codes are eating margin this renewal — comparison is worth 30 minutes.",
    "Finance/Insurance": "Talent retention in eastern NC finance is brutal — the master-plan benefits package punches at F500 weight without F500 cost.",
}
def vertical_hook(vertical_str: str) -> str:
    """Map free-text 'Vertical' string to a hook sentence; falls back to a generic line."""
    if not vertical_str: return ""
    v = vertical_str.strip()
    if v in VERTICAL_HOOKS: return VERTICAL_HOOKS[v]
    # Fuzzy bucket: substring match
    vlow = v.lower()
    for k, hook in VERTICAL_HOOKS.items():
        if k.lower().split()[0] in vlow: return hook
    return ""

# ─── INCUMBENT-PEO DISPLACEMENT OPENERS (per playbook §1.4) ─────────────────────
INCUMBENT_OPENERS = {
    "Insperity": {
        "email_subject": "Quick read — your Insperity renewal + the UHC contract pressure",
        "email_body": (
            "Hey {DM_NAME},\n\n"
            "Saw {COMPANY} is on the Insperity master plan. Two things worth knowing before your "
            "next anniversary:\n\n"
            "1. Insperity's stock has taken a public hit specifically on UHC contract pressure — "
            "the small-group network options on their master are getting squeezed.\n"
            "2. Most Insperity renewals are landing 12–22% this cycle, with an expanding admin "
            "carve-out on top.\n\n"
            "{VERTICAL_HOOK}\n\n"
            "I put together a 1-page side-by-side: TotalSource vs. Insperity, same {EE}-EE band, "
            "same county, same vertical. No commitment — just useful reading before your "
            "renewal letter shows up.\n\n"
            "— Nick\nADP TotalSource | [PHONE] | [CALENDLY]"
        ),
        "linkedin": (
            "Hey {DM_NAME} — quick one. Saw {COMPANY} runs through Insperity. Most clients I "
            "talk to in your size band are surprised when they see the Insperity vs. TotalSource "
            "side-by-side, especially on the medical side after the UHC contract pressure hit. "
            "Want me to send the comparable for {EE}-EE {VERTICAL}? No commitment."
        ),
        "drop_script": (
            "30-second drop. \"Hi — Nick with ADP TotalSource. Not pitching. Have a 1-page "
            "Insperity-vs-TotalSource comparison for {DM_NAME} ahead of their renewal. Card on "
            "top — please make sure she/he gets it.\""
        ),
        "call_script": (
            "{DM_NAME}, Nick with ADP TotalSource. Sent you the Insperity comparison and dropped "
            "by last week. Most owners on Insperity right now are seeing the UHC pressure show "
            "up on renewal — worth a 20-minute side-by-side before your anniversary letter lands?"
        ),
    },
    "TriNet": {
        "email_subject": "Quick read — TriNet master plan vs. ADP TotalSource for {COMPANY}",
        "email_body": (
            "Hey {DM_NAME},\n\n"
            "Confirmed {COMPANY} is on the TriNet MEP master plan via the 2024 5500 filing. "
            "Two notes:\n\n"
            "1. TriNet's PEO pricing structure breaks differently than Insperity/ADP's — the "
            "admin fee is bundled, but the medical underwriting is per-employer. That means the "
            "small-group renewal volatility hits TriNet clients more directly than the master-pool "
            "structure on TotalSource.\n"
            "2. TriNet contracts have a 12-month escape with proper notice — most clients don't "
            "realize they have an exit window.\n\n"
            "{VERTICAL_HOOK}\n\n"
            "I have a 1-page TriNet-vs-TotalSource for your size band. No commitment.\n\n"
            "— Nick"
        ),
        "linkedin": (
            "Hey {DM_NAME} — saw {COMPANY} is on the TriNet MEP. Quick FYI: most TriNet "
            "renewals this cycle have a 12-month escape window most clients don't know about. "
            "Want me to send the TotalSource comp for {EE}-EE {VERTICAL}?"
        ),
        "drop_script": (
            "30-second drop. \"Hi — Nick with ADP TotalSource. Have the TriNet-vs-TotalSource "
            "comparison for {DM_NAME}. Not pitching today. Card on top.\""
        ),
        "call_script": (
            "{DM_NAME}, Nick with ADP TotalSource. Followed up on the TriNet MEP comparison "
            "I sent + dropped. Worth 20 minutes to walk through the side-by-side before your "
            "next anniversary?"
        ),
    },
    "Paychex": {
        "email_subject": "Quick note on Paychex PEO Direct — pricing structure comparison",
        "email_body": (
            "Hey {DM_NAME},\n\n"
            "Saw {COMPANY} is on Paychex PEO Direct. One thing worth knowing: Paychex PEO Direct "
            "isn't a true master-plan structure the way ADP TotalSource and Insperity are — the "
            "medical pool is materially smaller, and small-group volatility passes through more "
            "directly on renewal.\n\n"
            "{VERTICAL_HOOK}\n\n"
            "I have a 1-page side-by-side for your size band — no commitment.\n\n"
            "— Nick"
        ),
        "linkedin": (
            "Hey {DM_NAME} — Paychex PEO Direct flag: it's not a master-plan structure, the pool "
            "is much smaller than TotalSource. The renewal economics break differently. Want the "
            "1-page comparison for {EE}-EE {VERTICAL}?"
        ),
        "drop_script": (
            "30-second drop. \"Hi — Nick with ADP TotalSource. Have a Paychex-vs-TotalSource "
            "comparison for {DM_NAME}. Card on top — not pitching today.\""
        ),
        "call_script": (
            "{DM_NAME}, Nick with ADP. Followed up on the Paychex PEO Direct comparison I sent. "
            "Worth 20 minutes on the master-plan-vs-direct difference?"
        ),
    },
    "Questco": {
        "email_subject": "Quick read — Questco vs. TotalSource on master-plan size + NC service depth",
        "email_body": (
            "Hey {DM_NAME},\n\n"
            "Saw {COMPANY} runs through Questco's MEP master plan via your 2024 5500 filing. "
            "Two structural things worth knowing:\n\n"
            "1. Questco's master pool is materially smaller than ADP TotalSource or Insperity — "
            "the risk-pool math just doesn't compound the same way. Small-group medical "
            "volatility hits Questco clients harder on renewal.\n"
            "2. Questco's NC field-service footprint is thin. Most of the actual HR consulting "
            "happens by phone from Texas — not on-site in eastern NC when NCDOL comes asking.\n\n"
            "{VERTICAL_HOOK}\n\n"
            "I have a 1-page Questco-vs-TotalSource side-by-side for your size band. No "
            "commitment, just useful prep before your renewal anniversary.\n\n"
            "— Nick\nADP TotalSource | [PHONE] | [CALENDLY]"
        ),
        "linkedin": (
            "Hey {DM_NAME} — saw {COMPANY} is on Questco's MEP. Two structural gaps worth a look: "
            "(1) Questco's medical pool is much smaller than TotalSource so renewal volatility "
            "hits harder, (2) thin NC field service — most HR consulting is phone-only from TX. "
            "Want the 1-page comparison for {EE}-EE {VERTICAL}?"
        ),
        "drop_script": (
            "30-second drop. \"Hi — Nick with ADP TotalSource. Have a Questco-vs-TotalSource "
            "comparison for {DM_NAME}. Card on top — not pitching today, just figured it'd be "
            "useful before the renewal letter.\""
        ),
        "call_script": (
            "{DM_NAME}, Nick with ADP TotalSource. Followed up on the Questco comparison I sent "
            "and dropped by. The master-plan size difference is the conversation most owners "
            "wish they'd had before signing. Worth 20 minutes?"
        ),
    },
    "Justworks": {
        "email_subject": "Quick read — Justworks vs. TotalSource on the eastern NC field-service side",
        "email_body": (
            "Hey {DM_NAME},\n\n"
            "Confirmed {COMPANY} is on Justworks. The structural gap: Justworks runs lean on "
            "eastern NC field service — there's no on-the-ground HR consultant for compliance "
            "issues that need a face. Most clients hit that wall the first time NCDOL comes "
            "asking questions.\n\n"
            "{VERTICAL_HOOK}\n\n"
            "TotalSource has 12 dedicated HR business partners assigned to NC. I have a 1-page "
            "service-model comparison + benefits side-by-side. No commitment.\n\n"
            "— Nick"
        ),
        "linkedin": (
            "Hey {DM_NAME} — Justworks flag for {COMPANY}: thin field service in eastern NC means "
            "compliance issues that need a face land hard. TotalSource has 12 NC HR business "
            "partners. Want the 1-page comparison?"
        ),
        "drop_script": (
            "30-second drop. \"Hi — Nick with ADP TotalSource. Have a Justworks-vs-TotalSource "
            "service-model comparison for {DM_NAME}. Card on top.\""
        ),
        "call_script": (
            "{DM_NAME}, Nick with ADP. Followed up on the Justworks comparison + dropped by. "
            "Worth 20 minutes on the field-service difference?"
        ),
    },
}
def incumbent_canonical(raw: str) -> str:
    """Map free-text current_peo string to one of the displacement playbooks (or 'Other')."""
    if not raw: return ""
    r = raw.strip().lower()
    if "insperity" in r: return "Insperity"
    if "trinet" in r: return "TriNet"
    if "questco" in r: return "Questco"
    if "justworks" in r: return "Justworks"
    # Paychex check comes after Questco because some rows say "Paychex (acquired Oasis) | Oasis Outsourcing".
    # Oasis-legacy clients on Paychex Direct: Paychex template fires, which is correct (Paychex bought Oasis).
    if "paychex" in r or "oasis" in r: return "Paychex"
    return "Other"


# ─── INDUSTRY-SIGNAL OPENER TEMPLATES (new triggers from 2026-05 audit) ─────────
INDUSTRY_TRIGGER_OPENERS = {
    "wc_lapsed": {
        "email_subject": "Heads-up on your NC workers' comp coverage",
        "email_body": (
            "Hey {FIRST_NAME},\n\n"
            "Saw on the NC Industrial Commission database that {COMPANY}'s workers' comp "
            "coverage lapsed recently. NC requires WC at 3+ employees, so this is the kind "
            "of compliance gap that creates real exposure if a claim hits between now and the "
            "next bind.\n\n"
            "I'm not pitching anything — just wanted to flag it before it became a problem. "
            "If you want, I can send a quick 1-pager on how {VERTICAL} owners in eastern NC "
            "are handling reinstatement and pricing this cycle.\n\n"
            "— Nick\nADP TotalSource | [PHONE] | [CALENDLY]"
        ),
        "linkedin": (
            "Hey {FIRST_NAME} — quick flag: looks like {COMPANY}'s WC coverage lapsed on the "
            "NCIC database. Not pitching, just wanted to make sure you knew before a claim "
            "lands. Happy to share what other {VERTICAL} owners are doing on reinstatement."
        ),
        "drop_script": (
            "30-second drop. \"Hi — Nick with ADP. Saw {COMPANY}'s WC coverage lapsed in the "
            "NCIC database. Just dropping a one-pager for {OWNER} — not pitching, just flagging. "
            "Card on top.\""
        ),
        "call_script": (
            "{FIRST_NAME}, Nick with ADP. Following up on the WC coverage gap — wanted to "
            "make sure {OWNER} saw the note. 15 minutes to walk through the reinstatement "
            "options before a claim creates a problem?"
        ),
    },
    "wc_recent_change": {
        "email_subject": "Saw you swapped WC carriers — {VERTICAL} benchmark inside",
        "email_body": (
            "Hey {FIRST_NAME},\n\n"
            "Caught that {COMPANY} changed WC carriers recently on the NCIC database. That "
            "usually means you were already shopping — which is exactly when the master-plan "
            "comparison is most useful.\n\n"
            "I put together a 1-page benchmark of {VERTICAL} mod rates in eastern NC for the "
            "same size band. No commitment — just useful intel for next year's renewal.\n\n"
            "— Nick"
        ),
        "linkedin": (
            "Hey {FIRST_NAME} — saw {COMPANY} switched WC carriers. Want the {VERTICAL} "
            "mod-rate benchmark for next year's renewal cycle? No commitment."
        ),
        "drop_script": (
            "30-second drop. \"Hi — Nick with ADP. Have a {VERTICAL} WC benchmark for {OWNER} "
            "since you just switched carriers. Card on top, not pitching today.\""
        ),
        "call_script": (
            "{FIRST_NAME}, Nick with ADP. Followed up on the WC carrier swap — worth 15 "
            "minutes on the master-plan comparison before next renewal?"
        ),
    },
    "osha_confirmed": OUTREACH_OPENERS["osha_recent"],  # alias to existing template
    "foreign_labor": {
        "email_subject": "Saw the LCA filings — {VERTICAL} talent benchmark inside",
        "email_body": (
            "Hey {FIRST_NAME},\n\n"
            "Pulled your DOL LCA filings — looks like {COMPANY} is using H-1B/visa-dependent "
            "hiring to fill {VERTICAL} roles. That's a tell: the local talent market for these "
            "positions is tight enough that you're paying the immigration premium.\n\n"
            "Inside TotalSource, the master-plan benefits package matches what F500 employers "
            "in your vertical offer — meaning your domestic-hire offers don't lose to the "
            "benefits gap. Net cost usually flat vs. current spend.\n\n"
            "Attached: 1-page benchmark of {VERTICAL} comp/benefits packages for the talent "
            "tier you're competing for.\n\n"
            "— Nick"
        ),
        "linkedin": (
            "Hey {FIRST_NAME} — saw the LCA filings for {COMPANY}. Visa hiring is the tell "
            "that your domestic offers can't compete on benefits. F500-grade benefits package "
            "inside a PEO usually closes the gap. Want the comp benchmark?"
        ),
        "drop_script": (
            "30-second drop. \"Hi — Nick with ADP. Have a {VERTICAL} talent benchmark for "
            "{OWNER} since you're hiring on H-1Bs. Card on top.\""
        ),
        "call_script": (
            "{FIRST_NAME}, Nick. The H-1B route says your domestic offers aren't winning. "
            "15 minutes on closing that gap with the master-plan benefits package?"
        ),
    },
    "carrier_consolidation": {
        "email_subject": "3+ health carriers on your 5500 — consolidation play",
        "email_body": (
            "Hey {FIRST_NAME},\n\n"
            "Saw {COMPANY}'s most recent 5500 — looks like you're running 3+ separate health "
            "carriers on the plan. That's a lot of broker spread + admin overhead for "
            "{EE} employees.\n\n"
            "Inside TotalSource, the master plan consolidates onto a single carrier network "
            "(UHC/Aetna/BCBS NC depending on geography) with the same plan options. One "
            "invoice, one broker, one renewal cycle. Net cost usually 8–15% lower.\n\n"
            "Attached: 1-page comparable for {EE}-EE {VERTICAL} employers in {COUNTY}.\n\n"
            "— Nick"
        ),
        "linkedin": (
            "Hey {FIRST_NAME} — 3+ carriers on the {COMPANY} 5500 = consolidation play. "
            "Want the side-by-side for {EE}-EE {VERTICAL}? One carrier, one invoice, lower spend."
        ),
        "drop_script": (
            "30-second drop. \"Hi — Nick with ADP. Have a benefits consolidation comparison "
            "for {OWNER}. Card on top, not pitching today.\""
        ),
        "call_script": (
            "{FIRST_NAME}, Nick. The 3-carrier setup is admin chaos for {EE} employees. "
            "15 minutes on consolidating to a single master plan?"
        ),
    },
    "shrinking_county": {
        "email_subject": "{COUNTY} county retention benchmark — {VERTICAL}",
        "email_body": (
            "Hey {FIRST_NAME},\n\n"
            "{COUNTY} county has been net-shedding employer jobs the last 18 months per BLS "
            "QCEW data. For {VERTICAL} owners that means two things:\n\n"
            "1. The candidates you DO have are getting recruited harder by out-of-county employers.\n"
            "2. Retention becomes cheaper than re-hiring — every 1% turnover reduction beats a "
            "$5K signing bonus on net cost.\n\n"
            "The master-plan benefits package is the retention lever that punches above your "
            "size. Attached: 1-page retention benchmark for {VERTICAL} in {COUNTY}.\n\n"
            "— Nick"
        ),
        "linkedin": (
            "Hey {FIRST_NAME} — {COUNTY} is shedding jobs per BLS. Retention is the cheapest "
            "insurance against the talent drain. Want the master-plan benefits benchmark?"
        ),
        "drop_script": (
            "30-second drop. \"Hi — Nick with ADP. Have a {COUNTY} retention benchmark for "
            "{OWNER}. Card on top.\""
        ),
        "call_script": (
            "{FIRST_NAME}, Nick. {COUNTY} job loss data — retention math gets ugly fast. "
            "15 minutes on the benefits-as-retention play?"
        ),
    },
    "dod_contractor": {
        "email_subject": "New DoD award + the SCA/FAR compliance load that comes with it",
        "email_body": (
            "Hey {FIRST_NAME},\n\n"
            "Saw {COMPANY}'s first DoD award land. Congrats — and a heads-up: federal "
            "contracting at your size adds Service Contract Act + FAR/DFARS compliance load "
            "your current HR setup probably isn't built for (52.222 clauses, prevailing wage "
            "audits, DCAA-ready timekeeping, EEO-1 reporting).\n\n"
            "TotalSource handles all of it inside the master service. Plus the SCA "
            "fringe benefits requirement (~$5/hr in benefits) is what a master plan covers "
            "naturally — most contractors over-pay because they don't structure it.\n\n"
            "Attached: 1-page DoD-contractor compliance checklist.\n\n"
            "— Nick"
        ),
        "linkedin": (
            "Hey {FIRST_NAME} — first DoD award? Congrats. The SCA + FAR compliance load is "
            "real at your size. Want the contractor compliance checklist?"
        ),
        "drop_script": (
            "30-second drop. \"Hi — Nick with ADP. Have a DoD-contractor compliance checklist "
            "for {OWNER}. Card on top.\""
        ),
        "call_script": (
            "{FIRST_NAME}, Nick. SCA fringe benefits compliance on the new DoD contract — "
            "15 minutes on getting it right before DCAA shows up?"
        ),
    },
    "warn_layoff": {
        "email_subject": "Saw the WARN filing — quick thought on the transition",
        "email_body": (
            "Hey {FIRST_NAME},\n\n"
            "Saw the WARN Act filing for {COMPANY}. Hope the team transition is going as "
            "smoothly as a transition can. I won't pretend that's not painful.\n\n"
            "The reason I'm reaching out: most companies post-WARN realize their remaining "
            "team is going to ask 'is my benefit package next?' Stability of benefits is the "
            "#1 retention lever after a layoff event — and the master-plan structure is "
            "specifically designed to insulate from this.\n\n"
            "Not pitching today. If/when a benefits review makes sense, I'm here.\n\n"
            "— Nick"
        ),
        "linkedin": (
            "Hey {FIRST_NAME} — saw the WARN filing. Wishing the team smooth transitions. "
            "If a benefits stability conversation makes sense down the line, I'm here."
        ),
        "drop_script": (
            "30-second drop. Quiet and respectful. \"Hi — Nick with ADP. Just dropping a "
            "benefits-stability one-pager for {OWNER}. Card on top, no pressure.\""
        ),
        "call_script": (
            "{FIRST_NAME}, Nick. Following up on the WARN filing note — when a benefits "
            "review fits the team's healing timeline, I'm here. 15 minutes whenever."
        ),
    },
}
# Wire into OUTREACH_OPENERS so build_opener_for finds them
OUTREACH_OPENERS.update(INDUSTRY_TRIGGER_OPENERS)


def primary_trigger_type(triggers):
    """Map a trigger label list to a cadence-type key."""
    if not triggers: return "default"
    label = triggers[0][0] if isinstance(triggers[0], (list, tuple)) else str(triggers[0])
    label_lower = label.lower()
    if "wc renewal" in label_lower: return "wc_renewal"
    if "health" in label_lower or "5500" in label_lower: return "health_renewal"
    if "osha" in label_lower: return "osha_recent"
    if "hiring" in label_lower: return "hiring_velocity"
    if "50-fte" in label_lower or "ace" in label_lower or "ala" in label_lower or "ale" in label_lower: return "ale_threshold"
    return "default"


# ─── OUTCOME ENUM ───────────────────────────────────────────────────
OUTCOMES = [
    "",                        # not yet logged
    "no_answer",
    "voicemail",
    "gatekeeper",
    "owner_convo",
    "meeting_booked",
    "meeting_held",
    "meeting_cancelled",
    "meeting_no_show",
    "disqualified",
    "not_interested",
    "dnc",
    "wrong_number",
    "dead",
    "acquired",
    "nurture_90d",
]
# Outcomes that count as a stage progression (rep moved the ball)
PROGRESSION_OUTCOMES = {"gatekeeper", "owner_convo", "meeting_booked", "meeting_held"}
# Terminal outcomes that should kill the cadence + write to pipeline.sqlite drops
KILL_OUTCOMES = {
    "not_interested": "NOT_INTERESTED",
    "dnc": "DNC",
    "wrong_number": "WRONG_NUMBER",
    "dead": "DEAD",
    "acquired": "ACQUIRED",
    "disqualified": "DOA_UNVERIFIABLE",
}
# Outcomes that move the cadence to nurture (re-fire at +90d)
NURTURE_OUTCOMES = {"nurture_90d", "meeting_cancelled", "meeting_no_show"}
# Outcome that triggers MEDDPICC tracking
MEDDPICC_OUTCOMES = {"meeting_booked", "meeting_held"}

DEFAULT_WEIGHTS = {
    "trigger": {"wc_renewal": 1.0, "health_renewal": 1.0, "osha_recent": 1.0,
                "hiring_velocity": 1.0, "ale_threshold": 1.0,
                "displacement": 1.0, "default": 1.0},
    "vertical": {},          # learned over time
    "channel": {"email": 1.0, "linkedin": 1.0, "drop": 1.0, "call": 1.0},
    "route_day": {"0": 1.0, "1": 1.0, "2": 1.0, "3": 1.0},
    "last_recomputed": None,
    "n_outcomes_at_recompute": 0,
}

# ─── STATE PERSISTENCE ──────────────────────────────────────────────
def load_state():
    state = None
    if STATE_FILE.exists():
        try:
            state = json.loads(STATE_FILE.read_text())
        except Exception:
            state = None
    if state is None:
        state = {
            "initialized_at": None,
            "last_run": None,
            "active_cadences": [],
            "completed": [],
            "weekly_metrics": {},
        }
    # Migrate-on-load: ensure new schema keys exist
    state.setdefault("active_cadences", [])
    state.setdefault("completed", [])
    state.setdefault("weekly_metrics", {})
    state.setdefault("weekly_outcomes", [])      # ledger of every recorded touch outcome
    state.setdefault("nurture_queue", [])        # cadences re-fire at +90d
    state.setdefault("buyer_cast", {})           # company_key → {owner, cfo, office_mom, broker, cpa, attorney}
    state.setdefault("meddpicc", {})             # company_key → 8-field score dict
    state.setdefault("channel_build", {"cpas": [], "brokers": [], "attorneys": []})
    if "weights" not in state:
        state["weights"] = json.loads(json.dumps(DEFAULT_WEIGHTS))  # deep copy
    else:
        # Top-up missing weight categories (forward compatibility)
        for k, v in DEFAULT_WEIGHTS.items():
            state["weights"].setdefault(k, v)
    # Each active cadence: ensure touches[] exists with full schema
    for c in state["active_cadences"]:
        if "touches" not in c or not c["touches"]:
            start = dt.date.fromisoformat(c["start_date"])
            c["touches"] = build_touch_schedule(start, c.get("route_day", 0))
        else:
            # Top-up missing fields
            for t in c["touches"]:
                t.setdefault("scheduled_for", "")
                t.setdefault("completed", False)
                t.setdefault("outcome", "")
                t.setdefault("notes", "")
                t.setdefault("broker_captured", "")
    return state

def save_state(state):
    state["last_run"] = TODAY_ISO
    STATE_FILE.write_text(json.dumps(state, indent=2, default=str))


# ─── DB / TRIGGER UNIVERSE LOAD ─────────────────────────────────────
def load_trigger_universe():
    """Read the bench from Trigger_Engine_Output.xlsx Daily_Top_25 + scoring tabs.
    Returns list of dicts sorted by score descending."""
    if not TRIGGER_OUT.exists():
        print(f"  [error] {TRIGGER_OUT} not found. Run scripts/build_trigger_engine.py first.", file=sys.stderr)
        sys.exit(1)
    wb = load_workbook(TRIGGER_OUT, read_only=True, data_only=True)

    excluded_keys = load_excluded_keys()
    naics_map = load_naics_map()
    wc_carrier_map = load_wc_carrier_map()
    print(f"  Excluding {len(excluded_keys)} previously-worked / drop-listed companies.")
    print(f"  WC carrier backfill: {len(wc_carrier_map)} carriers indexed from enrichments.")

    # Pull from WC_Renewals_120d (largest tab) — has the most data
    accounts = []
    seen_ids = set()
    n_excluded_dedupe = 0
    n_excluded_naics = 0
    n_excluded_name = 0
    n_excluded_ee = 0
    n_excluded_zip = 0
    for tab_name in ("WC_Renewals_120d", "Health_Renewals", "OSHA_Hot", "Hiring_Velocity"):
        if tab_name not in wb.sheetnames: continue
        ws = wb[tab_name]
        # Header row depends on tab — read first row that has "Company" in col A
        rows = list(ws.iter_rows(values_only=True))
        hdr_row_idx = next((i for i, r in enumerate(rows[:10]) if r and r[0] == "Company"), None)
        if hdr_row_idx is None: continue
        hdr = rows[hdr_row_idx]
        col = {h: i for i, h in enumerate(hdr) if h}
        for r in rows[hdr_row_idx+1:]:
            if not r or not r[col.get("Company", 0)]: continue
            company = r[col["Company"]]
            key = re.sub(r"[^a-z0-9]","", str(company).lower())
            if key in seen_ids: continue
            seen_ids.add(key)
            if key in excluded_keys:
                n_excluded_dedupe += 1
                continue
            naics_val = naics_map.get(key, "")
            if is_naics_excluded(naics_val):
                n_excluded_naics += 1
                continue
            company_upper = str(company).upper()
            if (EXCLUDE_NAME_REGEX.search(str(company))
                or any(s in company_upper for s in EXCLUDE_NAME_SUBSTR)
                or is_off_focus_by_name(str(company))):
                n_excluded_name += 1
                continue
            ee_val = r[col["EE"]] if "EE" in col and r[col["EE"]] not in (None, "") else None
            try:
                ee_int = int(ee_val) if ee_val is not None else None
            except (ValueError, TypeError):
                ee_int = None
            if ee_int is not None and (ee_int < ICP_EE_MIN or ee_int > ICP_EE_MAX):
                n_excluded_ee += 1
                continue
            vert_raw = (str(r[col["Vertical"]] or "") if "Vertical" in col else "").strip()
            if vert_raw in EXCLUDE_VERTICAL_EXACT or any(vert_raw.startswith(p) for p in EXCLUDE_VERTICAL_PREFIXES):
                n_excluded_naics += 1
                continue
            zip_raw = (str(r[col["ZIP"]] or "") if "ZIP" in col else "").strip().split(".")[0].zfill(5) if "ZIP" in col and r[col["ZIP"]] else ""
            if zip_raw in EXCLUDE_ZIPS:
                n_excluded_zip += 1
                continue
            rec = {
                "key": key,
                "company": str(company),
                "city": str(r[col["City"]] or "") if "City" in col else "",
                "zip": str(r[col["ZIP"]] or "") if "ZIP" in col else "",
                "phone": str(r[col["Phone"]] or "") if "Phone" in col else "",
                "ee": r[col["EE"]] if "EE" in col and r[col["EE"]] not in (None, "") else None,
                "vertical": str(r[col["Vertical"]] or "") if "Vertical" in col else "",
                "score": float(r[col["Score"]] or 0) if "Score" in col else 0.0,
                "trigger_source_tab": tab_name,
            }
            # Tab-specific extras
            if tab_name == "WC_Renewals_120d":
                src_carrier = str(r[col["WC Carrier"]] or "") if "WC Carrier" in col else ""
                rec["wc_carrier"]    = src_carrier or wc_carrier_map.get(key, "")
                rec["wc_renewal"]    = str(r[col["Renewal"]] or "") if "Renewal" in col else ""
                rec["days_out"]      = int(r[col["Days Out"]]) if "Days Out" in col and r[col["Days Out"]] not in (None, "") else 999
                rec["primary_trigger"] = "wc_renewal"
            elif tab_name == "Health_Renewals":
                rec["evidence"] = str(r[col["Carriers (Sched A)"]] or "") if "Carriers (Sched A)" in col else ""
                rec["plan_year"] = str(r[col["Plan Year"]] or "") if "Plan Year" in col else ""
                rec["primary_trigger"] = "health_renewal"
            elif tab_name == "OSHA_Hot":
                rec["evidence"] = str(r[col["Evidence"]] or "") if "Evidence" in col else ""
                rec["primary_trigger"] = "osha_recent"
            elif tab_name == "Hiring_Velocity":
                rec["evidence"] = str(r[col["Evidence"]] or "") if "Evidence" in col else ""
                rec["primary_trigger"] = "hiring_velocity"
            rec["naics"] = naics_val
            accounts.append(rec)
    wb.close()
    # Enrichment pass: backfill phone/EE/DM from pipeline.sqlite enrichments
    enrich_maps = load_enrichment_maps()
    n_enriched = 0
    for a in accounts:
        before = (a.get("phone"), a.get("ee"), a.get("dm_name"), a.get("dm_email"))
        enrich_record(a, enrich_maps)
        if (a.get("phone"), a.get("ee"), a.get("dm_name"), a.get("dm_email")) != before:
            n_enriched += 1
    if n_enriched:
        print(f"  Enriched {n_enriched} accounts (phone/EE/DM backfill from enrichments).")
    print(f"  Filtered: -{n_excluded_dedupe} dedupe  -{n_excluded_naics} NAICS  -{n_excluded_name} name  -{n_excluded_ee} EE band  -{n_excluded_zip} ZIP")
    return accounts


@functools.cache
def _load_displacement_universe_cached():
    """Internal cached read — see load_displacement_universe() wrapper below."""
    if not COMPETING_PEO_CSV.exists():
        return []
    import csv
    accounts = []
    seen = set()
    with COMPETING_PEO_CSV.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            company = (row.get("company_name") or "").strip()
            if not company: continue
            if (row.get("viable_for_totalsource","").strip().lower() in {"false","no","0",""}
                and row.get("disqualification_reason","").strip()):
                continue
            if (row.get("is_adp_existing_customer") or "").strip().lower() in {"true","yes","1"}:
                continue
            key = _norm_key(company)
            if not key or key in seen: continue
            # Apply same ICP guards
            if EXCLUDE_NAME_REGEX.search(company): continue
            company_upper = company.upper()
            if any(s in company_upper for s in EXCLUDE_NAME_SUBSTR): continue
            if is_off_focus_by_name(company): continue
            try:
                ee_val = row.get("ee_count","")
                ee_int = int(float(ee_val)) if ee_val not in (None,"") else None
            except (ValueError, TypeError):
                ee_int = None
            if ee_int is not None and (ee_int < ICP_EE_MIN or ee_int > ICP_EE_MAX):
                continue
            # NC zip prefix filter — kills foreign-state corruptions in the CSV
            zip_clean = (row.get("zip") or "").strip().split(".")[0].zfill(5) if row.get("zip") else ""
            if zip_clean:
                if not (zip_clean.startswith("27") or zip_clean.startswith("28")):
                    continue
                if zip_clean in EXCLUDE_ZIPS:
                    continue
            try:
                fitness = float(row.get("combined_fitness_score") or 0)
            except Exception:
                fitness = 0.0
            # Reclassify transition-signal rows: they don't have a confirmed PEO,
            # so the displacement opener would lie. Route as wc_renewal instead.
            current_peo_raw = (row.get("current_peo") or "").strip()
            evidence_raw = (row.get("current_peo_evidence") or "").strip()
            evidence_low = evidence_raw.lower()
            is_transition = (
                current_peo_raw.lower().startswith("unknown") or
                "transition signal" in current_peo_raw.lower() or
                "transition signal" in evidence_low
            )
            # Stale-evidence flag — when set, the opener softens from assertion to verification
            is_stale = any(w in evidence_low for w in ("expired", "lapsed", "cancelled", "canceled"))

            wc_carrier_extracted = ""
            if is_transition:
                primary_trigger = "wc_renewal"
                # Try to extract WC carrier from evidence (e.g., "carrier ACCIDENT FUND ...")
                m = re.search(r"carrier\s+([A-Z][A-Z0-9 ,&'.-]{3,80})", evidence_raw)
                if m: wc_carrier_extracted = m.group(1).strip().rstrip(",")
            else:
                primary_trigger = "displacement"

            seen.add(key)
            accounts.append({
                "key": key,
                "company": company,
                "city": (row.get("city") or "").strip(),
                "county": (row.get("county") or "").strip(),
                "zip": zip_clean,
                "phone": (row.get("phone") or "").strip().rstrip(".0"),
                "ee": ee_int,
                "vertical": (row.get("recommended_section") or "").replace("§", "").strip()[:30] or "",
                "score": fitness,
                "primary_trigger": primary_trigger,
                "incumbent_peo": current_peo_raw if not is_transition else "",
                "incumbent_evidence": evidence_raw,
                "incumbent_stale": is_stale,
                "wc_carrier": wc_carrier_extracted,
                "dm_name": (row.get("named_dm_name") or "").strip(),
                "dm_email": (row.get("named_dm_email") or "").strip(),
                "fitness_tier": (row.get("fitness_tier") or "").strip(),
                "talk_track": (row.get("talk_track") or "").strip(),
                "evidence": (row.get("signals_seen") or "").strip()[:200],
            })
    return accounts


def load_displacement_universe():
    """Public wrapper — returns shallow copies of cached records so callers can mutate
    (apply_weights_to_bench rewrites .score) without poisoning the cache.
    Also runs enrichment + unreachable filter on each call (cheap on a 500-row list)."""
    cached = _load_displacement_universe_cached()
    out = [dict(rec) for rec in cached]
    enrich_maps = load_enrichment_maps()
    n_enriched = 0
    n_dropped_unreachable = 0
    cleaned = []
    for a in out:
        before = (a.get("phone"), a.get("dm_name"), a.get("dm_email"))
        enrich_record(a, enrich_maps)
        if (a.get("phone"), a.get("dm_name"), a.get("dm_email")) != before:
            n_enriched += 1
        # Unreachable filter: if no DM + no DM email + no phone, we can't actually touch them. Drop.
        if not a.get("dm_name") and not a.get("dm_email") and not a.get("phone"):
            n_dropped_unreachable += 1
            continue
        cleaned.append(a)
    if n_enriched:
        print(f"  Displacement enriched: {n_enriched} accounts (phone/DM backfill).")
    if n_dropped_unreachable:
        print(f"  Displacement dropped: {n_dropped_unreachable} unreachable (no DM + no phone).")
    return cleaned


@functools.cache
def _county_lookup_cached():
    """{normalized_name → (county, zip)} from companies table. Cached across calls."""
    db = sqlite3.connect(str(DB))
    name_lookup = {}
    for r in db.execute("SELECT name_normalized, name_display, county, zip FROM companies").fetchall():
        n_norm, n_disp, county, zp = r
        if not county: continue
        key = re.sub(r"[^a-z0-9]","", (n_disp or "").lower())
        if key and key not in name_lookup:
            name_lookup[key] = (county, zp)
    db.close()
    return name_lookup


# ─── INDUSTRY SIGNAL TRIGGERS (compliance / talent / consolidation) ─────────────
# Maps pipeline.sqlite pitch_signal_primary values + supplemental fields to our trigger keys.
# Each has a base score weight tuned by signal strength + playbook conversion expectation.
INDUSTRY_SIGNAL_TRIGGERS = {
    "WC_LAPSED":             ("wc_lapsed",            5.5),   # Compliance violation (NC requires WC at 3+ EE) — highest urgency
    "WC_RECENT_CHANGE":      ("wc_recent_change",     4.5),   # Carrier swap last 90 days — they're already shopping
    "OSHA_CONFIRMED":        ("osha_confirmed",       4.5),   # Playbook §1.2 trigger 5
    "FOREIGN_LABOR":         ("foreign_labor",        4.0),   # H1B/LCA filings = proven talent shortage
    "CARRIER_CONSOLIDATION": ("carrier_consolidation",3.5),   # 3+ health carriers = admin pain
    "SHRINKING_COUNTY":      ("shrinking_county",     3.0),   # County losing jobs → retention play
    "DOD_FIRST_AWARD":       ("dod_contractor",       3.5),   # New DoD contract = SCA/FAR compliance burden = PEO need
    "COMPETING_PEO":         ("displacement",         3.0),   # Backstop for COMPETING_PEO signal not in displacement CSV
    "ABC_RENEWAL":           ("abc_renewal",          0.0),   # Hospitality — off-focus per 2026-05; intentionally 0 weight
    "FMCSA_DRIVERS":         ("fmcsa",                0.0),   # Trucking — out of focus; 0 weight
    "ADP_401K_UPSELL":       ("adp_upsell",           0.0),   # Already a customer — different play
}

@functools.cache
def load_industry_signal_universe():
    """Pull from pipeline.sqlite pre-computed pitch signals.
    Returns list of accounts with primary_trigger ∈ {wc_lapsed, osha_confirmed, foreign_labor, ...}.
    Adds WARN-filed (workforce upheaval) and federal contractor (SCA compliance) as separate triggers."""
    if not DB.exists(): return []
    db = sqlite3.connect(str(DB))
    accounts = []
    seen = set()

    # Pre-fetch helpers (single query each, cached at module level)
    excluded_keys = load_excluded_keys()
    naics_map = load_naics_map()
    enrich_maps = load_enrichment_maps()
    zip_to_county = _zip_to_county_cached()

    # Pull every company that has a non-DEFAULT pitch_signal_primary in our trigger map
    placeholders = ",".join(["?"] * len(INDUSTRY_SIGNAL_TRIGGERS))
    rows = db.execute(f"""
        SELECT c.company_id, c.name_display, c.city, c.zip, c.county, e.value AS signal,
               (SELECT value FROM enrichments WHERE company_id=c.company_id AND field='ee_count' LIMIT 1) AS ee,
               (SELECT value FROM enrichments WHERE company_id=c.company_id AND field='pitch_angle' LIMIT 1) AS angle,
               (SELECT value FROM enrichments WHERE company_id=c.company_id AND field='phone' LIMIT 1) AS phone,
               (SELECT value FROM enrichments WHERE company_id=c.company_id AND field='primary_contact_name' LIMIT 1) AS dm,
               (SELECT value FROM enrichments WHERE company_id=c.company_id AND field='contact_email' LIMIT 1) AS dm_email
        FROM enrichments e JOIN companies c ON c.company_id = e.company_id
        WHERE e.field='pitch_signal_primary' AND e.value IN ({placeholders})
          AND c.state = 'NC'
    """, tuple(INDUSTRY_SIGNAL_TRIGGERS.keys())).fetchall()

    n_dropped = {"ee": 0, "zip": 0, "name": 0, "excluded": 0, "unreachable": 0, "zero_weight": 0}
    for cid, name, city, zp, county, signal, ee_raw, angle, phone, dm, dm_email in rows:
        if not name: continue
        key = _norm_key(name)
        if not key or key in seen: continue

        trigger_key, weight = INDUSTRY_SIGNAL_TRIGGERS[signal]
        if weight == 0:
            n_dropped["zero_weight"] += 1; continue

        # Excluded keys (worked / drops / registry)
        if key in excluded_keys:
            n_dropped["excluded"] += 1; continue

        # Name filter
        if (EXCLUDE_NAME_REGEX.search(name)
            or any(s in name.upper() for s in EXCLUDE_NAME_SUBSTR)
            or is_off_focus_by_name(name)):
            n_dropped["name"] += 1; continue

        # NAICS filter
        naics = naics_map.get(key, "")
        if is_naics_excluded(naics):
            n_dropped["name"] += 1; continue

        # EE band
        try: ee_int = int(float(ee_raw)) if ee_raw else None
        except (ValueError, TypeError): ee_int = None
        if ee_int is not None and (ee_int < ICP_EE_MIN or ee_int > ICP_EE_MAX):
            n_dropped["ee"] += 1; continue

        # NC zip
        zip_clean = ""
        if zp:
            zip_clean = str(zp).strip().split(".")[0].zfill(5)[:5]
            if not (zip_clean.startswith("27") or zip_clean.startswith("28")):
                n_dropped["zip"] += 1; continue
            if zip_clean in EXCLUDE_ZIPS:
                n_dropped["zip"] += 1; continue

        # County fallback
        if not county and zip_clean and zip_clean in zip_to_county:
            county = zip_to_county[zip_clean]

        # Reachable filter
        phone_clean = re.sub(r"\D", "", phone or "")[:11]
        if not dm and not dm_email and not phone_clean:
            n_dropped["unreachable"] += 1; continue

        seen.add(key)
        accounts.append({
            "key": key,
            "company": name,
            "city": city or "",
            "county": county or "",
            "zip": zip_clean,
            "phone": phone_clean,
            "ee": ee_int,
            "vertical": "",   # falls through to NAICS-based / enrichment-driven inference
            "naics": naics,
            "score": weight,
            "primary_trigger": trigger_key,
            "evidence": (angle or signal)[:200],
            "dm_name": dm or "",
            "dm_email": dm_email or "",
        })

    db.close()

    # Also pull WARN Act filings (workforce upheaval) — playbook signal not in pitch_signal_primary
    db = sqlite3.connect(str(DB))
    warn_cutoff = (TODAY - dt.timedelta(days=180)).isoformat()  # last 6 months
    warn_rows = db.execute("""
        SELECT c.company_id, c.name_display, c.city, c.zip, c.county, e.value AS warn_date,
               (SELECT value FROM enrichments WHERE company_id=c.company_id AND field='ee_count' LIMIT 1),
               (SELECT value FROM enrichments WHERE company_id=c.company_id AND field='phone' LIMIT 1),
               (SELECT value FROM enrichments WHERE company_id=c.company_id AND field='primary_contact_name' LIMIT 1),
               (SELECT value FROM enrichments WHERE company_id=c.company_id AND field='contact_email' LIMIT 1)
        FROM enrichments e JOIN companies c ON c.company_id = e.company_id
        WHERE e.field='warn_filed_date' AND e.value >= ? AND c.state='NC'
    """, (warn_cutoff,)).fetchall()
    for cid, name, city, zp, county, warn_date, ee_raw, phone, dm, dm_email in warn_rows:
        if not name: continue
        key = _norm_key(name)
        if not key or key in seen or key in excluded_keys: continue
        if EXCLUDE_NAME_REGEX.search(name) or is_off_focus_by_name(name): continue
        try: ee_int = int(float(ee_raw)) if ee_raw else None
        except (ValueError, TypeError): ee_int = None
        if ee_int is not None and (ee_int < ICP_EE_MIN or ee_int > ICP_EE_MAX): continue
        zip_clean = str(zp or "").strip().split(".")[0].zfill(5)[:5] if zp else ""
        if zip_clean and not (zip_clean.startswith("27") or zip_clean.startswith("28")): continue
        if zip_clean in EXCLUDE_ZIPS: continue
        if not county and zip_clean and zip_clean in zip_to_county:
            county = zip_to_county[zip_clean]
        phone_clean = re.sub(r"\D", "", phone or "")[:11]
        if not dm and not dm_email and not phone_clean: continue
        seen.add(key)
        accounts.append({
            "key": key, "company": name, "city": city or "", "county": county or "",
            "zip": zip_clean, "phone": phone_clean, "ee": ee_int, "vertical": "",
            "naics": naics_map.get(key, ""), "score": 5.0,
            "primary_trigger": "warn_layoff",
            "evidence": f"WARN Act filing {warn_date} — mass layoff in progress",
            "dm_name": dm or "", "dm_email": dm_email or "",
        })
    db.close()

    print(f"  Industry signals loaded: {len(accounts)} accounts (dropped: {n_dropped})")
    return accounts
    """{vertical_label → {'wc_top':[(carrier,n)...], 'health_top':[...], 'peo_top':[...]}}.
    Built from pipeline.sqlite: joins companies → enrichments (naics_inferred + wc_carrier + competing_peo)
    + carriers table for health. Used to power the Industry_Trends tab."""
    out = defaultdict(lambda: {"wc_top": Counter(), "health_top": Counter(), "peo_top": Counter(), "n_companies": 0})
    if not DB.exists(): return out
    db = sqlite3.connect(str(DB))
    # Step 1: company_id → vertical
    company_to_vert = {}
    for cid, naics in db.execute(
        "SELECT company_id, value FROM enrichments WHERE field='naics_inferred'"
    ).fetchall():
        v = _naics_to_vertical(naics or "")
        if v: company_to_vert[cid] = v
    for v in company_to_vert.values():
        out[v]["n_companies"] += 1
    # Step 2: WC carrier per company
    for cid, val in db.execute(
        "SELECT company_id, value FROM enrichments WHERE field='wc_carrier' AND value IS NOT NULL"
    ).fetchall():
        v = company_to_vert.get(cid)
        if v and val: out[v]["wc_top"][str(val).strip().upper()[:50]] += 1
    # Step 3: Health carriers from carriers table
    for cid, name in db.execute("SELECT company_id, carrier_name FROM carriers WHERE carrier_name IS NOT NULL").fetchall():
        v = company_to_vert.get(cid)
        if v and name: out[v]["health_top"][str(name).strip().upper()[:50]] += 1
    # Step 4: Incumbent PEO per company (EXCLUDING ADP TotalSource — those are Nick's own clients,
    # not displacement targets)
    for cid, val in db.execute(
        "SELECT company_id, value FROM enrichments WHERE field='competing_peo_brand' AND value IS NOT NULL"
    ).fetchall():
        v = company_to_vert.get(cid)
        if not v or not val: continue
        sval = str(val).strip()
        if "adp" in sval.lower() or "totalsource" in sval.lower(): continue
        out[v]["peo_top"][sval[:30]] += 1
    db.close()
    return out


@functools.cache
def _industry_carrier_rollup_cached():
    """{vertical_label → {'wc_top':Counter, 'health_top':Counter, 'peo_top':Counter, 'n_companies':int}}
    from pipeline.sqlite. Used by the Industry_Trends tab."""
    out = defaultdict(lambda: {"wc_top": Counter(), "health_top": Counter(), "peo_top": Counter(), "n_companies": 0})
    if not DB.exists(): return out
    db = sqlite3.connect(str(DB))
    company_to_vert = {}
    for cid, naics in db.execute(
        "SELECT company_id, value FROM enrichments WHERE field='naics_inferred'"
    ).fetchall():
        v = _naics_to_vertical(naics or "")
        if v: company_to_vert[cid] = v
    for v in company_to_vert.values():
        out[v]["n_companies"] += 1
    for cid, val in db.execute(
        "SELECT company_id, value FROM enrichments WHERE field='wc_carrier' AND value IS NOT NULL"
    ).fetchall():
        v = company_to_vert.get(cid)
        if v and val: out[v]["wc_top"][str(val).strip().upper()[:50]] += 1
    for cid, name in db.execute("SELECT company_id, carrier_name FROM carriers WHERE carrier_name IS NOT NULL").fetchall():
        v = company_to_vert.get(cid)
        if v and name: out[v]["health_top"][str(name).strip().upper()[:50]] += 1
    # Incumbent PEO (excluding ADP — those are won territory, not targets)
    for cid, val in db.execute(
        "SELECT company_id, value FROM enrichments WHERE field='competing_peo_brand' AND value IS NOT NULL"
    ).fetchall():
        v = company_to_vert.get(cid)
        if not v or not val: continue
        sval = str(val).strip()
        if "adp" in sval.lower() or "totalsource" in sval.lower(): continue
        out[v]["peo_top"][sval[:30]] += 1
    db.close()
    return out


def _naics_to_vertical(naics: str) -> str:
    """Map NAICS prefix to a human-readable vertical bucket."""
    if not naics: return ""
    n = str(naics).strip()
    if n.startswith(("31","32","33")): return "Manufacturing"
    if n.startswith("23822") or n.startswith("23823"): return "HVAC"
    if n.startswith("23821"): return "Electrical"
    if n.startswith("5413"): return "Engineering"
    if n.startswith("5415") or n.startswith("5112") or n.startswith("5182"): return "Tech / IT"
    if n.startswith("5416"): return "Mgmt Consulting"
    if n.startswith("5419") or n.startswith("5418") or n.startswith("5414"): return "Other Prof Svcs"
    if n.startswith("5411"): return "Legal"
    if n.startswith("5417"): return "Biotech / R&D"
    if n.startswith("62"): return "Health Care"
    if n.startswith("52"): return "Finance/Insurance"
    if n.startswith("42"): return "Wholesale"
    if n.startswith("488") or n.startswith("493"): return "Logistics"
    if n.startswith("7211"): return "Hospitality"
    return ""


@functools.cache
def _zip_to_county_cached():
    """{zip5 → most-common-county} from companies table. Used to backfill county when only ZIP is known."""
    from collections import Counter, defaultdict
    db = sqlite3.connect(str(DB))
    by_zip = defaultdict(Counter)
    for zp, ct in db.execute("SELECT zip, county FROM companies WHERE zip IS NOT NULL AND county IS NOT NULL").fetchall():
        if zp and ct:
            z5 = str(zp).strip().split(".")[0].zfill(5)[:5]
            by_zip[z5][ct.strip()] += 1
    db.close()
    return {z: ct.most_common(1)[0][0] for z, ct in by_zip.items()}


def enrich_with_county(accounts):
    """Fill in county from sqlite for each account where missing.
    Two-pass: (1) name match against companies table, (2) ZIP → mode county fallback."""
    name_lookup = _county_lookup_cached()
    zip_lookup = _zip_to_county_cached()
    n_by_name = n_by_zip = 0
    for a in accounts:
        if a.get("county"): continue
        hit = name_lookup.get(a["key"])
        if hit:
            a["county"] = hit[0]
            if not a.get("zip"): a["zip"] = hit[1] or ""
            n_by_name += 1
            continue
        # Fallback: infer from ZIP
        zp = (a.get("zip") or "").strip().split(".")[0].zfill(5)[:5]
        if zp and zp in zip_lookup:
            a["county"] = zip_lookup[zp]
            n_by_zip += 1
        else:
            a["county"] = ""
    if n_by_name or n_by_zip:
        print(f"  County enriched: +{n_by_name} by name, +{n_by_zip} by ZIP")


# ─── DAILY ACTION QUEUE GENERATION ──────────────────────────────────
def what_action_today(start_date_iso: str):
    """Given a cadence start date, return (day_offset, channel, label, time_block) for today,
    or None if today is not an action day."""
    start = dt.date.fromisoformat(start_date_iso)
    days = (TODAY - start).days
    for d, ch, lbl, tb in CADENCE:
        if d == days:
            return (d, ch, lbl, tb)
    return None


def build_opener_for(rec, channel):
    """Fill in template variables for the given trigger type + channel.
    Displacement triggers route through the incumbent-specific opener bank."""
    trigger_key = rec.get("primary_trigger", "default")

    # Route displacement to incumbent-specific bank
    if trigger_key == "displacement":
        incumbent = incumbent_canonical(rec.get("incumbent_peo", ""))
        if incumbent in INCUMBENT_OPENERS:
            template_set = INCUMBENT_OPENERS[incumbent]
        else:
            template_set = OUTREACH_OPENERS.get("default")
    else:
        template_set = OUTREACH_OPENERS.get(trigger_key, OUTREACH_OPENERS["default"])

    if channel == "email":
        body_template = template_set["email_body"]
        subject_template = template_set["email_subject"]
    elif channel == "linkedin":
        body_template = template_set["linkedin"]
        subject_template = ""
    elif channel == "drop":
        body_template = template_set["drop_script"]
        subject_template = ""
    elif channel == "call":
        body_template = template_set["call_script"]
        subject_template = ""
    else:
        return ("", "")

    # Variable substitution
    company = rec.get("company", "")
    first_name = rec.get("first_name") or "[OWNER FIRST NAME — verify]"
    owner = first_name if first_name and not first_name.startswith("[") else "[OWNER]"
    vertical = rec.get("vertical") or "your vertical"
    county = rec.get("county") or "your area"
    ee = rec.get("ee") or "[N]"
    carrier = rec.get("wc_carrier") or "[carrier]"
    renewal_date = rec.get("wc_renewal") or "[renewal date]"
    days_to_renewal = rec.get("days_out") or "[N]"
    plan_year = rec.get("plan_year") or "[plan year]"
    n_jobs = rec.get("n_jobs") or "several"
    dm_name = rec.get("dm_name") or first_name
    dm_email = rec.get("dm_email") or "[email — verify]"
    v_hook = vertical_hook(rec.get("vertical","")) or (
        "Most owners in your size band are seeing renewal volatility this cycle that the "
        "master-plan structure directly insulates against."
    )

    # WC opener gracefully degrades when carrier is missing (the trigger source has gaps)
    has_carrier = bool(rec.get("wc_carrier"))
    has_renewal = bool(rec.get("wc_renewal"))
    if has_carrier and has_renewal:
        wc_lead = (f"Pulled your workers' comp coverage from the NC Industrial Commission database — "
                   f"looks like your policy with {carrier} renews around {renewal_date}.")
        wc_linkedin = f"I also pulled your specific carrier and expiration ({carrier}, renewal {renewal_date})."
    elif has_renewal:
        wc_lead = (f"Saw your workers' comp policy is up for renewal around {renewal_date}. "
                   f"Most {vertical} owners I talk to in eastern NC are starting to look at options 60–90 days out.")
        wc_linkedin = f"Saw your renewal hits around {renewal_date}."
    else:
        wc_lead = (f"Most {vertical} owners I'm talking to in eastern NC are seeing WC modifier "
                   f"and rate movement this cycle worth a 5-minute read.")
        wc_linkedin = ""

    subs = dict(
        FIRST_NAME=first_name, OWNER=owner, COMPANY=company, VERTICAL=vertical, COUNTY=county,
        EE=ee, CARRIER=carrier, RENEWAL_DATE=renewal_date, DAYS_TO_RENEWAL=days_to_renewal,
        PLAN_YEAR=plan_year, N_JOBS=n_jobs, N=ee, DATE=rec.get("evidence_date","[date]"),
        DM_NAME=dm_name, DM_EMAIL=dm_email, VERTICAL_HOOK=v_hook,
        INCUMBENT=rec.get("incumbent_peo","[incumbent]"),
        WC_LEAD_LINE=wc_lead, WC_LINKEDIN_REF=wc_linkedin,
    )
    # Robust .format with missing keys → leaves placeholder visible rather than crashing
    class _Safe(dict):
        def __missing__(self, key): return "[" + key + "]"
    safe = _Safe(subs)
    subject = subject_template.format_map(safe) if subject_template else ""
    body = body_template.format_map(safe)
    return (subject, body)


# ─── PORTFOLIO SELECTION + INITIALIZATION ───────────────────────────
def init_portfolio(state, accounts):
    """Day-1 init: seed 10 cadences per active route day for the next 4 weekdays."""
    print("\nInitializing portfolio…")
    print("  Seeding stagger-started cadences across Mon-Thu route days.")
    if state.get("active_cadences"):
        print(f"  [warn] state already has {len(state['active_cadences'])} active cadences. Skipping init. Use --reset-state to start over.")
        return state

    # For each route day this week + next, pick top-scored accounts whose county matches
    # that day's route and start their cadences on that day.
    # Lane balance: cap displacement at 6/10 per route-day so trigger lane stays alive.
    # Per Nick 2026-05-11: 50 active total, distributed 12/12/13/13 across Mon/Tue/Wed/Thu.
    SLOTS_PER_DAY = {0: 12, 1: 12, 2: 13, 3: 13}   # Mon Wake / Tue Pitt / Wed Northern / Thu Cumberland
    DISP_CAP = 7   # max displacement per route day (lane balance: 7 disp + 5-6 other)
    route_buckets = {0: [], 1: [], 2: [], 3: []}
    overflow = []
    bucket_disp = {0: [], 1: [], 2: [], 3: []}
    bucket_trig = {0: [], 1: [], 2: [], 3: []}
    overflow_disp, overflow_trig = [], []
    for a in accounts:
        rd = county_to_route_day(a.get("county",""))
        is_disp = a.get("primary_trigger") == "displacement"
        if rd in (0, 1, 2, 3):
            (bucket_disp if is_disp else bucket_trig)[rd].append(a)
        else:
            (overflow_disp if is_disp else overflow_trig).append(a)
    for rd in (0, 1, 2, 3):
        bucket_disp[rd].sort(key=lambda x: -x["score"])
        bucket_trig[rd].sort(key=lambda x: -x["score"])
    overflow_disp.sort(key=lambda x: -x["score"])
    overflow_trig.sort(key=lambda x: -x["score"])

    # STRICT GEO: only seed from accounts whose county is actually in the route-day's cluster.
    # If we can't fill 10 in-cluster, take fewer. Better fewer accounts than cross-county drops.
    for rd in (0, 1, 2, 3):
        target = SLOTS_PER_DAY[rd]
        # take up to DISP_CAP displacement IN-CLUSTER only (no overflow)
        d_take = list(bucket_disp[rd][:DISP_CAP])
        # take remaining trigger IN-CLUSTER only
        slots_left = target - len(d_take)
        t_take = list(bucket_trig[rd][:slots_left])
        # if displacement bucket was under-full, top with more trigger (still in-cluster)
        combined = d_take + t_take
        if len(combined) < target:
            extras = bucket_trig[rd][slots_left:slots_left + (target - len(combined))]
            combined.extend(extras)
        if len(combined) < target:
            extras = bucket_disp[rd][DISP_CAP:DISP_CAP + (target - len(combined))]
            combined.extend(extras)
        route_buckets[rd] = combined
        print(f"  RD{rd} ({ROUTE_BY_WEEKDAY[rd][0][:24]}): {len(combined)}/{target} in-cluster accounts seeded")

    # Seed 10 per route day, starting on the next occurrence of that weekday
    seeded = 0
    for rd in (0, 1, 2, 3):
        # Find the next occurrence of weekday rd from today (inclusive if today matches, else next)
        days_ahead = (rd - WEEKDAY) % 7
        if days_ahead == 0 and WEEKDAY > rd:
            days_ahead = 7
        start_date = TODAY + dt.timedelta(days=days_ahead)
        for a in route_buckets[rd][:SLOTS_PER_DAY[rd]]:
            touches = build_touch_schedule(start_date, rd)
            cadence = {
                "key": a["key"],
                "company": a["company"],
                "city": a.get("city",""),
                "county": a.get("county",""),
                "zip": a.get("zip",""),
                "phone": a.get("phone",""),
                "ee": a.get("ee"),
                "vertical": a.get("vertical",""),
                "naics": a.get("naics",""),
                "score": a.get("score",0),
                "score_raw": a.get("score_raw", 0),
                "score_overlay": a.get("score_overlay", 0),
                "weight_mult": a.get("weight_mult", 1.0),
                "has_health_benefits": a.get("has_health_benefits", False),
                "multi_state_likely": a.get("multi_state_likely", False),
                "growth_tier": a.get("growth_tier", ""),
                "multi_carrier_consolidation": a.get("multi_carrier_consolidation", 0),
                "primary_trigger": a.get("primary_trigger","default"),
                "wc_carrier": a.get("wc_carrier",""),
                "wc_renewal": a.get("wc_renewal",""),
                "days_out": a.get("days_out"),
                "evidence": a.get("evidence",""),
                "plan_year": a.get("plan_year",""),
                "incumbent_peo": a.get("incumbent_peo",""),
                "incumbent_stale": a.get("incumbent_stale", False),
                "dm_name": a.get("dm_name",""),
                "dm_email": a.get("dm_email",""),
                "fitness_tier": a.get("fitness_tier",""),
                "talk_track": a.get("talk_track",""),
                "start_date": start_date.isoformat(),
                "route_day": rd,
                "status": "active",
                "touches": touches,
                "first_name": "",
                "tier": "Top" if a["score"] >= 8.0 else "Mid",
            }
            # Ensure buyer_cast slot exists for this account
            state["buyer_cast"].setdefault(a["key"], {
                "owner": "", "cfo": "", "office_mom": "",
                "broker": "", "cpa": "", "attorney": "",
            })
            state["active_cadences"].append(cadence)
            seeded += 1

    state["initialized_at"] = TODAY_ISO
    print(f"  Seeded {seeded} cadences across the next 4 active route days.")
    return state


def ingest_weekly_wrap(state):
    """Read prior workbook's Weekly_Wrap tab → apply outcomes to state.
    Idempotent: only ingests un-completed touches with non-blank outcome."""
    if not OUT.exists(): return 0
    try:
        wb = load_workbook(OUT, read_only=True, data_only=True)
    except Exception as e:
        print(f"  [warn] could not open prior workbook for ingest: {e}", file=sys.stderr)
        return 0
    if "Weekly_Wrap" not in wb.sheetnames:
        wb.close(); return 0
    ws = wb["Weekly_Wrap"]
    rows = list(ws.iter_rows(min_row=5, values_only=True))
    wb.close()

    n_ingested = 0
    valid_outcomes = set(o for o in OUTCOMES if o)
    for r in rows:
        if not r or len(r) < 9: continue
        outcome = (r[5] or "").strip() if isinstance(r[5], str) else ""
        coord = (r[8] or "").strip() if isinstance(r[8], str) else ""
        if not outcome or not coord or outcome not in valid_outcomes:
            continue
        try:
            ci_str, ti_str = coord.split(":")
            ci, ti = int(ci_str), int(ti_str)
        except Exception:
            continue
        if ci >= len(state["active_cadences"]): continue
        c = state["active_cadences"][ci]
        if ti >= len(c["touches"]): continue
        t = c["touches"][ti]
        if t.get("completed"): continue   # already ingested

        broker = (r[6] or "").strip() if isinstance(r[6], str) else ""
        notes = (r[7] or "").strip() if isinstance(r[7], str) else ""

        t["completed"] = True
        t["outcome"] = outcome
        t["broker_captured"] = broker
        t["notes"] = notes

        # Capture broker into buyer_cast AND auto-add to channel_build recruitment list
        if broker:
            cast = state["buyer_cast"].setdefault(c["key"], {})
            if not cast.get("broker"):
                cast["broker"] = broker
            # Auto-recruit: if broker not yet in channel_build.brokers, add
            existing_brokers = {b["name"].lower() for b in state["channel_build"]["brokers"]}
            if broker.lower() not in existing_brokers:
                state["channel_build"]["brokers"].append({
                    "name": broker,
                    "first_seen": TODAY_ISO,
                    "first_seen_via": c["company"],
                    "county": c.get("county", ""),
                    "phone": "",
                    "email": "",
                    "last_brief_sent": "",
                    "lunch_status": "not yet",
                    "n_clients_observed": 1,
                    "notes": f"First captured during {c.get('primary_trigger','')} cadence on {c['company']}.",
                })
            else:
                # Increment client count for existing broker
                for b in state["channel_build"]["brokers"]:
                    if b["name"].lower() == broker.lower():
                        b["n_clients_observed"] = b.get("n_clients_observed", 0) + 1
                        break

        # Append to ledger
        week_start = (TODAY - dt.timedelta(days=TODAY.weekday())).isoformat()
        state["weekly_outcomes"].append({
            "week_start": week_start,
            "logged_at": TODAY_ISO,
            "company_key": c["key"],
            "company": c["company"],
            "trigger": c.get("primary_trigger",""),
            "vertical": c.get("vertical",""),
            "channel": t["channel"],
            "day_offset": t["day"],
            "route_day": c.get("route_day"),
            "scheduled_for": t.get("scheduled_for",""),
            "outcome": outcome,
            "broker_captured": broker,
            "notes": notes,
        })
        n_ingested += 1

    if n_ingested:
        print(f"  Ingested {n_ingested} outcomes from last week's Weekly_Wrap.")
    return n_ingested


def recompute_weights(state, min_n=10):
    """Update state['weights'] from the weekly_outcomes ledger.
    multiplier = clamp((rep_rate / overall_rate), 0.5, 2.0). Skips dimensions with <min_n samples."""
    ledger = state.get("weekly_outcomes", [])
    n_total = len(ledger)
    if n_total < min_n:
        print(f"  Learning engine: only {n_total} outcomes logged — need ≥{min_n} before recomputing weights. Skipping.")
        return
    n_progressions = sum(1 for o in ledger if o["outcome"] in PROGRESSION_OUTCOMES)
    overall_rate = n_progressions / n_total if n_total else 0.0
    if overall_rate == 0:
        print(f"  Learning engine: 0 progressions in ledger — keeping default weights.")
        return

    def _rate(filter_fn):
        sub = [o for o in ledger if filter_fn(o)]
        if len(sub) < 5: return None  # not enough samples for this segment
        prog = sum(1 for o in sub if o["outcome"] in PROGRESSION_OUTCOMES)
        return prog / len(sub)

    def _mult(r):
        if r is None: return 1.0
        m = r / overall_rate if overall_rate > 0 else 1.0
        return max(0.5, min(2.0, m))

    # Per-trigger
    triggers = sorted({o["trigger"] for o in ledger if o.get("trigger")})
    for t in triggers:
        r = _rate(lambda o, _t=t: o["trigger"] == _t)
        state["weights"]["trigger"][t] = round(_mult(r), 3)
    # Per-vertical
    verticals = sorted({o["vertical"] for o in ledger if o.get("vertical")})
    for v in verticals:
        r = _rate(lambda o, _v=v: o["vertical"] == _v)
        state["weights"]["vertical"][v] = round(_mult(r), 3)
    # Per-channel
    for ch in ("email", "linkedin", "drop", "call"):
        r = _rate(lambda o, _ch=ch: o["channel"] == _ch)
        state["weights"]["channel"][ch] = round(_mult(r), 3)
    # Per-route-day (route_day stored as int — coerce both sides to str)
    for rd in ("0", "1", "2", "3"):
        r = _rate(lambda o, _rd=rd: str(o.get("route_day","")) == _rd)
        state["weights"]["route_day"][rd] = round(_mult(r), 3)

    state["weights"]["last_recomputed"] = TODAY_ISO
    state["weights"]["n_outcomes_at_recompute"] = n_total
    state["weights"]["overall_rate"] = round(overall_rate, 4)
    print(f"  Learning engine: recomputed from {n_total} outcomes (overall progression rate {overall_rate:.1%}).")


def _icp_focus_multiplier(rec):
    """Returns 1.5 if account matches ICP focus (Mfg/HVAC/Electric/Engineering/Tech/Prof Services),
    0.4 if clearly off-focus (incl. by name), 1.0 if unknown. Per Nick 2026-05-10."""
    vert = (rec.get("vertical") or "").lower()
    name = (rec.get("company") or "").lower()
    naics = (rec.get("naics") or "").strip()
    # Positive signal: vertical or NAICS or name match
    if any(p in vert for p in ICP_FOCUS_VERTICAL_PATTERNS):
        return 1.5
    if naics and any(naics.startswith(p) for p in ICP_FOCUS_NAICS_PREFIXES):
        return 1.5
    if any(k in name for k in ("hvac", "heating", "air condition", "mechanical",
                               "electric", "engineering", "technologies", "tech ",
                               "consulting", "consultants", "manufacturing", "mfg",
                               "machining", "machine works", "fabrication", "machinery")):
        return 1.5
    # Negative signal: vertical OR name suggests off-focus
    if any(p in vert for p in ICP_OFFOCUS_VERTICAL_PATTERNS):
        return 0.4
    if any(k in name for k in (
        "biotech", "pharmacy", "wine", "winery", "vineyard", "spirits",
        "therapy", "behavioral", "psychiatric", "developmental",
        "marketplace", "pet ", "naturally", "organic ",
        "salon", "spa ", "boutique", "retail",
        "dry clean", "laundry", "cleaners",
        "optometry", "vision ", "eye care", "eyecare", "optical",
        "olive oil", "olive ", "specialty food", "gourmet",
        "ministry", "church", "fellowship",
    )):
        return 0.4
    return 1.0


def apply_weights_to_bench(accounts, weights):
    """Two-layer scoring (per Nick 2026-05-11):
    1. Additive overlay — adds PEO buy signals (health/multi-state/growth) and demotes pure WC.
    2. Multiplicative — learned trigger/route/vertical weights × ICP focus multiplier.
    Final score = (raw + overlay) * learned_mult * focus_mult."""
    tw = weights.get("trigger", {})
    rw = weights.get("route_day", {})
    vw = weights.get("vertical", {})
    peo_signals = load_peo_signal_maps()
    for a in accounts:
        a["score_raw"] = a.get("score", 0)
        # Layer 1: additive PEO buy signal overlay
        a["score_overlay"] = round(score_overlay(a, peo_signals), 3)
        base = a["score_raw"] + a["score_overlay"]
        # Layer 2: multiplicative learned + focus
        m = 1.0
        m *= tw.get(a.get("primary_trigger",""), 1.0)
        rd = county_to_route_day(a.get("county",""))
        if rd >= 0:
            m *= rw.get(str(rd), 1.0)
        m *= vw.get(a.get("vertical",""), 1.0)
        m *= _icp_focus_multiplier(a)
        a["score"] = round(base * m, 3)
        a["weight_mult"] = round(m, 3)


def topup_replacements(state, accounts):
    """For every active route-day bucket below target, top up from the bench (highest weighted score, dedup vs active)."""
    SLOTS_PER_DAY = {0: 12, 1: 12, 2: 13, 3: 13}   # Mon Wake / Tue Pitt / Wed Northern / Thu Cumberland — 50 total
    active_keys = {c["key"] for c in state["active_cadences"]}
    nurture_keys = {c["key"] for c in state.get("nurture_queue", [])}
    by_rd = defaultdict(list)
    for c in state["active_cadences"]:
        by_rd[c["route_day"]].append(c)

    # STRICT GEO top-up: only fill empty slots from accounts ACTUALLY in the route-day cluster
    bench_by_rd = defaultdict(list)
    for a in accounts:
        if a["key"] in active_keys or a["key"] in nurture_keys: continue
        rd = county_to_route_day(a.get("county",""))
        if rd in (0, 1, 2, 3):
            bench_by_rd[rd].append(a)
    for rd in bench_by_rd: bench_by_rd[rd].sort(key=lambda x: -x.get("score",0))

    n_added = 0
    for rd in (0, 1, 2, 3):
        days_ahead = (rd - WEEKDAY) % 7
        if days_ahead == 0 and WEEKDAY > rd:
            days_ahead = 7
        start_date = TODAY + dt.timedelta(days=days_ahead)
        slots = max(0, SLOTS_PER_DAY[rd] - len(by_rd[rd]))   # clamp ≥0 — fix neg-slice bug 2026-05-11
        if slots == 0: continue
        candidates = bench_by_rd[rd][:slots]
        for a in candidates:
            if a["key"] in active_keys: continue
            touches = build_touch_schedule(start_date, rd)
            cadence = {
                "key": a["key"], "company": a["company"],
                "city": a.get("city",""), "county": a.get("county",""),
                "zip": a.get("zip",""), "phone": a.get("phone",""),
                "ee": a.get("ee"), "vertical": a.get("vertical",""),
                "naics": a.get("naics",""),
                "score": a.get("score",0),
                "primary_trigger": a.get("primary_trigger","default"),
                "wc_carrier": a.get("wc_carrier",""),
                "wc_renewal": a.get("wc_renewal",""),
                "days_out": a.get("days_out"),
                "evidence": a.get("evidence",""),
                "plan_year": a.get("plan_year",""),
                "incumbent_peo": a.get("incumbent_peo",""),
                "incumbent_stale": a.get("incumbent_stale", False),
                "dm_name": a.get("dm_name",""),
                "dm_email": a.get("dm_email",""),
                "fitness_tier": a.get("fitness_tier",""),
                "talk_track": a.get("talk_track",""),
                "start_date": start_date.isoformat(),
                "route_day": rd, "status": "active", "touches": touches,
                "first_name": "", "tier": "Top" if a.get("score",0) >= 8.0 else "Mid",
            }
            state["buyer_cast"].setdefault(a["key"], {
                "owner": "", "cfo": "", "office_mom": "",
                "broker": "", "cpa": "", "attorney": "",
            })
            state["active_cadences"].append(cadence)
            active_keys.add(a["key"])
            n_added += 1
    if n_added:
        print(f"  Top-up: added {n_added} replacement cadences from bench.")


def _write_drop_to_sqlite(company_key, company_display, reason_code, reason_detail):
    """INSERT a row into pipeline.sqlite drops (idempotent: looks up company_id by name)."""
    if not DB.exists(): return False
    try:
        db = sqlite3.connect(str(DB))
        cur = db.cursor()
        # Find company_id by normalized name
        row = cur.execute(
            "SELECT company_id FROM companies WHERE name_normalized = ? OR name_loose = ? LIMIT 1",
            (company_key, company_key),
        ).fetchone()
        if not row:
            # Best effort match on display name
            row = cur.execute(
                "SELECT company_id FROM companies WHERE LOWER(REPLACE(REPLACE(REPLACE(name_display,' ',''),'.',''),',',''))= ? LIMIT 1",
                (company_key,),
            ).fetchone()
        if not row:
            db.close()
            return False
        company_id = row[0]
        # Avoid duplicate drop with same reason on same day
        existing = cur.execute(
            "SELECT 1 FROM drops WHERE company_id = ? AND reason_code = ? AND DATE(dropped_at) = DATE(?) LIMIT 1",
            (company_id, reason_code, TODAY_ISO),
        ).fetchone()
        if existing:
            db.close()
            return True
        cur.execute(
            "INSERT INTO drops (company_id, reason_code, reason_detail, dropped_in_run, dropped_at) VALUES (?, ?, ?, ?, ?)",
            (company_id, reason_code, reason_detail, "sales_os_disposition", TODAY_ISO),
        )
        db.commit()
        db.close()
        return True
    except Exception as e:
        print(f"  [warn] could not write drop for {company_display}: {e}", file=sys.stderr)
        return False


def apply_dispositions(state):
    """Walk every active cadence's most-recent completed touch and route based on outcome.
    Kill outcomes → write to pipeline.sqlite drops + remove from active.
    Nurture outcomes → move to nurture_queue (re-fire at +90d).
    MEDDPICC outcomes → initialize MEDDPICC slot if missing."""
    n_killed = 0
    n_nurtured = 0
    n_meddpicc = 0
    for c in state["active_cadences"][:]:
        # Find the latest outcome
        latest = None
        for t in c["touches"]:
            if t.get("completed") and t.get("outcome"):
                latest = t
        if not latest: continue
        outcome = latest["outcome"]

        if outcome in KILL_OUTCOMES:
            reason_code = KILL_OUTCOMES[outcome]
            detail = f"sales_os: {outcome} after touch d{latest['day']} {latest['channel']}"
            if latest.get("notes"):
                detail += f" — {latest['notes'][:120]}"
            _write_drop_to_sqlite(c["key"], c["company"], reason_code, detail)
            c["status"] = f"killed_{outcome}"
            c["killed_at"] = TODAY_ISO
            state["completed"].append(c)
            state["active_cadences"].remove(c)
            n_killed += 1
            continue

        if outcome in NURTURE_OUTCOMES:
            c["status"] = f"nurture_{outcome}"
            c["nurture_until"] = (TODAY + dt.timedelta(days=90)).isoformat()
            state["nurture_queue"].append(c)
            state["active_cadences"].remove(c)
            n_nurtured += 1
            continue

        if outcome in MEDDPICC_OUTCOMES:
            slot = state["meddpicc"].setdefault(c["key"], {
                "company": c["company"],
                "M_metrics": "", "E_econ_buyer": "", "D1_decision_criteria": "",
                "D2_decision_process": "", "P_paper_process": "", "I_pain": "",
                "C_champion": "", "Cmp_competition": "",
                "first_meeting_date": "", "stage": "discovery", "next_action": "",
            })
            if outcome == "meeting_held" and not slot.get("first_meeting_date"):
                slot["first_meeting_date"] = latest.get("scheduled_for") or TODAY_ISO
            n_meddpicc += 1

    # Reactivate nurture-queue cadences whose +90d window has passed
    n_reactivated = 0
    for c in state["nurture_queue"][:]:
        nu = c.get("nurture_until")
        if not nu: continue
        try:
            nu_d = dt.date.fromisoformat(nu)
        except Exception:
            continue
        if nu_d <= TODAY:
            # Restart cadence: today is the new D0
            c["status"] = "active"
            c["start_date"] = TODAY_ISO
            c["route_day"] = county_to_route_day(c.get("county","")) if county_to_route_day(c.get("county","")) >= 0 else c.get("route_day", 0)
            c["touches"] = build_touch_schedule(TODAY, c["route_day"])
            c.pop("nurture_until", None)
            state["active_cadences"].append(c)
            state["nurture_queue"].remove(c)
            n_reactivated += 1

    if n_killed:    print(f"  Disposition: {n_killed} cadences killed → pipeline.sqlite drops.")
    if n_nurtured:  print(f"  Disposition: {n_nurtured} cadences moved to nurture_queue (+90d).")
    if n_meddpicc:  print(f"  Disposition: {n_meddpicc} cadences advanced to MEDDPICC tracking.")
    if n_reactivated: print(f"  Disposition: {n_reactivated} nurture-queue accounts re-activated.")


def advance_cadences(state):
    """Mark cadences as completed if past day 14 with no further actions."""
    for c in state["active_cadences"][:]:
        start = dt.date.fromisoformat(c["start_date"])
        days = (TODAY - start).days
        if days > 14 and c["status"] == "active":
            # Only auto-complete if no kill/nurture outcome already routed it elsewhere
            c["status"] = "completed_or_nurture"
            state["completed"].append(c)
            state["active_cadences"].remove(c)


# ─── BUILD WORKBOOK ─────────────────────────────────────────────────
def build_workbook(state):
    wb = Workbook()
    title_font = Font(bold=True, size=16, color="1F4E78")
    h_font     = Font(bold=True, size=11, color="FFFFFF")
    h_fill     = PatternFill("solid", fgColor="1F4E78")
    sub_fill   = PatternFill("solid", fgColor="D9E2F3")
    metric     = Font(bold=True, size=18, color="1F4E78")

    # Channel-specific row fills
    fill_email    = PatternFill("solid", fgColor="DAE3F3")
    fill_linkedin = PatternFill("solid", fgColor="E2EFD9")
    fill_drop     = PatternFill("solid", fgColor="FFE699")
    fill_call     = PatternFill("solid", fgColor="F4B084")

    # ═══ TAB 1: TODAY ═══
    ws = wb.active; ws.title = "TODAY"
    route_label, _ = ROUTE_BY_WEEKDAY.get(WEEKDAY, ("(weekend)", set()))
    weekday_name = ["Monday","Tuesday","Wednesday","Thursday","Friday","Saturday","Sunday"][WEEKDAY]
    ws["A1"] = f"TODAY — {weekday_name}, {TODAY_ISO}    |    Route: {route_label}"
    ws["A1"].font = title_font; ws.merge_cells("A1:H1")

    # Day-specific run order — Nick's REAL schedule (2026-05): training daily 14:30–16:00, field Tue/Wed/Thu
    DAY_RUN_ORDER = {
        0: "MON OFFICE · 09:00–10:00 internal mtg · 10:00–11:30 EMAIL BATCH (D0 sends + last week's D8 LinkedIn DMs + D15 calls) · 11:30–12:30 cold-dial power · 12:30–13:30 lunch + CRM · 13:30–14:30 LinkedIn outbound + Tue route plan · 14:30–16:00 TRAINING (locked) · 16:00+ wrap",
        1: "TUE FIELD · 10:00–10:30 prep · 10:30–14:30 ★ Pitt-cluster route (drops + scheduled meetings) — D2 drops for Mon's Pitt batch + any booked Tue meetings · in-transit phone calls between stops · 14:30–16:00 TRAINING (locked) · 16:00+ CRM log all drops",
        2: "WED FIELD · 10:00–10:30 prep · 10:30–14:30 ★ Northern-cluster route (drops + meetings) — D3 drops for Mon's Northern batch + booked meetings · 14:30–16:00 TRAINING (locked) · 16:00+ CRM log",
        3: "THU FIELD · 10:00–10:30 prep · 10:30–14:30 ★ Cumberland/Coastal route (drops + meetings) — D4 drops for Mon's Cumberland batch + booked meetings · 14:30–16:00 TRAINING (locked) · 16:00+ event night (chamber/trade — optional)",
        4: "FRI OFFICE · 09:00–10:00 office mtg · 10:00–11:00 BREAKUP EMAILS (D22 batch) · 11:00–12:00 Weekly_Wrap fill (Mon–Thu outcomes) · 12:00–13:00 lunch · 13:00–14:30 ★ NEXT MONDAY PREP (next 12 cadences researched + emails pre-drafted) · 14:30–16:00 TRAINING (locked) · 16:00+ Sun re-run reminder",
        5: "SAT · Off. Optional: drive-by recon if you're near a Tier-1 account.",
        6: "SUN · 20:00 → RUN: python3 scripts/build_sales_os.py — ingests Friday's Weekly_Wrap, recomputes weights, tops up to 50 active. Read Coaching_Card before bed.",
    }
    ws["A2"] = DAY_RUN_ORDER.get(WEEKDAY, "Run build_sales_os.py to refresh.")
    ws["A2"].font = Font(italic=True, color="606060"); ws.merge_cells("A2:H2")
    ws.row_dimensions[2].height = 32
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")

    # Generate today's actions from active cadences
    actions = []
    for c in state["active_cadences"]:
        action = what_action_today(c["start_date"])
        if action is None: continue
        day, channel, label, tb = action
        subject, body = build_opener_for(c, channel)
        actions.append({
            "channel": channel, "time_block": tb, "day": day,
            "company": c["company"], "city": c.get("city",""), "phone": c.get("phone",""),
            "vertical": c.get("vertical",""), "ee": c.get("ee"),
            "trigger": c.get("primary_trigger",""),
            "score": c.get("score",0),
            "subject": subject, "body": body,
            "wc_carrier": c.get("wc_carrier",""), "wc_renewal": c.get("wc_renewal",""),
            "address": f"{c.get('city','')}, {c.get('zip','')}",
            "label": label,
        })

    # Sort: emails first (early AM), then call (cold-power), then drops (field), then linkedin
    sort_order = {"email": 0, "call": 1, "drop": 2, "linkedin": 3}
    actions.sort(key=lambda x: (sort_order.get(x["channel"], 9), -x["score"]))

    hdr = ["#","Time","Ch","Co","City","Phone","Trigger","Score","Day","Subject / Script (use as opener — fill in any [BRACKETS])"]
    for i, h in enumerate(hdr, 1):
        c = ws.cell(row=4, column=i, value=h); c.font = h_font; c.fill = h_fill
    ws.freeze_panes = "A5"
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 9
    ws.column_dimensions["D"].width = 36
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 6
    ws.column_dimensions["I"].width = 5
    ws.column_dimensions["J"].width = 110

    if not actions:
        ws.cell(row=5, column=1, value="No active cadence actions due today.")
        ws.cell(row=5, column=1).font = Font(italic=True, color="C00000")
        ws.cell(row=6, column=1, value=f"State has {len(state.get('active_cadences',[]))} active cadences. "
                                       f"If 0, run with --init flag to seed.")
    else:
        for i, a in enumerate(actions, 5):
            rank = i - 4
            ws.cell(row=i, column=1, value=rank)
            ws.cell(row=i, column=2, value=a["time_block"])
            ws.cell(row=i, column=3, value=a["channel"].upper())
            ws.cell(row=i, column=4, value=a["company"][:50])
            ws.cell(row=i, column=5, value=a["city"][:18])
            ws.cell(row=i, column=6, value=a["phone"])
            ws.cell(row=i, column=7, value=a["trigger"])
            ws.cell(row=i, column=8, value=round(a["score"],1))
            ws.cell(row=i, column=9, value=f"D{a['day']}")
            content = (a["subject"] + "\n\n" if a["subject"] else "") + a["body"]
            cell = ws.cell(row=i, column=10, value=content[:32000])
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[i].height = max(60, min(280, 20 + content.count("\n") * 16))
            # Color row by channel
            fill = {"email": fill_email, "linkedin": fill_linkedin,
                    "drop": fill_drop, "call": fill_call}.get(a["channel"])
            if fill:
                for col in range(1, 11):
                    ws.cell(row=i, column=col).fill = fill

    # Footer summary
    summary_row = 5 + len(actions) + 2
    ws.cell(row=summary_row, column=1, value="TODAY'S TOTALS:").font = Font(bold=True)
    counts = defaultdict(int)
    for a in actions: counts[a["channel"]] += 1
    parts = [f"{counts.get(ch,0)} {ch}s" for ch in ("email","call","drop","linkedin")]
    ws.cell(row=summary_row, column=2, value="  |  ".join(parts) + f"  |  TOTAL = {len(actions)}")
    ws.merge_cells(start_row=summary_row, start_column=2, end_row=summary_row, end_column=10)

    # ═══ TAB 2: THIS_WEEK ═══
    ws2 = wb.create_sheet("THIS_WEEK")
    ws2["A1"] = f"THIS WEEK — Mon-Fri lookahead from {TODAY_ISO}"
    ws2["A1"].font = title_font; ws2.merge_cells("A1:G1")
    h2 = ["Day","Date","Route","Email","LinkedIn","Drop","Call","Total"]
    for i, h in enumerate(h2, 1):
        c = ws2.cell(row=3, column=i, value=h); c.font = h_font; c.fill = h_fill
    # Compute what's due each day this week (or next, if today is weekend)
    if WEEKDAY >= 5:
        monday = TODAY + dt.timedelta(days=(7 - WEEKDAY))
    else:
        monday = TODAY - dt.timedelta(days=WEEKDAY)
    for d_offset in range(5):
        target = monday + dt.timedelta(days=d_offset)
        if target < TODAY: continue  # skip past days (this week)
        day_counts = defaultdict(int)
        for c in state["active_cadences"]:
            start = dt.date.fromisoformat(c["start_date"])
            for cd_day, ch, _, _ in CADENCE:
                if start + dt.timedelta(days=cd_day) == target:
                    day_counts[ch] += 1
        row = 4 + d_offset
        wn = ["Monday","Tuesday","Wednesday","Thursday","Friday"][d_offset]
        ws2.cell(row=row, column=1, value=wn)
        ws2.cell(row=row, column=2, value=target.isoformat())
        ws2.cell(row=row, column=3, value=ROUTE_BY_WEEKDAY[d_offset][0])
        ws2.cell(row=row, column=4, value=day_counts["email"])
        ws2.cell(row=row, column=5, value=day_counts["linkedin"])
        ws2.cell(row=row, column=6, value=day_counts["drop"])
        ws2.cell(row=row, column=7, value=day_counts["call"])
        ws2.cell(row=row, column=8, value=sum(day_counts.values()))
        if target == TODAY:
            for col in range(1, 9):
                ws2.cell(row=row, column=col).fill = fill_drop
    for col, w in zip("ABCDEFGH", [12, 12, 50, 9, 11, 9, 9, 9]):
        ws2.column_dimensions[col].width = w

    # ═══ TAB 2b: Account_Pipeline (one row per active cadence — touch history + next action) ═══
    ws_ap = wb.create_sheet("Account_Pipeline")
    week_start = TODAY - dt.timedelta(days=TODAY.weekday())
    week_end_iso = (week_start + dt.timedelta(days=6)).isoformat()

    # Compute weekly target metrics from state/ledger
    ledger = state.get("weekly_outcomes", [])
    this_week_outcomes = [o for o in ledger
                          if o.get("logged_at","") >= week_start.isoformat()
                          and o.get("logged_at","") <= week_end_iso]
    n_touches_this_week = len(this_week_outcomes)
    n_meetings_this_week = sum(1 for o in this_week_outcomes if o.get("outcome") in {"meeting_booked","meeting_held"})

    # ── Weekly Targets banner ──
    ws_ap["A1"] = f"ACCOUNT PIPELINE — week of {week_start.isoformat()}    |    {len(state['active_cadences'])} active / target 50"
    ws_ap["A1"].font = title_font; ws_ap.merge_cells("A1:K1")

    ws_ap.cell(row=2, column=1, value="TARGETS:").font = Font(bold=True)
    ws_ap.cell(row=2, column=2, value="Touches:").font = Font(bold=True)
    cell = ws_ap.cell(row=2, column=3, value=f"{n_touches_this_week} / 45"); cell.font = metric
    if n_touches_this_week >= 45: cell.fill = fill_linkedin
    elif n_touches_this_week >= 30: cell.fill = fill_drop
    else: cell.fill = PatternFill("solid", fgColor="FFCCCC")
    ws_ap.cell(row=2, column=4, value="Active:").font = Font(bold=True)
    cell = ws_ap.cell(row=2, column=5, value=f"{len(state['active_cadences'])} / 50"); cell.font = metric
    if len(state['active_cadences']) >= 50: cell.fill = fill_linkedin
    elif len(state['active_cadences']) >= 40: cell.fill = fill_drop
    else: cell.fill = PatternFill("solid", fgColor="FFCCCC")
    ws_ap.cell(row=2, column=6, value="Meetings:").font = Font(bold=True)
    cell = ws_ap.cell(row=2, column=7, value=f"{n_meetings_this_week} / 3"); cell.font = metric
    if n_meetings_this_week >= 3: cell.fill = fill_linkedin
    elif n_meetings_this_week >= 1: cell.fill = fill_drop
    else: cell.fill = PatternFill("solid", fgColor="FFCCCC")

    # ── Per-account row ──
    h_ap = ["#","Company","County","EE","Trigger","D-in","History","Next Action","Date","DM?","Broker?"]
    for i, h in enumerate(h_ap, 1):
        c = ws_ap.cell(row=4, column=i, value=h); c.font = h_font; c.fill = h_fill
    ws_ap.freeze_panes = "A5"

    def _touch_symbol(t):
        ch = t.get("channel","")
        done = t.get("completed", False)
        sym = {"email":"E", "linkedin":"L", "drop":"D", "call":"P", "breakup":"B"}.get(ch, "?")
        return sym + ("✓" if done else "·")

    def _next_action(c):
        """Find the next un-completed touch."""
        for t in c.get("touches", []):
            if not t.get("completed"):
                sd = t.get("scheduled_for","")
                ch = t.get("channel","")
                day = t.get("day","?")
                return (f"D{day} {ch.upper()}", sd)
        return ("done — disposition", "")

    sorted_active = sorted(state["active_cadences"], key=lambda x: -x.get("score",0))
    for i, c in enumerate(sorted_active, 5):
        rank = i - 4
        try:
            start = dt.date.fromisoformat(c["start_date"])
            days_in = (TODAY - start).days
        except Exception:
            days_in = "?"
        history = " ".join(_touch_symbol(t) for t in c.get("touches", []))
        action, date = _next_action(c)
        cast = state.get("buyer_cast", {}).get(c["key"], {})
        dm_ok = "✓" if c.get("dm_name") else "—"
        broker_ok = "✓" if cast.get("broker") else "—"

        ws_ap.cell(row=i, column=1, value=rank)
        ws_ap.cell(row=i, column=2, value=c["company"][:40])
        ws_ap.cell(row=i, column=3, value=c.get("county",""))
        ws_ap.cell(row=i, column=4, value=c.get("ee"))
        ws_ap.cell(row=i, column=5, value=c.get("primary_trigger",""))
        ws_ap.cell(row=i, column=6, value=f"D{days_in}")
        ws_ap.cell(row=i, column=7, value=history).font = Font(name="Courier New", size=10)
        ws_ap.cell(row=i, column=8, value=action)
        ws_ap.cell(row=i, column=9, value=date)
        dm_cell = ws_ap.cell(row=i, column=10, value=dm_ok)
        if not c.get("dm_name"): dm_cell.fill = PatternFill("solid", fgColor="FFE0E0")
        broker_cell = ws_ap.cell(row=i, column=11, value=broker_ok)
        if not cast.get("broker"): broker_cell.fill = PatternFill("solid", fgColor="FFE0E0")
        # Highlight next-action-today rows
        if date == TODAY_ISO:
            for col in range(1, 12):
                ws_ap.cell(row=i, column=col).fill = fill_drop

    # Legend
    last_row = 5 + len(sorted_active) + 2
    ws_ap.cell(row=last_row, column=1, value="LEGEND:").font = Font(bold=True)
    ws_ap.cell(row=last_row, column=2, value="E=Email  L=LinkedIn  D=Drop  P=Phone  B=Breakup    ✓=done  ·=scheduled").font = Font(size=10, italic=True, color="606060")
    ws_ap.merge_cells(start_row=last_row, start_column=2, end_row=last_row, end_column=11)

    for col, w in zip("ABCDEFGHIJK", [4, 38, 12, 5, 18, 6, 22, 18, 12, 6, 8]):
        ws_ap.column_dimensions[col].width = w

    # ═══ TAB 3: Active_Portfolio_50 ═══
    ws3 = wb.create_sheet("Active_Portfolio")
    ws3["A1"] = f"ACTIVE PORTFOLIO — {len(state['active_cadences'])} cadences in flight"
    ws3["A1"].font = title_font; ws3.merge_cells("A1:K1")
    h3 = ["#","Company","City","County","EE","Vertical","Score","Trigger","Start","Days In","Status"]
    for i, h in enumerate(h3, 1):
        c = ws3.cell(row=3, column=i, value=h); c.font = h_font; c.fill = h_fill
    ws3.freeze_panes = "A4"
    sorted_active = sorted(state["active_cadences"], key=lambda x: -x.get("score",0))
    for i, c in enumerate(sorted_active, 4):
        rank = i - 3
        start = dt.date.fromisoformat(c["start_date"])
        days_in = (TODAY - start).days
        ws3.cell(row=i, column=1, value=rank)
        ws3.cell(row=i, column=2, value=c["company"][:50])
        ws3.cell(row=i, column=3, value=c.get("city",""))
        ws3.cell(row=i, column=4, value=c.get("county",""))
        ws3.cell(row=i, column=5, value=c.get("ee"))
        ws3.cell(row=i, column=6, value=c.get("vertical",""))
        ws3.cell(row=i, column=7, value=round(c.get("score",0),1))
        ws3.cell(row=i, column=8, value=c.get("primary_trigger",""))
        ws3.cell(row=i, column=9, value=c["start_date"])
        ws3.cell(row=i, column=10, value=days_in)
        ws3.cell(row=i, column=11, value=c.get("status","active"))
    for col, w in zip("ABCDEFGHIJK", [4, 38, 16, 14, 6, 22, 7, 18, 11, 8, 18]):
        ws3.column_dimensions[col].width = w

    # ═══ TAB 4: Cadence_Tracker ═══
    ws4 = wb.create_sheet("Cadence_Tracker")
    ws4["A1"] = "CADENCE TRACKER — every active account's day-0/4/9/14 progress"
    ws4["A1"].font = title_font; ws4.merge_cells("A1:H1")
    h4 = ["Company","Trigger","Start","Today's Day","D0 (Email)","D4 (LinkedIn)","D9 (Drop)","D14 (Call)"]
    for i, h in enumerate(h4, 1):
        c = ws4.cell(row=3, column=i, value=h); c.font = h_font; c.fill = h_fill
    ws4.freeze_panes = "A4"
    for i, c in enumerate(sorted_active, 4):
        start = dt.date.fromisoformat(c["start_date"])
        days_in = (TODAY - start).days
        ws4.cell(row=i, column=1, value=c["company"][:40])
        ws4.cell(row=i, column=2, value=c.get("primary_trigger",""))
        ws4.cell(row=i, column=3, value=c["start_date"])
        ws4.cell(row=i, column=4, value=f"D{days_in}")
        for j, (cd_day, _, _, _) in enumerate(CADENCE, 5):
            mark = "✓" if days_in > cd_day else ("► TODAY" if days_in == cd_day else f"{(start + dt.timedelta(days=cd_day) - TODAY).days}d")
            ws4.cell(row=i, column=j, value=mark)
            if days_in == cd_day:
                ws4.cell(row=i, column=j).font = Font(bold=True, color="C00000")
    for col, w in zip("ABCDEFGH", [38, 18, 12, 11, 12, 14, 12, 12]):
        ws4.column_dimensions[col].width = w

    # ═══ TAB 5: Routes_Today ═══
    ws5 = wb.create_sheet("Routes_Today")
    today_drops = [a for a in actions if a["channel"] == "drop"]
    ws5["A1"] = f"TODAY'S DROP ROUTE — {len(today_drops)} stops in {route_label}"
    ws5["A1"].font = title_font; ws5.merge_cells("A1:E1")
    if today_drops:
        ws5["A2"] = "Pre-load all stops in Google Maps the prior afternoon. Stack 4-6 drops per route. Never drive 90 min for one drop."
        ws5["A2"].font = Font(italic=True, color="606060"); ws5.merge_cells("A2:E2")
        h5 = ["Stop","Company","City / ZIP","Trigger","30-Sec Drop Script (verbatim)"]
        for i, h in enumerate(h5, 1):
            c = ws5.cell(row=4, column=i, value=h); c.font = h_font; c.fill = h_fill
        for i, drop in enumerate(today_drops, 5):
            stop = i - 4
            ws5.cell(row=i, column=1, value=stop)
            ws5.cell(row=i, column=2, value=drop["company"])
            ws5.cell(row=i, column=3, value=drop["address"])
            ws5.cell(row=i, column=4, value=drop["trigger"])
            cell = ws5.cell(row=i, column=5, value=drop["body"])
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws5.row_dimensions[i].height = max(60, min(180, 20 + drop["body"].count("\n") * 16))
            for col in range(1, 6):
                ws5.cell(row=i, column=col).fill = fill_drop
    else:
        ws5.cell(row=3, column=1, value="No drops scheduled today.").font = Font(italic=True, color="606060")
        ws5.cell(row=4, column=1, value=f"Today is {weekday_name} ({route_label}).")
        ws5.cell(row=5, column=1, value="Drops fall on Day 9 of a cadence — wait for active cadences to mature, "
                                         "or seed more via --init.")
    for col, w in zip("ABCDE", [5, 38, 22, 18, 100]):
        ws5.column_dimensions[col].width = w

    # ═══ TAB 6: Operator_Dashboard ═══
    ws6 = wb.create_sheet("Operator_Dashboard")
    ws6["A1"] = f"OPERATOR DASHBOARD — Week of {(TODAY - dt.timedelta(days=WEEKDAY)).isoformat()}"
    ws6["A1"].font = title_font; ws6.merge_cells("A1:F1")
    ws6["A2"] = "Update Friday afternoon. Single number to optimize: STAGE-PROGRESSIONS-PER-WEEK (target 8-12)."
    ws6["A2"].font = Font(italic=True, color="606060"); ws6.merge_cells("A2:F2")

    ws6.cell(row=4, column=1, value="WEEKLY LEADING INDICATORS (inputs)").font = h_font
    ws6.cell(row=4, column=1).fill = h_fill
    ws6.merge_cells(start_row=4, start_column=1, end_row=4, end_column=5)
    li = [("Total touches", 50), ("Cold/trigger touches", 25), ("Warm touches (referral/CPA/broker)", 15),
          ("Field drops", 10), ("New triggered accounts identified", 5),
          ("★ Stage progressions (≥1 stage advance)", "8-12"),
          ("Discovery conversations", 3), ("Business Profiles collected", 1)]
    ws6.cell(row=5, column=1, value="Metric"); ws6.cell(row=5, column=2, value="Target")
    ws6.cell(row=5, column=3, value="This Wk"); ws6.cell(row=5, column=4, value="Last Wk"); ws6.cell(row=5, column=5, value="Trend")
    for c in "ABCDE": ws6[f"{c}5"].font = Font(bold=True); ws6[f"{c}5"].fill = sub_fill
    for i, (m, t) in enumerate(li, 6):
        ws6.cell(row=i, column=1, value=m); ws6.cell(row=i, column=2, value=t)
        if "Stage progressions" in m:
            ws6.cell(row=i, column=1).font = Font(bold=True, color="C00000")

    row = 6 + len(li) + 2
    ws6.cell(row=row, column=1, value="WEEKLY OUTCOME METRICS (outputs)").font = h_font
    ws6.cell(row=row, column=1).fill = h_fill
    ws6.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 1
    om = [("First meetings booked", 3), ("First meetings held (vs cancelled)", 2.5),
          ("Proposals/quotes delivered", 0.5), ("Closes signed", "—")]
    ws6.cell(row=row, column=1, value="Metric").font = Font(bold=True)
    ws6.cell(row=row, column=2, value="Target").font = Font(bold=True)
    ws6.cell(row=row, column=3, value="This Wk").font = Font(bold=True)
    for c_ in (1,2,3): ws6.cell(row=row, column=c_).fill = sub_fill
    row += 1
    for m, t in om:
        ws6.cell(row=row, column=1, value=m); ws6.cell(row=row, column=2, value=t); row += 1

    row += 2
    ws6.cell(row=row, column=1, value="THE SINGLE NUMBER TO OPTIMIZE").font = title_font
    ws6.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 1
    ws6.cell(row=row, column=1, value="Stage progressions per week").font = Font(bold=True, size=12)
    ws6.cell(row=row, column=2, value="Target 8-12").font = metric
    ws6.cell(row=row, column=3, value="—").font = metric

    for col, w in zip("ABCDE", [42, 14, 14, 14, 14]):
        ws6.column_dimensions[col].width = w

    # ═══ TAB 5b: Displacement_Pipeline (incumbent-PEO accounts in active cadences) ═══
    disp_active = [c for c in state["active_cadences"] if c.get("primary_trigger") == "displacement"]
    ws_disp = wb.create_sheet("Displacement_Pipeline")
    ws_disp["A1"] = f"DISPLACEMENT PIPELINE — {len(disp_active)} active incumbent-PEO conversions"
    ws_disp["A1"].font = title_font; ws_disp.merge_cells("A1:I1")
    ws_disp["A2"] = ("Per §1.4: incumbent-specific displacement angles. Insperity → UHC contract pressure. "
                     "TriNet → 12-month escape clause + per-employer underwriting. "
                     "Paychex PEO Direct → not a true master-pool. "
                     "Justworks → thin eastern-NC field service.")
    ws_disp["A2"].font = Font(italic=True, color="606060"); ws_disp.merge_cells("A2:I2")
    ws_disp.row_dimensions[2].height = 30
    ws_disp["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    h_disp = ["#","Company","City","County","EE","Incumbent","DM Name","DM Email","Start"]
    for i, h in enumerate(h_disp, 1):
        c = ws_disp.cell(row=4, column=i, value=h); c.font = h_font; c.fill = h_fill
    ws_disp.freeze_panes = "A5"
    sorted_disp = sorted(disp_active, key=lambda x: -x.get("score",0))
    for i, c in enumerate(sorted_disp, 5):
        rank = i - 4
        ws_disp.cell(row=i, column=1, value=rank)
        ws_disp.cell(row=i, column=2, value=c["company"][:42])
        ws_disp.cell(row=i, column=3, value=c.get("city",""))
        ws_disp.cell(row=i, column=4, value=c.get("county",""))
        ws_disp.cell(row=i, column=5, value=c.get("ee"))
        inc = incumbent_canonical(c.get("incumbent_peo",""))
        cell_inc = ws_disp.cell(row=i, column=6, value=inc or c.get("incumbent_peo","")[:14])
        # Color code incumbent
        inc_fills = {
            "Insperity": PatternFill("solid", fgColor="F4B084"),
            "TriNet":    PatternFill("solid", fgColor="DAE3F3"),
            "Paychex":   PatternFill("solid", fgColor="FFE699"),
            "Justworks": PatternFill("solid", fgColor="E2EFD9"),
        }
        if inc in inc_fills: cell_inc.fill = inc_fills[inc]
        ws_disp.cell(row=i, column=7, value=c.get("dm_name",""))
        ws_disp.cell(row=i, column=8, value=c.get("dm_email",""))
        ws_disp.cell(row=i, column=9, value=c["start_date"])
    if not sorted_disp:
        ws_disp.cell(row=5, column=1, value="No displacement cadences active. (Run --init to seed.)")
        ws_disp.cell(row=5, column=1).font = Font(italic=True, color="A06060")
        ws_disp.merge_cells("A5:I5")
    for col, w in zip("ABCDEFGHI", [4, 38, 16, 14, 6, 14, 22, 32, 12]):
        ws_disp.column_dimensions[col].width = w

    # ═══ TAB 5c: Buyer_Cast (track Owner / CFO / Office_Mom / BROKER / CPA / Attorney per account) ═══
    ws_cast = wb.create_sheet("Buyer_Cast")
    ws_cast["A1"] = "BUYER CAST — broker is the unknown deal-killer. CAPTURE on every drop."
    ws_cast["A1"].font = title_font; ws_cast.merge_cells("A1:H1")
    ws_cast["A2"] = ("Per §1.5: Owner signs, but CFO and INSURANCE BROKER can kill. Every drop must capture broker. "
                     "Enter what you learn here → flows back into the system.")
    ws_cast["A2"].font = Font(italic=True, color="606060"); ws_cast.merge_cells("A2:H2")
    h_cast = ["Company","Owner / CEO","CFO","Office Mom (HR)","★ Broker","CPA","Attorney","Trigger / Incumbent"]
    for i, h in enumerate(h_cast, 1):
        c = ws_cast.cell(row=4, column=i, value=h); c.font = h_font; c.fill = h_fill
    ws_cast.freeze_panes = "A5"
    fill_broker_unknown = PatternFill("solid", fgColor="FFCCCC")
    for i, c in enumerate(sorted(state["active_cadences"], key=lambda x: -x.get("score",0)), 5):
        cast = state["buyer_cast"].get(c["key"], {})
        ws_cast.cell(row=i, column=1, value=c["company"][:38])
        ws_cast.cell(row=i, column=2, value=cast.get("owner",""))
        ws_cast.cell(row=i, column=3, value=cast.get("cfo",""))
        ws_cast.cell(row=i, column=4, value=cast.get("office_mom",""))
        broker = cast.get("broker","")
        bc = ws_cast.cell(row=i, column=5, value=broker or "(unknown — capture on next drop)")
        if not broker:
            bc.fill = fill_broker_unknown
            bc.font = Font(italic=True, color="A04040")
        ws_cast.cell(row=i, column=6, value=cast.get("cpa",""))
        ws_cast.cell(row=i, column=7, value=cast.get("attorney",""))
        trig_label = c.get("primary_trigger","")
        if trig_label == "displacement":
            trig_label = f"displ → {incumbent_canonical(c.get('incumbent_peo',''))}"
        ws_cast.cell(row=i, column=8, value=trig_label)
    for col, w in zip("ABCDEFGH", [34, 22, 22, 22, 32, 22, 22, 22]):
        ws_cast.column_dimensions[col].width = w

    # ═══ TAB 5d: Discovery_Flight_Plan (PRINT THIS — carry to every 20-min discovery) ═══
    ws_disc = wb.create_sheet("Discovery_Flight_Plan")
    ws_disc["A1"] = "DISCOVERY FLIGHT PLAN — print, carry to every 20-min meeting"
    ws_disc["A1"].font = title_font; ws_disc.merge_cells("A1:C1")
    ws_disc["A2"] = ("Per playbook §1.3 + §2.3. Question stack + 5 Whys + Sandler pain funnel + "
                     "Challenger TCO reframe + Gap Selling Current/Future/Impact. "
                     "Goal: 4–6 named pains, an economic-buyer commitment, and a next step.")
    ws_disc["A2"].font = Font(italic=True, color="606060"); ws_disc.merge_cells("A2:C2")
    ws_disc.row_dimensions[2].height = 32
    ws_disc["A2"].alignment = Alignment(wrap_text=True, vertical="top")

    sections = [
        ("STAGE 1 — OPENING (2 min)", [
            "\"Thanks for the time. I have 20 minutes blocked. Goal of the call: I learn how you run HR / payroll / benefits today and you walk away with at least one specific number you didn't have when we started. Fair?\"",
            "Set the up-front contract (Sandler): mutual permission to say 'no' and to give a definitive next step.",
        ]),
        ("STAGE 2 — CURRENT STATE (Gap Selling, 5 min)", [
            "Walk me through how you run payroll today. Who does it? How long does it take?",
            "What does your benefits package look like today? Who's the broker? When's renewal?",
            "If you had a NCDOL or DOL audit Monday morning — how prepared would you say you are, on a 1–10?",
            "What's your workers' comp setup — carrier, mod, when's renewal?",
            "When something employment-law-flavored comes up (termination, harassment claim, leave), where do you turn?",
        ]),
        ("STAGE 3 — FUTURE STATE + IMPACT (5 Whys, 5 min)", [
            "If those 5 things ran themselves — what would change for you personally? For the business?",
            "If we don't fix [pain #1] in the next 12 months, what does it cost you? (5 Whys: keep going till the answer is dollars or hours or owner-time.)",
            "If you had a 1–10 confidence number on [topic] going into this call, where does it need to be 12 months from now?",
            "Whose attention does this currently steal? (Office Mom? CFO? You?)",
        ]),
        ("STAGE 4 — TCO REFRAME (Challenger, 3 min)", [
            "\"Most owners I work with at your size think their HR cost is their HR person's salary. The actual True Cost of In-House HR — when you add WC mod overpayment, broker spread on benefits, owner-time on compliance, and turnover from a sub-F500 benefits package — runs 8–14% of payroll for an 11–50 EE company. The TotalSource structure typically lands at 6–9%. That's the conversation worth having.\"",
            "Challenger insight #2: \"Carriers shopping the small-group market are quietly thinning networks AND raising rates. Your renewal is going to look better than it actually is — the network exclusion is buried.\"",
            "Challenger insight #3: \"The biggest hidden cost isn't the renewal — it's the 1.5–2× turnover penalty when your benefits package can't compete with the F500 anchor your candidate is also interviewing with.\"",
        ]),
        ("STAGE 5 — DECISION CRITERIA + PROCESS (MEDDPICC, 3 min)", [
            "\"If we do another conversation after this one, who else needs to be in the room?\" → Economic Buyer.",
            "\"What's the bar a solution would have to clear for you to say yes?\" → Decision Criteria.",
            "\"How do decisions like this get made here? Renewal-driven? Budget-cycle?\" → Decision Process.",
            "\"Who's the contract — owner sign-off, or board? Any procurement / legal review?\" → Paper Process.",
            "\"What's your current-state fix — broker quote shopping? Stay put? Switch to a competitor?\" → Competition.",
        ]),
        ("STAGE 6 — CLOSE TO NEXT STEP (2 min)", [
            "Specific next step (NEVER 'let me think about it' as the close): \"Tuesday at 10am, I'll come back with [the WC mod benchmark / health renewal side-by-side / TCO model] and your CFO/broker can be in the room. Yes or no?\"",
            "If yes: confirm at end of call + send calendar invite within 1 hour.",
            "If no: Sandler funnel one more time — \"What would have to be true for that to be a yes?\"",
        ]),
        ("SANDLER PAIN FUNNEL (deploy any time pain surfaces)", [
            "Tell me more about that.",
            "Can you be more specific? Give me an example.",
            "How long has this been a problem?",
            "What have you tried to do about it?",
            "Did it work?",
            "How much do you think this has cost you?",
            "How do you feel about that?",
            "Have you given up trying to fix this on your own?",
        ]),
    ]
    row = 4
    for sec_title, items in sections:
        c = ws_disc.cell(row=row, column=1, value=sec_title)
        c.font = h_font; c.fill = h_fill
        ws_disc.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1
        for q in items:
            cell = ws_disc.cell(row=row, column=1, value="  " + q)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.font = Font(size=11)
            ws_disc.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            ws_disc.row_dimensions[row].height = 32
            row += 1
        row += 1
    for col, w in zip("ABC", [44, 44, 44]):
        ws_disc.column_dimensions[col].width = w

    # ═══ TAB 5e: MEDDPICC_Pipeline (deals past D14 → engaged or beyond) ═══
    ws_med = wb.create_sheet("MEDDPICC_Pipeline")
    ws_med["A1"] = "MEDDPICC PIPELINE — every account that's had a meeting booked or held"
    ws_med["A1"].font = title_font; ws_med.merge_cells("A1:K1")
    ws_med["A2"] = ("Per §2.3. Score 0/1/2 in each column. ≥10 = forecast-able. <6 = needs work. "
                    "Updated automatically when an outcome of 'meeting_booked' or 'meeting_held' "
                    "is logged in Weekly_Wrap.")
    ws_med["A2"].font = Font(italic=True, color="606060"); ws_med.merge_cells("A2:K2")
    h_med = ["Company","First mtg","Stage","M Metrics","E Econ Buyer","D1 Criteria","D2 Process","P Paper","I Pain","C Champion","Cmp Competition"]
    for i, h in enumerate(h_med, 1):
        c = ws_med.cell(row=4, column=i, value=h); c.font = h_font; c.fill = h_fill
    ws_med.freeze_panes = "A5"
    meddpicc_rows = list(state.get("meddpicc", {}).items())
    if not meddpicc_rows:
        ws_med.cell(row=5, column=1, value="No meetings booked yet. (When you log meeting_booked or meeting_held in Weekly_Wrap, the account appears here automatically.)")
        ws_med.cell(row=5, column=1).font = Font(italic=True, color="A06060")
        ws_med.merge_cells("A5:K5")
    else:
        for i, (key, rec) in enumerate(meddpicc_rows, 5):
            ws_med.cell(row=i, column=1, value=rec.get("company","")[:36])
            ws_med.cell(row=i, column=2, value=rec.get("first_meeting_date",""))
            ws_med.cell(row=i, column=3, value=rec.get("stage",""))
            for j, fld in enumerate(["M_metrics","E_econ_buyer","D1_decision_criteria",
                                     "D2_decision_process","P_paper_process","I_pain",
                                     "C_champion","Cmp_competition"], 4):
                ws_med.cell(row=i, column=j, value=rec.get(fld,""))
    for col, w in zip("ABCDEFGHIJK", [32, 12, 14, 26, 22, 22, 22, 22, 22, 22, 22]):
        ws_med.column_dimensions[col].width = w

    # ═══ TAB 5f: Channel_Build_Program (recruit a partner channel from zero) ═══
    ws_chan = wb.create_sheet("Channel_Build_Program")
    ws_chan["A1"] = "CHANNEL BUILD — recruit 25 CPAs / 10 brokers / 4 attorneys (you have 0 today)"
    ws_chan["A1"].font = title_font; ws_chan.merge_cells("A1:G1")
    ws_chan["A2"] = (
        "Per §1.5 + §2.6: brokers and CPAs are the silent deal-killer or accelerator. "
        "Brokers auto-populate when you capture them on drops (column G of Weekly_Wrap). "
        "Quarterly compliance brief = 1-page recap of NC enforcement actions / new law / IRS rulings, "
        "sent to every contact every 90 days. Reciprocity compounds at month 6."
    )
    ws_chan["A2"].font = Font(italic=True, color="606060")
    ws_chan.merge_cells("A2:G2")
    ws_chan.row_dimensions[2].height = 36
    ws_chan["A2"].alignment = Alignment(wrap_text=True, vertical="top")

    # ── Targets summary ──
    ws_chan["A4"] = "TARGETS vs CURRENT (build to these counts over months 1–12)"
    ws_chan["A4"].font = h_font; ws_chan["A4"].fill = h_fill
    ws_chan.merge_cells("A4:G4")
    counts = [
        ("Brokers (insurance — primary deal-killer)", 10, len(state["channel_build"]["brokers"])),
        ("CPAs (small-biz tax — referral source)",     25, len(state["channel_build"]["cpas"])),
        ("Employment attorneys",                        4, len(state["channel_build"]["attorneys"])),
    ]
    for i, (label, target, current) in enumerate(counts, 5):
        ws_chan.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws_chan.cell(row=i, column=2, value="Target:")
        ws_chan.cell(row=i, column=3, value=target).font = metric
        ws_chan.cell(row=i, column=4, value="Current:")
        cell_cur = ws_chan.cell(row=i, column=5, value=current); cell_cur.font = metric
        if current == 0:
            cell_cur.fill = PatternFill("solid", fgColor="FFCCCC")
        elif current >= target:
            cell_cur.fill = PatternFill("solid", fgColor="C6EFCE")
        else:
            cell_cur.fill = PatternFill("solid", fgColor="FFEB9C")
        ws_chan.cell(row=i, column=6, value=f"{current/target:.0%}" if target else "")

    # ── Recruitment monthly quota ──
    ws_chan["A9"] = "MONTHLY QUOTA (per playbook §2.6)"
    ws_chan["A9"].font = h_font; ws_chan["A9"].fill = h_fill
    ws_chan.merge_cells("A9:G9")
    quotas = [
        "2 CPA lunches / month — walk in with a 1-pager: \"how a TotalSource referral works for your client.\"",
        "1 broker lunch / month — when you discover a friendly broker, lunch is the qualifier.",
        "1 employment attorney lunch / quarter — pick from the NC State Bar L&E section roster.",
        "Quarterly compliance brief drop to every contact (NCDOL enforcement, ACA deadlines, new law).",
        "When you close a CPA-/broker-sourced deal: introduce that partner to your other clients (Givers Gain).",
    ]
    for i, q in enumerate(quotas, 10):
        cell = ws_chan.cell(row=i, column=1, value="  • " + q)
        cell.alignment = Alignment(wrap_text=True, vertical="top"); cell.font = Font(size=11)
        ws_chan.merge_cells(start_row=i, start_column=1, end_row=i, end_column=7)
        ws_chan.row_dimensions[i].height = 28

    # ── Brokers roster (auto-populated from drops) ──
    row = 16
    ws_chan.cell(row=row, column=1, value="BROKERS (auto-populated as you capture them on drops)").font = h_font
    ws_chan.cell(row=row, column=1).fill = h_fill
    ws_chan.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    row += 1
    h_brk = ["Name", "First seen via", "County", "Phone", "Email", "# clients", "Last brief sent"]
    for i, h in enumerate(h_brk, 1):
        c = ws_chan.cell(row=row, column=i, value=h); c.font = Font(bold=True); c.fill = sub_fill
    row += 1
    if not state["channel_build"]["brokers"]:
        ws_chan.cell(row=row, column=1, value="(none yet — capture on every drop in Weekly_Wrap col G)").font = Font(italic=True, color="A06060")
        ws_chan.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        row += 1
    else:
        for b in state["channel_build"]["brokers"]:
            ws_chan.cell(row=row, column=1, value=b.get("name","")[:30])
            ws_chan.cell(row=row, column=2, value=b.get("first_seen_via","")[:30])
            ws_chan.cell(row=row, column=3, value=b.get("county",""))
            ws_chan.cell(row=row, column=4, value=b.get("phone",""))
            ws_chan.cell(row=row, column=5, value=b.get("email",""))
            ws_chan.cell(row=row, column=6, value=b.get("n_clients_observed", 1))
            ws_chan.cell(row=row, column=7, value=b.get("last_brief_sent",""))
            row += 1

    # ── CPA recruitment seed list (pull from territory CRM ICP-PASS verticals) ──
    row += 1
    ws_chan.cell(row=row, column=1, value="CPA RECRUITMENT TARGETS (manual fill — NC CPA Board search by county)").font = h_font
    ws_chan.cell(row=row, column=1).fill = h_fill
    ws_chan.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    row += 1
    h_cpa = ["Firm name", "Lead CPA", "County", "Phone", "Email", "Lunch status", "Last brief sent"]
    for i, h in enumerate(h_cpa, 1):
        c = ws_chan.cell(row=row, column=i, value=h); c.font = Font(bold=True); c.fill = sub_fill
    row += 1
    # Seed 8 empty rows for manual fill
    if not state["channel_build"]["cpas"]:
        for _ in range(8):
            for col in range(1, 8):
                ws_chan.cell(row=row, column=col, value="")
                ws_chan.cell(row=row, column=col).border = Border(
                    bottom=Side(style="thin", color="C0C0C0"),
                )
            row += 1
    else:
        for cpa in state["channel_build"]["cpas"]:
            ws_chan.cell(row=row, column=1, value=cpa.get("firm",""))
            ws_chan.cell(row=row, column=2, value=cpa.get("lead",""))
            ws_chan.cell(row=row, column=3, value=cpa.get("county",""))
            ws_chan.cell(row=row, column=4, value=cpa.get("phone",""))
            ws_chan.cell(row=row, column=5, value=cpa.get("email",""))
            ws_chan.cell(row=row, column=6, value=cpa.get("lunch_status",""))
            ws_chan.cell(row=row, column=7, value=cpa.get("last_brief_sent",""))
            row += 1

    # ── Attorney recruitment (4 slots) ──
    row += 1
    ws_chan.cell(row=row, column=1, value="EMPLOYMENT ATTORNEY TARGETS (NC State Bar L&E section by county — 4 slots)").font = h_font
    ws_chan.cell(row=row, column=1).fill = h_fill
    ws_chan.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    row += 1
    h_atty = ["Firm", "Attorney name", "County", "Phone", "Email", "Quarterly lunch?", "Notes"]
    for i, h in enumerate(h_atty, 1):
        c = ws_chan.cell(row=row, column=i, value=h); c.font = Font(bold=True); c.fill = sub_fill
    row += 1
    if not state["channel_build"]["attorneys"]:
        for _ in range(4):
            for col in range(1, 8):
                ws_chan.cell(row=row, column=col, value="")
            row += 1

    for col, w in zip("ABCDEFG", [40, 28, 16, 16, 28, 14, 16]):
        ws_chan.column_dimensions[col].width = w

    # ═══ TAB 5g: Industry_Trends (weekly per-vertical rollup — what's converting, top carriers, top incumbents) ═══
    ws_ind = wb.create_sheet("Industry_Trends")
    ws_ind["A1"] = f"INDUSTRY TRENDS — week of {(TODAY - dt.timedelta(days=TODAY.weekday())).isoformat()}    |    per-vertical rollup"
    ws_ind["A1"].font = title_font; ws_ind.merge_cells("A1:J1")
    ws_ind["A2"] = (
        "Per-industry view: active cadences, mean score, last-week conversion, "
        "top WC carrier, top health carrier, top incumbent PEO. "
        "Use this to pick which industry to lean into when seeding new cadences."
    )
    ws_ind["A2"].font = Font(italic=True, color="606060")
    ws_ind.merge_cells("A2:J2")
    ws_ind.row_dimensions[2].height = 32
    ws_ind["A2"].alignment = Alignment(wrap_text=True, vertical="top")

    # ── Build per-vertical rollup from active state + ledger ──
    industry_carrier = _industry_carrier_rollup_cached()
    ledger = state.get("weekly_outcomes", [])
    week_start = TODAY - dt.timedelta(days=TODAY.weekday())
    last_week_start = week_start - dt.timedelta(days=7)
    last_week_outcomes = [o for o in ledger
                          if o.get("logged_at","") >= last_week_start.isoformat()
                          and o.get("logged_at","") < week_start.isoformat()]

    # Normalize each active cadence's vertical/name into a coarse industry bucket
    def _bucket_for(c):
        name = (c.get("company") or "").lower()
        vert = (c.get("vertical") or "").lower()
        if "engineering" in vert or "engineer" in name or "engineers" in name: return "Engineering"
        if "manufactur" in vert or "mfg" in name or "machining" in name or "packaging" in name or "fabric" in name: return "Manufacturing"
        if "hvac" in name or "air condition" in name or "heating" in name or "air purif" in name or "a/c " in name: return "HVAC"
        if "electric" in name: return "Electrical"
        if "tech" in vert or "saas" in vert or "tech" in name or "software" in name or "data solutions" in name or "digital" in name: return "Tech / IT"
        if "consulting" in vert or "consult" in name: return "Mgmt Consulting"
        if "hatfield" in name or "law" in name or "& wall" in name or "purks" in name: return "Legal / Prof Svcs"
        if "medical" in name or "biotech" in name or "labs" in name or "pharma" in name: return "Biotech / Medical Device"
        if "creative" in name or "marketing" in name: return "Creative / Marketing"
        return "Other / Unclassified"

    by_industry = defaultdict(list)
    for c in state["active_cadences"]:
        by_industry[_bucket_for(c)].append(c)

    # Conversion stats per industry from ledger
    def _industry_for_ledger(o):
        # Reconstruct bucket from the ledger row's vertical + company
        fake = {"vertical": o.get("vertical",""), "company": o.get("company","")}
        return _bucket_for(fake)
    ledger_by_industry = defaultdict(list)
    for o in last_week_outcomes:
        ledger_by_industry[_industry_for_ledger(o)].append(o)

    # ── Headers ──
    h_ind = ["Industry","# Active","Median Score","Touches LW","Progress LW","Conv % LW",
             "Top WC Carrier","Top Health Carrier","Top Incumbent PEO","Trend"]
    for i, h in enumerate(h_ind, 1):
        c = ws_ind.cell(row=4, column=i, value=h); c.font = h_font; c.fill = h_fill
    ws_ind.freeze_panes = "A5"

    # Sort industries by # active desc
    sorted_inds = sorted(by_industry.items(), key=lambda x: -len(x[1]))
    row = 5
    for industry, cs in sorted_inds:
        n_active = len(cs)
        scores = [c.get("score", 0) for c in cs]
        median_score = sorted(scores)[len(scores)//2] if scores else 0
        lw = ledger_by_industry.get(industry, [])
        n_touches_lw = len(lw)
        n_progress_lw = sum(1 for o in lw if o.get("outcome") in PROGRESSION_OUTCOMES)
        conv_lw = (n_progress_lw / n_touches_lw) if n_touches_lw else None
        # Carrier / incumbent rollup from sqlite
        carrier_data = industry_carrier.get(industry, {})
        top_wc = ", ".join(f"{n}({k})" for n,k in [(c[0], c[1]) for c in carrier_data.get("wc_top", {}).most_common(2)]) if carrier_data.get("wc_top") else "—"
        top_hl = ", ".join(f"{n}({k})" for n,k in [(c[0], c[1]) for c in carrier_data.get("health_top", {}).most_common(2)]) if carrier_data.get("health_top") else "—"
        top_peo = ", ".join(f"{n}({k})" for n,k in [(c[0], c[1]) for c in carrier_data.get("peo_top", {}).most_common(2)]) if carrier_data.get("peo_top") else "—"

        # Trend arrow: compare conv_lw to ledger 2-weeks-ago for same industry
        prior_start = last_week_start - dt.timedelta(days=7)
        prior_outcomes = [o for o in ledger
                          if o.get("logged_at","") >= prior_start.isoformat()
                          and o.get("logged_at","") < last_week_start.isoformat()
                          and _industry_for_ledger(o) == industry]
        prior_conv = (sum(1 for o in prior_outcomes if o.get("outcome") in PROGRESSION_OUTCOMES) /
                      len(prior_outcomes)) if prior_outcomes else None
        if conv_lw is not None and prior_conv is not None:
            if conv_lw > prior_conv * 1.1: trend = "↑"
            elif conv_lw < prior_conv * 0.9: trend = "↓"
            else: trend = "→"
        elif conv_lw is not None and prior_conv is None: trend = "(new)"
        else: trend = "—"

        ws_ind.cell(row=row, column=1, value=industry).font = Font(bold=True)
        ws_ind.cell(row=row, column=2, value=n_active)
        ws_ind.cell(row=row, column=3, value=round(median_score, 1))
        ws_ind.cell(row=row, column=4, value=n_touches_lw)
        ws_ind.cell(row=row, column=5, value=n_progress_lw)
        ws_ind.cell(row=row, column=6, value=f"{conv_lw:.0%}" if conv_lw is not None else "—")
        ws_ind.cell(row=row, column=7, value=top_wc[:60])
        ws_ind.cell(row=row, column=8, value=top_hl[:60])
        ws_ind.cell(row=row, column=9, value=top_peo[:50])
        ws_ind.cell(row=row, column=10, value=trend)
        # Highlight focus industries
        if industry in {"Manufacturing","HVAC","Electrical","Engineering","Tech / IT",
                        "Mgmt Consulting","Legal / Prof Svcs"}:
            for col in range(1, 11):
                ws_ind.cell(row=row, column=col).fill = fill_linkedin   # green-ish
        elif industry == "Other / Unclassified":
            for col in range(1, 11):
                ws_ind.cell(row=row, column=col).fill = sub_fill
        row += 1

    # ── Macro context block (territory-wide trends from pipeline.sqlite) ──
    row += 2
    ws_ind.cell(row=row, column=1, value="TERRITORY-WIDE MACRO (from pipeline.sqlite)").font = h_font
    ws_ind.cell(row=row, column=1).fill = h_fill
    ws_ind.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    row += 1

    db = sqlite3.connect(str(DB))
    macro_rows = []
    # WC renewals firing in next 90 days by industry
    wc_by_ind = Counter()
    for cid, val in db.execute(
        "SELECT company_id, value FROM enrichments WHERE field='wc_expiration_date'"
    ).fetchall():
        d = parse_date(val) if 'parse_date' in dir() else None
        # quick parse:
        try:
            from datetime import datetime as _dt
            for fmt in ("%Y-%m-%d","%m/%d/%Y","%Y-%m-%dT%H:%M:%SZ"):
                try: d = _dt.strptime(val, fmt).date(); break
                except Exception: d = None
        except Exception: d = None
        if d and 0 <= (d - TODAY).days <= 90:
            v = ""
            r = db.execute("SELECT value FROM enrichments WHERE company_id=? AND field='naics_inferred' LIMIT 1",(cid,)).fetchone()
            if r: v = _naics_to_vertical(r[0] or "")
            if v: wc_by_ind[v] += 1
    macro_rows.append(("WC renewals firing in next 90 days (territory)",
                       ", ".join(f"{k}: {n}" for k, n in wc_by_ind.most_common(8))))

    # Health carrier concentration
    hc = Counter()
    for r in db.execute("SELECT carrier_name FROM carriers").fetchall():
        if r[0]: hc[r[0]] += 1
    macro_rows.append(("Top 5 health carriers in territory (PEO master plan should beat these)",
                       ", ".join(f"{k}: {n}" for k, n in hc.most_common(5))))

    # Incumbent PEO concentration (excludes ADP's own — those are won territory, not targets)
    pc = Counter()
    adp_own = 0
    for r in db.execute("SELECT value FROM enrichments WHERE field='competing_peo_brand'").fetchall():
        if not r[0]: continue
        if "adp" in r[0].lower() or "totalsource" in r[0].lower():
            adp_own += 1
            continue
        pc[r[0]] += 1
    macro_rows.append(("Top incumbent PEOs in territory (displacement targets, ex-ADP)",
                       ", ".join(f"{k}: {n}" for k, n in pc.most_common(8))))
    macro_rows.append(("Territory ADP TotalSource penetration (already-won, NOT targets)",
                       f"{adp_own} accounts identified as current ADP customers"))
    db.close()

    for label, val in macro_rows:
        ws_ind.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws_ind.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        ws_ind.cell(row=row, column=4, value=val)
        ws_ind.merge_cells(start_row=row, start_column=4, end_row=row, end_column=10)
        ws_ind.cell(row=row, column=4).alignment = Alignment(wrap_text=True, vertical="top")
        ws_ind.row_dimensions[row].height = 28
        row += 1

    for col, w in zip("ABCDEFGHIJ", [24, 9, 12, 11, 12, 10, 32, 32, 28, 8]):
        ws_ind.column_dimensions[col].width = w

    # ═══ TAB 6a: Coaching_Card (MONDAY READ — adjustments from last week's data) ═══
    ws_coach = wb.create_sheet("Coaching_Card")
    ws_coach["A1"] = f"COACHING CARD — Monday {TODAY_ISO}    |    what last week's data is telling you"
    ws_coach["A1"].font = title_font; ws_coach.merge_cells("A1:F1")

    ledger = state.get("weekly_outcomes", [])
    week_start = TODAY - dt.timedelta(days=TODAY.weekday())
    last_week_start = week_start - dt.timedelta(days=7)
    last_week_end = week_start - dt.timedelta(days=1)
    last_week_outcomes = [o for o in ledger
                          if o.get("logged_at","") >= last_week_start.isoformat()
                          and o.get("logged_at","") <= last_week_end.isoformat()]

    n_touches_lw = len(last_week_outcomes)
    n_progressions_lw = sum(1 for o in last_week_outcomes if o["outcome"] in PROGRESSION_OUTCOMES)
    n_meetings_booked_lw = sum(1 for o in last_week_outcomes if o["outcome"] == "meeting_booked")
    n_meetings_held_lw = sum(1 for o in last_week_outcomes if o["outcome"] == "meeting_held")
    n_dead_lw = sum(1 for o in last_week_outcomes if o["outcome"] in KILL_OUTCOMES)
    n_broker_captured_lw = sum(1 for o in last_week_outcomes if o.get("broker_captured"))
    n_drops_lw = sum(1 for o in last_week_outcomes if o["channel"] == "drop")
    broker_rate = (n_broker_captured_lw / n_drops_lw) if n_drops_lw > 0 else None

    ws_coach["A3"] = "LAST WEEK SCOREBOARD"
    ws_coach["A3"].font = h_font; ws_coach["A3"].fill = h_fill
    ws_coach.merge_cells("A3:F3")
    metrics = [
        ("Touches logged",         n_touches_lw,         "target 50"),
        ("Stage progressions",     n_progressions_lw,    "target 8–12 (THE single number)"),
        ("Meetings booked",        n_meetings_booked_lw, "target 3"),
        ("Meetings held",          n_meetings_held_lw,   "target 2 (≤ booked, cancellation gap)"),
        ("Killed (DEAD/DNC/etc)",  n_dead_lw,            "fed back to pipeline.sqlite drops"),
        ("Broker discovery rate",
            f"{broker_rate:.0%}" if broker_rate is not None else "n/a",
            "drops only — every drop should capture broker"),
    ]
    for i, (label, val, note) in enumerate(metrics, 4):
        ws_coach.cell(row=i, column=1, value=label).font = Font(bold=True)
        c = ws_coach.cell(row=i, column=2, value=val); c.font = metric
        ws_coach.cell(row=i, column=3, value=note).font = Font(italic=True, color="606060")

    # ── What's converting / what's not ──
    weights = state.get("weights", {})
    def _sorted_weights(d, only_seen_in_ledger=None):
        items = list(d.items())
        if only_seen_in_ledger is not None:
            items = [(k, v) for k, v in items if k in only_seen_in_ledger]
        items.sort(key=lambda x: -x[1])
        return items

    ws_coach.cell(row=11, column=1, value="WHAT'S CONVERTING (weight > 1.0 = do MORE)").font = h_font
    ws_coach.cell(row=11, column=1).fill = h_fill
    ws_coach.merge_cells("A11:F11")
    triggers_seen = {o["trigger"] for o in ledger if o.get("trigger")}
    verticals_seen = {o["vertical"] for o in ledger if o.get("vertical")}
    sections = [
        ("Trigger",   _sorted_weights(weights.get("trigger", {}), triggers_seen)),
        ("Vertical",  _sorted_weights(weights.get("vertical", {}), verticals_seen)),
        ("Channel",   _sorted_weights(weights.get("channel", {}))),
        ("Route day", [(["Mon Wake","Tue Pitt","Wed Northern","Thu Cumberland"][int(k)], v)
                       for k, v in _sorted_weights(weights.get("route_day", {}))]),
    ]
    row = 12
    for label, items in sections:
        ws_coach.cell(row=row, column=1, value=label).font = Font(bold=True)
        if not items:
            ws_coach.cell(row=row, column=2, value="— not enough data yet —").font = Font(italic=True, color="A06060")
            row += 1
            continue
        for k, v in items[:6]:
            ws_coach.cell(row=row, column=2, value=str(k)[:30])
            wcell = ws_coach.cell(row=row, column=3, value=f"{v:.2f}×")
            if v >= 1.3: wcell.fill = fill_linkedin
            elif v <= 0.7: wcell.fill = fill_drop
            row += 1
        row += 1

    # ── 3 prescribed adjustments ──
    ws_coach.cell(row=row, column=1, value="THIS WEEK'S 3 ADJUSTMENTS").font = h_font
    ws_coach.cell(row=row, column=1).fill = h_fill
    ws_coach.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1
    adjustments = []
    if weights.get("last_recomputed"):
        # Always emit a top-1 / bottom-1 recommendation per dimension that has data
        ledger_triggers = sorted({o["trigger"] for o in ledger if o.get("trigger")})
        seen_trigger_w = [(k, weights["trigger"].get(k, 1.0)) for k in ledger_triggers]
        if len(seen_trigger_w) >= 2:
            best = max(seen_trigger_w, key=lambda x: x[1]); worst = min(seen_trigger_w, key=lambda x: x[1])
            if best[1] >= 1.05:
                adjustments.append(f"Lean into {best[0]} ({best[1]:.2f}×) — your highest-converting trigger so far. Replacements should bias here.")
            if worst[1] <= 0.95 and worst[0] != best[0]:
                adjustments.append(f"Audit {worst[0]} cadences ({worst[1]:.2f}×) — converting below your average. Tighten the opener or rotate down.")
        all_ch_w = [(k, v) for k, v in weights.get("channel", {}).items()]
        if all_ch_w:
            ch_best = max(all_ch_w, key=lambda x: x[1]); ch_worst = min(all_ch_w, key=lambda x: x[1])
            if ch_best[1] >= 1.05:
                adjustments.append(f"Your {ch_best[0]} channel is at {ch_best[1]:.2f}× — load this channel as the anchor touch on this week's new cadences.")
            if ch_worst[1] <= 0.95 and ch_worst[0] != ch_best[0]:
                adjustments.append(f"Re-look the {ch_worst[0]} copy ({ch_worst[1]:.2f}×) — the message isn't landing. Try a sharper trigger anchor.")
        all_rd_w = [(k, v) for k, v in weights.get("route_day", {}).items()]
        if all_rd_w:
            day_names = {"0":"Mon Wake", "1":"Tue Pitt", "2":"Wed Northern", "3":"Thu Cumberland"}
            rd_best = max(all_rd_w, key=lambda x: x[1])
            if rd_best[1] >= 1.05:
                adjustments.append(f"Route day {day_names.get(rd_best[0], rd_best[0])} is hitting {rd_best[1]:.2f}× — protect that block, no calendar conflicts.")
        if broker_rate is not None:
            if broker_rate < 0.5:
                adjustments.append(f"Broker discovery rate is {broker_rate:.0%} on drops — every drop MUST capture broker. The gatekeeper question: 'Who's your workers' comp through right now?'")
            elif broker_rate >= 0.7:
                adjustments.append(f"Broker discovery rate {broker_rate:.0%} — solid. Now use that intel: when you know the broker, your displacement angle changes.")
        cancel_rate = (sum(1 for o in last_week_outcomes if o["outcome"] in {"meeting_cancelled","meeting_no_show"})
                       / max(1, n_meetings_booked_lw + sum(1 for o in last_week_outcomes if o["outcome"] in {"meeting_cancelled","meeting_no_show"})))
        if n_meetings_booked_lw + n_meetings_held_lw > 0 and cancel_rate > 0.25:
            adjustments.append(f"Cancellation rate {cancel_rate:.0%} > 25% — confirm every meeting 24 hr before with reference to prep work.")
        if n_progressions_lw > 0 and n_progressions_lw < 8:
            adjustments.append(f"Stage progressions {n_progressions_lw} below target 8–12. Push 5–8 more touches per day this week.")
        elif n_progressions_lw >= 12:
            adjustments.append(f"Stage progressions {n_progressions_lw} at/above target — protect the cadence discipline. Don't slack on D14 calls.")
    if not adjustments:
        adjustments = [
            "Need ≥10 logged outcomes before learning engine kicks in. Run cadences, fill Weekly_Wrap Friday, re-run Sun night.",
            "Every drop must capture: gatekeeper name, owner schedule, current HR/payroll, and WC carrier (the 4 critical fields).",
            "On day-14 calls: ask for a 20-minute meeting with a specific outcome (mod-rate review, benefits comp, etc) — never an open-ended 'chat.'",
        ]
    for a in adjustments[:5]:
        ws_coach.cell(row=row, column=1, value=f"  • {a}").font = Font(size=11)
        ws_coach.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws_coach.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws_coach.row_dimensions[row].height = 30
        row += 1

    # Footer note
    row += 1
    last_recompute = weights.get("last_recomputed") or "never"
    ws_coach.cell(row=row, column=1, value=f"Weights last recomputed: {last_recompute} on {weights.get('n_outcomes_at_recompute',0)} outcomes.")
    ws_coach.cell(row=row, column=1).font = Font(italic=True, color="606060", size=9)
    ws_coach.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)

    for col, w in zip("ABCDEF", [34, 18, 11, 38, 12, 12]):
        ws_coach.column_dimensions[col].width = w

    # ═══ TAB 6b: Weekly_Wrap (FRIDAY INPUT — rep fills outcomes here) ═══
    from openpyxl.worksheet.datavalidation import DataValidation
    ws_wrap = wb.create_sheet("Weekly_Wrap")
    week_end = TODAY
    week_start = week_end - dt.timedelta(days=6)
    ws_wrap["A1"] = f"WEEKLY WRAP — {week_start.isoformat()} → {week_end.isoformat()}    |    fill OUTCOME for every row, save, re-run script Sun night"
    ws_wrap["A1"].font = title_font; ws_wrap.merge_cells("A1:I1")
    ws_wrap["A2"] = (
        "Outcome dropdown is in column F. Broker name (col G) — capture every drop. "
        "On next run, outcomes flow into the learning engine and adjust next week's plan."
    )
    ws_wrap["A2"].font = Font(italic=True, color="606060"); ws_wrap.merge_cells("A2:I2")

    h_wrap = ["Company", "Trigger", "Channel", "Day", "Scheduled", "Outcome", "Broker captured", "Notes", "Cadence row #"]
    for i, h in enumerate(h_wrap, 1):
        c = ws_wrap.cell(row=4, column=i, value=h); c.font = h_font; c.fill = h_fill
    ws_wrap.freeze_panes = "A5"

    # Build the row list: every touch whose scheduled_for is in [week_start, week_end] AND not yet completed
    wrap_rows = []
    for ci, c in enumerate(state["active_cadences"]):
        for ti, t in enumerate(c["touches"]):
            sf = t.get("scheduled_for","")
            if not sf: continue
            try:
                sf_d = dt.date.fromisoformat(sf)
            except Exception:
                continue
            if week_start <= sf_d <= week_end and not t.get("completed"):
                wrap_rows.append((ci, ti, c, t, sf_d))
    wrap_rows.sort(key=lambda x: (x[4], x[2]["company"]))

    for r_idx, (ci, ti, c, t, sf_d) in enumerate(wrap_rows, 5):
        ws_wrap.cell(row=r_idx, column=1, value=c["company"][:40])
        ws_wrap.cell(row=r_idx, column=2, value=c.get("primary_trigger",""))
        ws_wrap.cell(row=r_idx, column=3, value=t["channel"])
        ws_wrap.cell(row=r_idx, column=4, value=f"D{t['day']}")
        ws_wrap.cell(row=r_idx, column=5, value=sf_d.isoformat())
        ws_wrap.cell(row=r_idx, column=6, value=t.get("outcome",""))   # rep fills
        ws_wrap.cell(row=r_idx, column=7, value=t.get("broker_captured",""))
        ws_wrap.cell(row=r_idx, column=8, value=t.get("notes",""))
        ws_wrap.cell(row=r_idx, column=9, value=f"{ci}:{ti}")
        ws_wrap.cell(row=r_idx, column=9).font = Font(color="909090", size=9)
        # Channel-coded fill on each row for at-a-glance scanning
        ch_fill = {"email": fill_email, "linkedin": fill_linkedin, "drop": fill_drop, "call": fill_call}.get(t["channel"])
        if ch_fill:
            for col in range(1, 10):
                ws_wrap.cell(row=r_idx, column=col).fill = ch_fill

    if not wrap_rows:
        ws_wrap.cell(row=5, column=1, value="No touches scheduled this week. (Run --init or wait for cadences to start firing.)")
        ws_wrap.cell(row=5, column=1).font = Font(italic=True, color="A06060")
        ws_wrap.merge_cells("A5:I5")

    # Outcome dropdown on column F
    n_data_rows = max(len(wrap_rows), 1)
    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(o for o in OUTCOMES if o) + '"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.error = "Pick one from the dropdown."
    dv.errorTitle = "Invalid outcome"
    ws_wrap.add_data_validation(dv)
    dv.add(f"F5:F{4 + n_data_rows}")

    for col, w in zip("ABCDEFGHI", [40, 18, 11, 6, 12, 18, 28, 50, 12]):
        ws_wrap.column_dimensions[col].width = w

    # ═══ TAB 7: All_Triggers_Bench ═══
    ws7 = wb.create_sheet("All_Triggers_Bench")
    ws7["A1"] = "BENCH — full trigger universe (replacements when portfolio runs dry)"
    ws7["A1"].font = title_font; ws7.merge_cells("A1:G1")
    ws7["A2"] = "Pull from here when a cadence completes. Full universe lives in Trigger_Engine_Output.xlsx."
    ws7["A2"].font = Font(italic=True, color="606060"); ws7.merge_cells("A2:G2")
    h7 = ["#","Company","City","County","EE","Score","Trigger"]
    for i, h in enumerate(h7, 1):
        c = ws7.cell(row=4, column=i, value=h); c.font = h_font; c.fill = h_fill
    # Re-load from trigger universe (top 200, excluding accounts already in active cadences)
    accounts = load_trigger_universe()
    enrich_with_county(accounts)
    active_keys = {c["key"] for c in state["active_cadences"]}
    bench = [a for a in accounts if a["key"] not in active_keys]
    bench.sort(key=lambda x: -x["score"])
    for i, a in enumerate(bench[:200], 5):
        rank = i - 4
        ws7.cell(row=i, column=1, value=rank)
        ws7.cell(row=i, column=2, value=a["company"][:40])
        ws7.cell(row=i, column=3, value=a["city"][:16])
        ws7.cell(row=i, column=4, value=a.get("county",""))
        ws7.cell(row=i, column=5, value=a.get("ee"))
        ws7.cell(row=i, column=6, value=round(a["score"],1))
        ws7.cell(row=i, column=7, value=a.get("primary_trigger",""))
    for col, w in zip("ABCDEFG", [4, 38, 16, 14, 6, 7, 18]):
        ws7.column_dimensions[col].width = w

    # ═══ TAB 8: README ═══
    ws8 = wb.create_sheet("README")
    readme = [
        ("Sales Operating System — README", title_font),
        ("", None),
        (f"Generated: {TODAY_ISO}", None),
        ("", None),
        ("WHAT THIS IS", Font(bold=True, size=12, color="1F4E78")),
        ("", None),
        ("Daily orchestrator. Reads your trigger data, computes your cadence positions,", None),
        ("and outputs THE LIST of actions for today in execution order.", None),
        ("", None),
        ("MORNING WORKFLOW (every weekday)", Font(bold=True, size=12, color="1F4E78")),
        ("", None),
        ("  1. Run: python3 scripts/build_sales_os.py", None),
        ("  2. Open: Sales_OS_MASTER.xlsx → TODAY tab", None),
        ("  3. Execute top-to-bottom. Don't think — execute.", None),
        ("     06:00-07:30 — emails (touch-1 of new cadences)", None),
        ("     07:30-09:00 — phone calls (touch-4, the ASK)", None),
        ("     09:00-15:00 — field route + drops (touch-3)", None),
        ("     15:00-16:00 — LinkedIn DMs (touch-2)", None),
        ("     16:00-17:00 — CRM hygiene + log everything in Worked_Accounts_Registry", None),
        ("", None),
        ("  4. Friday: open Operator_Dashboard tab, fill in week's metrics.", None),
        ("", None),
        ("ARCHITECTURE", Font(bold=True, size=12, color="1F4E78")),
        ("", None),
        ("  Each account flows through the give-give-give-ask 4-touch cadence:", None),
        ("    Day 0  → email (give: industry brief)", None),
        ("    Day 4  → LinkedIn DM (give: their specific data)", None),
        ("    Day 9  → drop-by (give: printed analysis)", None),
        ("    Day 14 → phone call (ASK: 20-min meeting)", None),
        ("", None),
        ("  Cadences are stagger-started by territory route day:", None),
        ("    Mon → Wake/Triangle accounts (their day-9 drops fall on future Mondays)", None),
        ("    Tue → Pitt/Greenville accounts", None),
        ("    Wed → Northern (Nash, Edgecombe, Wilson, Halifax)", None),
        ("    Thu → Cumberland/Sandhills (alt-week Coastal Carteret/Craven)", None),
        ("    Fri → no new starts (CRM, planning, give-blocks)", None),
        ("", None),
        ("  10 new cadences/active-day × 4 active days = 40 new accounts/wk.", None),
        ("  Steady state: ~80 active cadences after day 14.", None),
        ("", None),
        ("INITIALIZATION", Font(bold=True, size=12, color="1F4E78")),
        ("", None),
        ("  Day 1: python3 scripts/build_sales_os.py --init", None),
        ("    Seeds 10 accounts per route day (Mon-Thu) = 40 active cadences from your bench.", None),
        ("    By day 4 → first LinkedIn touches.", None),
        ("    By day 9 → first drops.", None),
        ("    By day 14 → first calls (the ASK).", None),
        ("    By week 3 → at steady state, ~10-15 actions/day on every active weekday.", None),
        ("", None),
        ("STATE", Font(bold=True, size=12, color="1F4E78")),
        ("", None),
        ("  Persists in _sales_os_state.json (next to this workbook).", None),
        ("  Hand-edit if you need to remove an account, change a status, or fix a typo.", None),
        ("  Backup before edits: cp _sales_os_state.json _sales_os_state.bak", None),
        ("", None),
        ("TUNING", Font(bold=True, size=12, color="1F4E78")),
        ("", None),
        ("  After 4 weeks of running, recalibrate by computing actual conversion-to-meeting", None),
        ("  per trigger type. Edit OUTREACH_OPENERS templates as you find what's working.", None),
        ("  Trigger weights live in scripts/build_trigger_engine.py.", None),
        ("", None),
        ("LIMITATIONS / FOLLOW-UPS", Font(bold=True, size=12, color="1F4E78")),
        ("", None),
        ("  • First-name capture is manual (during drops). Update _sales_os_state.json", None),
        ("    with first_name once captured, so future LinkedIn / call copy uses the real name.", None),
        ("  • CPA/broker channel tracking lives in a separate workbook (next build).", None),
        ("  • Trade event calendar lives in a separate workbook (next build).", None),
        ("  • Closed-loop conversion feedback (re-weighting trigger scores) — manual quarterly.", None),
        ("", None),
    ]
    for i, (t, f) in enumerate(readme, 1):
        cell = ws8.cell(row=i, column=1, value=t)
        if f: cell.font = f
    ws8.column_dimensions["A"].width = 110

    wb.save(OUT)
    print(f"\nSaved: {OUT}")
    print(f"Tabs: {wb.sheetnames}")
    return wb


# ─── MAIN ───────────────────────────────────────────────────────────
def main():
    args = sys.argv[1:]
    state = load_state()

    if "--reset-state" in args:
        print("Resetting state to empty.")
        # Wipe the file then reload through migrate-on-load so all schema keys are present
        if STATE_FILE.exists(): STATE_FILE.unlink()
        state = load_state()

    if "--init" in args or not state.get("initialized_at"):
        print("Loading trigger universe…")
        accounts = load_trigger_universe()
        enrich_with_county(accounts)
        displ = load_displacement_universe()
        if displ:
            print(f"  + {len(displ)} displacement (incumbent-PEO) accounts loaded")
            existing_keys = {a["key"] for a in accounts}
            accounts.extend(d for d in displ if d["key"] not in existing_keys)
        industry_signals = load_industry_signal_universe()
        if industry_signals:
            print(f"  + {len(industry_signals)} industry-signal triggers (WC lapsed / OSHA / talent / etc.)")
            existing_keys = {a["key"] for a in accounts}
            accounts.extend(s for s in industry_signals if s["key"] not in existing_keys)
        apply_weights_to_bench(accounts, state["weights"])
        print(f"  {len(accounts)} accounts on the bench")
        state = init_portfolio(state, accounts)

    # Ingest last week's outcomes from the Weekly_Wrap tab BEFORE advancing/scoring
    if "--no-ingest" not in args:
        ingest_weekly_wrap(state)

    # Disposition routing: kill outcomes → drops table; nurture → +90d queue; meeting → MEDDPICC
    apply_dispositions(state)

    # Learning engine: recompute weights from the outcome ledger
    recompute_weights(state)

    advance_cadences(state)

    # Top-up: refill any bucket that dropped below 10 (cadences disqualified, completed, nurtured)
    if state.get("initialized_at") and "--no-topup" not in args:
        accounts_for_topup = load_trigger_universe()
        enrich_with_county(accounts_for_topup)
        displ_topup = load_displacement_universe()
        if displ_topup:
            existing_keys = {a["key"] for a in accounts_for_topup}
            accounts_for_topup.extend(d for d in displ_topup if d["key"] not in existing_keys)
        industry_signals_topup = load_industry_signal_universe()
        if industry_signals_topup:
            existing_keys = {a["key"] for a in accounts_for_topup}
            accounts_for_topup.extend(s for s in industry_signals_topup if s["key"] not in existing_keys)
        apply_weights_to_bench(accounts_for_topup, state["weights"])
        topup_replacements(state, accounts_for_topup)

    print(f"\nState: {len(state['active_cadences'])} active cadences, {len(state['completed'])} completed.")

    build_workbook(state)
    save_state(state)

    # Also print today's summary to console
    weekday_name = ["Mon","Tue","Wed","Thu","Fri","Sat","Sun"][WEEKDAY]
    print(f"\n=== TODAY ({weekday_name} {TODAY_ISO}) ===")
    counts = defaultdict(int)
    for c in state["active_cadences"]:
        action = what_action_today(c["start_date"])
        if action:
            counts[action[1]] += 1
    for ch in ("email","linkedin","drop","call"):
        if counts[ch] > 0:
            print(f"  {counts[ch]:3} × {ch}")
    print(f"  ─────")
    print(f"  {sum(counts.values()):3} total actions")
    print(f"\n  → Open Sales_OS_MASTER.xlsx → TODAY tab. Execute.")

if __name__ == "__main__":
    main()
