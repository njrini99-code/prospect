#!/usr/bin/env python3
"""migrate_to_neon.py — full ETL of ADP PEO prospect data → Neon Postgres.

Sources (read locally; never call CRM / SaaS APIs):
  1. enrichment/db/pipeline.sqlite  (companies + 1M+ enrichments + carriers + drops)
  2. MASTER/uploads_to_drive/ALL_COMPETING_PEO_ACCOUNTS.csv  (674 displacement targets)
  3. _sales_os_state.json  (active cadences + buyer_cast + weights)
  4. Trigger_Engine_Output.xlsx  (WC / Health / OSHA / Hiring trigger tabs)
  5. verification_V2.json  (fact-check verdicts)
  6. ADP_Territory_CRM_Master.xlsx  (territory CRM, "All Prospects" + "By County")

Destination: Neon Postgres (DATABASE_URL in workspace .env — never logged).

Schema: 12 tables + 3 views (see CREATE statements in SCHEMA_SQL below).

Usage:
  python3 scripts/migrate_to_neon.py            # full migration
  python3 scripts/migrate_to_neon.py --fresh    # DROP + recreate schema first
  python3 scripts/migrate_to_neon.py --dry-run  # parse + dedupe, but skip all writes
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import os
import re
import sqlite3
import sys
import time
from collections import defaultdict
from pathlib import Path
from typing import Any

# ─── workspace setup ────────────────────────────────────────────────────────────
ROOT = Path("/Users/ricknini/Documents/ADP PEO")
SQLITE_PATH = ROOT / "enrichment" / "db" / "pipeline.sqlite"
COMPETING_PEO_CSV = ROOT / "MASTER" / "uploads_to_drive" / "ALL_COMPETING_PEO_ACCOUNTS.csv"
SALES_OS_JSON = ROOT / "_sales_os_state.json"
TRIGGER_XLSX = ROOT / "Trigger_Engine_Output.xlsx"
VERIFICATION_JSON = ROOT / "verification_V2.json"
CRM_XLSX = ROOT / "ADP_Territory_CRM_Master.xlsx"

# Load DATABASE_URL from .env without printing the value
def _load_env():
    env_path = ROOT / ".env"
    if not env_path.exists():
        print("✗ .env not found at workspace root", file=sys.stderr)
        sys.exit(1)
    for line in env_path.read_text().splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        os.environ.setdefault(k.strip(), v.strip())
_load_env()

DSN = os.environ.get("DATABASE_URL")
if not DSN:
    print("✗ DATABASE_URL missing in .env", file=sys.stderr)
    sys.exit(1)

try:
    import psycopg
except ImportError:
    print("✗ psycopg not installed. Run: pip install --break-system-packages 'psycopg[binary]'", file=sys.stderr)
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("✗ openpyxl not installed. Run: pip install --break-system-packages openpyxl", file=sys.stderr)
    sys.exit(1)

# ─── ICP exclusion rules (mirrored from scripts/build_sales_os.py) ──────────────
EXCLUDE_NAICS_PREFIXES = (
    "236", "237",
    "23811", "23812", "23813", "23814", "23815", "23816", "23817", "23818", "23819",
    "23829", "23831", "23832", "23833", "23834", "23835", "23839", "23891", "23899",
    "484", "562", "711", "712", "722", "812", "445", "446", "448",
)
EXCLUDE_NAICS_EXACT = {"561730"}
EXCLUDE_NAME_REGEX = re.compile(
    r"\b(construction|contractor|contracting|builders?|"
    r"trucking|freight|hauling|truck\s*stop|truck\s*leasing|dry\s*bulk|"
    r"excavat\w*|paving|roof(?:ing|ers?)|plumb\w+|"
    r"drywall|concrete|mason\w*|carpent\w*|paint(?:ing|ers)|janitorial|"
    r"landscap\w*|lawn\s+care|"
    r"catering|cuisine|restaurant\w*|cafe\w*|bistro|pizz\w+|food\s+truck|"
    r"brewery|brewing\s+co|winery|distillery|"
    r"theater|theatre|cinema|gallery|salon|barber|tanning|nail\s+spa|"
    r"jewel(?:ers|ry)|wine\s+(?:shop|store|cellar)|"
    r"insurance\s+agency|wealth\s+management|"
    r"public\s+school|county\s+(?:public\s+)?school|"
    r"city\s+of|town\s+of|department\s+of|university|community\s+college|"
    r"\bhospital\b|medical\s+center|health\s+system|"
    r"\w+\s+association\b|trade\s+association|chamber\s+of\s+commerce|"
    r"roundtable|"
    r"\bnails?\b|\boptometry\b|\bvision\b|eye\s+care|eyecare|optical|"
    r"olive\s+oil|\bgourmet\b|specialty\s+food|"
    r"auto\s+group|auto\s+sales|auto\s+dealer|car\s+dealer|"
    r"collection\s+service|debt\s+collect|"
    r"learning\s+center|tutoring|"
    r"behavioral\s+(?:services|health)|psychiatric|"
    r"\bpharmacy\b|optometr\w*|\bdds\b|dental(?:\s+(?:office|clinic|practice))?"
    r")\b",
    re.I,
)
OFF_FOCUS_NAME_PATTERNS = (
    "sylvan", "kumon", "mathnasium", "tutoring", "learning center",
    " mart", "quality mart", "convenience store", "gas station", "truck stop",
    "furniture", "appliance", "mattress", "carpet",
    "coins", "pawn", "hobby", "thrift", "antique",
    "grocery", "supermarket", "produce ", " soap ", "soap bar", "soap co",
    "museum", "aquarium",
    " bank", "credit union", "savings & loan", "mortgage",
    " finance ", "finance inc", "financial services", "financial group",
    "logistics", "freight", "fulfillment center",
    "upholstery", "furnishings",
    "auto group", "auto sales", "auto dealer", "car dealer",
    "automotive service", "tire ", "muffler", "transmission",
    "holdings, llc", " holdings llc",
    "behavioral services", "behavioral health", "psychiatric",
    "therapy", "developmental",
    "optometry", "optical", "vision care",
    "pharmacy", "drug store",
    "olive oil", "gourmet", "specialty food",
    "wine ", "winery", "vineyard", "distillery", "spirits",
    "salon", " spa ", "boutique", "tanning",
    "dry clean", "laundry", "cleaners",
    "marketplace", "naturally", " organic ", "pet ",
    "ministry", " church", "fellowship",
)
EXCLUDE_NAME_SUBSTR = (
    "LENOVO", "PEP BOYS", "ABC SUPPLY", "CATALENT", "AISIN", "AMAZON", "WALMART",
    "TARGET CORP", "HOME DEPOT", "LOWE'S", "FEDEX", "UPS INC", "GOODWILL",
    "MCDONALD", "STARBUCKS", "DOLLAR GENERAL", "DOLLAR TREE",
    "COCA-COLA", "COCA COLA", "PEPSI", "DR PEPPER", "ANHEUSER",
    "CHERRY HOSPITAL", "DOROTHEA DIX", "BROUGHTON HOSPITAL",
)
# UNC/Duke mail-drop ZIPs and RTP-proper — out of territory even if 27/28
UNC_DUKE_MAILDROPS = {"27599", "27710", "27705", "27708", "27704"}
RTP_PROPER_ZIP = "27709"
EXCLUDE_ZIPS = UNC_DUKE_MAILDROPS | {RTP_PROPER_ZIP}

def is_disqualified(name: str | None, naics: str | None, zip5: str | None) -> tuple[bool, str | None]:
    """Return (disqualified, reason) for an ICP off-focus check."""
    n = (name or "").strip()
    if not n:
        return True, "NO_NAME"
    nu = n.upper()
    nl = n.lower()
    if EXCLUDE_NAME_REGEX.search(n):
        return True, "NAME_REGEX"
    if any(s in nu for s in EXCLUDE_NAME_SUBSTR):
        return True, "NAME_SUBSTR"
    if any(p in nl for p in OFF_FOCUS_NAME_PATTERNS):
        return True, "OFF_FOCUS_NAME"
    if naics:
        nx = str(naics).strip()
        if nx in EXCLUDE_NAICS_EXACT:
            return True, "NAICS_EXACT"
        if any(nx.startswith(p) for p in EXCLUDE_NAICS_PREFIXES):
            return True, "NAICS_PREFIX"
    if zip5 and zip5 in EXCLUDE_ZIPS:
        return True, "ZIP_OUT_OF_TERRITORY"
    return False, None

def is_nc_territory(state: str | None, zip5: str | None) -> bool:
    """Per the workspace memory: NC territory = state='NC' OR zip starts with 27/28."""
    if state and state.strip().upper() == "NC":
        return True
    if zip5 and (zip5.startswith("27") or zip5.startswith("28")):
        return True
    return False

# ─── normalization helpers ──────────────────────────────────────────────────────
def norm_name(s: str | None) -> str:
    if not s:
        return ""
    return re.sub(r"[^a-z0-9]", "", str(s).lower())

def clean_zip(z: Any) -> str | None:
    if z is None or z == "":
        return None
    s = str(z).strip()
    # strip ".0" tails from float-as-string in CSV
    if s.endswith(".0"):
        s = s[:-2]
    s = re.sub(r"[^0-9]", "", s)
    if not s:
        return None
    return s[:5].zfill(5) if len(s) <= 5 else s[:5]

def clean_phone(p: Any) -> str | None:
    if p is None or p == "":
        return None
    s = re.sub(r"[^0-9]", "", str(p))
    if not s:
        return None
    # strip trailing .0 already handled by non-digit strip
    if len(s) == 11 and s.startswith("1"):
        s = s[1:]
    return s[:10] if len(s) >= 10 else s

def clean_int(v: Any) -> int | None:
    if v is None or v == "":
        return None
    try:
        return int(float(str(v).replace(",", "").strip()))
    except (ValueError, TypeError):
        return None

def clean_float(v: Any) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(",", "").strip())
    except (ValueError, TypeError):
        return None

def clean_date(v: Any) -> str | None:
    if v is None or v == "":
        return None
    if isinstance(v, (dt.date, dt.datetime)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%Y-%m-%dT%H:%M:%S"):
        try:
            return dt.datetime.strptime(s[:len(fmt)+5], fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None

def truthy(v: Any) -> bool:
    if v is None or v == "":
        return False
    s = str(v).strip().lower()
    return s in ("1", "true", "t", "yes", "y", "x")

# ─── schema ─────────────────────────────────────────────────────────────────────
SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS companies (
  id              SERIAL PRIMARY KEY,
  name            TEXT NOT NULL,
  name_normalized TEXT NOT NULL,
  city            TEXT,
  county          TEXT,
  state           TEXT DEFAULT 'NC',
  zip             TEXT,
  ein             TEXT,
  domain          TEXT,
  website         TEXT,
  linkedin_url    TEXT,
  naics           TEXT,
  vertical        TEXT,
  ee              INTEGER,
  multi_state_likely BOOLEAN DEFAULT FALSE,
  fed_contractor  BOOLEAN DEFAULT FALSE,
  has_5500        BOOLEAN DEFAULT FALSE,
  has_health_carriers BOOLEAN DEFAULT FALSE,
  growth_signal   TEXT,
  pitch_signal    TEXT,
  pitch_angle     TEXT,
  disqualified    BOOLEAN DEFAULT FALSE,
  disqualified_reason TEXT,
  source          TEXT,
  first_seen      TIMESTAMPTZ DEFAULT NOW(),
  last_updated    TIMESTAMPTZ DEFAULT NOW(),
  UNIQUE (name_normalized, zip)
);
CREATE INDEX IF NOT EXISTS idx_companies_name_norm ON companies(name_normalized);
CREATE INDEX IF NOT EXISTS idx_companies_zip       ON companies(zip);
CREATE INDEX IF NOT EXISTS idx_companies_state     ON companies(state);
CREATE INDEX IF NOT EXISTS idx_companies_vertical  ON companies(vertical);
CREATE INDEX IF NOT EXISTS idx_companies_disqual   ON companies(disqualified);

CREATE TABLE IF NOT EXISTS contacts (
  id           SERIAL PRIMARY KEY,
  company_id   INTEGER NOT NULL REFERENCES companies(id) ON DELETE CASCADE,
  name         TEXT,
  title        TEXT,
  email        TEXT,
  phone        TEXT,
  linkedin_url TEXT,
  is_primary   BOOLEAN DEFAULT FALSE,
  source       TEXT,
  created_at   TIMESTAMPTZ DEFAULT NOW()
);
CREATE INDEX IF NOT EXISTS idx_contacts_company ON contacts(company_id);
CREATE INDEX IF NOT EXISTS idx_contacts_email   ON contacts(email);

CREATE TABLE IF NOT EXISTS carriers (
  id           SERIAL PRIMARY KEY,
  company_id   INTEGER NOT NULL REFERENCES companies(id) ON DELETE CASCADE,
  carrier_name TEXT NOT NULL,
  benefit_type TEXT,
  naic         TEXT,
  premium      NUMERIC,
  covered_lives INTEGER,
  plan_year    INTEGER,
  source       TEXT,
  observed_at  TIMESTAMPTZ DEFAULT NOW()
);
CREATE INDEX IF NOT EXISTS idx_carriers_company ON carriers(company_id);
CREATE INDEX IF NOT EXISTS idx_carriers_name    ON carriers(carrier_name);

CREATE TABLE IF NOT EXISTS triggers (
  id           SERIAL PRIMARY KEY,
  company_id   INTEGER NOT NULL REFERENCES companies(id) ON DELETE CASCADE,
  trigger_type TEXT NOT NULL,
  score        NUMERIC,
  evidence     TEXT,
  trigger_date DATE,
  source       TEXT,
  created_at   TIMESTAMPTZ DEFAULT NOW()
);
CREATE INDEX IF NOT EXISTS idx_triggers_company ON triggers(company_id);
CREATE INDEX IF NOT EXISTS idx_triggers_type    ON triggers(trigger_type);

CREATE TABLE IF NOT EXISTS incumbent_peo (
  id           SERIAL PRIMARY KEY,
  company_id   INTEGER NOT NULL REFERENCES companies(id) ON DELETE CASCADE,
  peo_brand    TEXT NOT NULL,
  confidence   TEXT,
  evidence     TEXT,
  filing_year  INTEGER,
  source       TEXT,
  observed_at  TIMESTAMPTZ DEFAULT NOW()
);
CREATE INDEX IF NOT EXISTS idx_incumbent_company ON incumbent_peo(company_id);
CREATE INDEX IF NOT EXISTS idx_incumbent_brand   ON incumbent_peo(peo_brand);

CREATE TABLE IF NOT EXISTS cadences (
  id              SERIAL PRIMARY KEY,
  company_id      INTEGER REFERENCES companies(id) ON DELETE CASCADE,
  company_key     TEXT UNIQUE NOT NULL,
  status          TEXT,
  start_date      DATE,
  route_day       INTEGER,
  score           NUMERIC,
  score_raw       NUMERIC,
  score_overlay   NUMERIC,
  weight_mult     NUMERIC,
  primary_trigger TEXT,
  tier            TEXT,
  fitness_tier    TEXT,
  enrichment_notes TEXT,
  evidence        TEXT,
  talk_track      TEXT,
  vertical        TEXT,
  created_at      TIMESTAMPTZ DEFAULT NOW()
);
CREATE INDEX IF NOT EXISTS idx_cadences_company ON cadences(company_id);
CREATE INDEX IF NOT EXISTS idx_cadences_status  ON cadences(status);

CREATE TABLE IF NOT EXISTS touches (
  id              SERIAL PRIMARY KEY,
  cadence_id      INTEGER NOT NULL REFERENCES cadences(id) ON DELETE CASCADE,
  day_offset      INTEGER,
  channel         TEXT,
  scheduled_for   DATE,
  completed       BOOLEAN DEFAULT FALSE,
  outcome         TEXT,
  notes           TEXT,
  broker_captured TEXT,
  created_at      TIMESTAMPTZ DEFAULT NOW()
);
CREATE INDEX IF NOT EXISTS idx_touches_cadence ON touches(cadence_id);
CREATE INDEX IF NOT EXISTS idx_touches_sched   ON touches(scheduled_for);

CREATE TABLE IF NOT EXISTS outcomes_ledger (
  id            BIGSERIAL PRIMARY KEY,
  touch_id      INTEGER REFERENCES touches(id) ON DELETE SET NULL,
  cadence_id    INTEGER REFERENCES cadences(id) ON DELETE SET NULL,
  company_key   TEXT,
  week_start    DATE,
  logged_at     DATE,
  trigger_type  TEXT,
  vertical      TEXT,
  channel       TEXT,
  day_offset    INTEGER,
  route_day     INTEGER,
  scheduled_for DATE,
  outcome       TEXT,
  broker_captured TEXT,
  notes         TEXT,
  created_at    TIMESTAMPTZ DEFAULT NOW()
);
CREATE INDEX IF NOT EXISTS idx_outcomes_company ON outcomes_ledger(company_key);

CREATE TABLE IF NOT EXISTS notes (
  id          SERIAL PRIMARY KEY,
  company_id  INTEGER NOT NULL REFERENCES companies(id) ON DELETE CASCADE,
  body        TEXT NOT NULL,
  source      TEXT,
  created_at  TIMESTAMPTZ DEFAULT NOW()
);
CREATE INDEX IF NOT EXISTS idx_notes_company ON notes(company_id);

CREATE TABLE IF NOT EXISTS tasks (
  id          SERIAL PRIMARY KEY,
  company_id  INTEGER REFERENCES companies(id) ON DELETE CASCADE,
  title       TEXT NOT NULL,
  body        TEXT,
  due_date    DATE,
  status      TEXT DEFAULT 'open',
  created_at  TIMESTAMPTZ DEFAULT NOW()
);
CREATE INDEX IF NOT EXISTS idx_tasks_company ON tasks(company_id);
CREATE INDEX IF NOT EXISTS idx_tasks_status  ON tasks(status);

CREATE TABLE IF NOT EXISTS meddpicc (
  id                  SERIAL PRIMARY KEY,
  company_id          INTEGER REFERENCES companies(id) ON DELETE CASCADE,
  company_key         TEXT UNIQUE,
  first_meeting_date  DATE,
  stage               TEXT,
  m_metrics           TEXT,
  e_econ_buyer        TEXT,
  d1_decision_criteria TEXT,
  d2_decision_process TEXT,
  p_paper_process     TEXT,
  i_pain              TEXT,
  c_champion          TEXT,
  cmp_competition     TEXT,
  next_action         TEXT,
  updated_at          TIMESTAMPTZ DEFAULT NOW()
);

CREATE TABLE IF NOT EXISTS brokers (
  id                  SERIAL PRIMARY KEY,
  name                TEXT UNIQUE NOT NULL,
  county              TEXT,
  phone               TEXT,
  email               TEXT,
  first_seen          DATE,
  first_seen_via      TEXT,
  last_brief_sent     DATE,
  lunch_status        TEXT,
  n_clients_observed  INTEGER,
  notes               TEXT,
  created_at          TIMESTAMPTZ DEFAULT NOW()
);

CREATE TABLE IF NOT EXISTS weights_current (
  id                  SERIAL PRIMARY KEY,
  dim                 TEXT NOT NULL,
  key                 TEXT NOT NULL,
  mult                NUMERIC NOT NULL,
  last_recomputed     DATE,
  n_outcomes          INTEGER DEFAULT 0,
  UNIQUE (dim, key)
);

CREATE TABLE IF NOT EXISTS weekly_metrics (
  id                  SERIAL PRIMARY KEY,
  week_start          DATE UNIQUE,
  touches             INTEGER DEFAULT 0,
  progressions        INTEGER DEFAULT 0,
  meetings_booked     INTEGER DEFAULT 0,
  meetings_held       INTEGER DEFAULT 0,
  killed              INTEGER DEFAULT 0,
  created_at          TIMESTAMPTZ DEFAULT NOW()
);

-- ─── views ────────────────────────────────────────────────────────────────────
CREATE OR REPLACE VIEW v_today_actions AS
SELECT
  c.id AS company_id,
  c.name AS company,
  c.city, c.county, c.zip,
  cd.company_key,
  cd.score,
  cd.primary_trigger,
  cd.tier,
  t.id AS touch_id,
  t.channel,
  t.day_offset,
  t.scheduled_for
FROM touches t
JOIN cadences cd ON cd.id = t.cadence_id
LEFT JOIN companies c ON c.id = cd.company_id
WHERE NOT t.completed
  AND t.scheduled_for <= CURRENT_DATE
ORDER BY cd.score DESC NULLS LAST, t.scheduled_for ASC;

CREATE OR REPLACE VIEW v_active_pipeline AS
SELECT
  c.id AS company_id,
  c.name AS company,
  c.city, c.county, c.zip, c.ee, c.vertical,
  cd.score, cd.primary_trigger, cd.tier, cd.status,
  cd.start_date,
  (SELECT MIN(t.scheduled_for) FROM touches t
   WHERE t.cadence_id = cd.id AND NOT t.completed) AS next_touch_date,
  (SELECT t.channel FROM touches t
   WHERE t.cadence_id = cd.id AND NOT t.completed
   ORDER BY t.scheduled_for ASC LIMIT 1) AS next_touch_channel
FROM cadences cd
LEFT JOIN companies c ON c.id = cd.company_id
WHERE cd.status = 'active'
ORDER BY cd.score DESC NULLS LAST;

CREATE OR REPLACE VIEW v_bench_top100 AS
SELECT
  c.id, c.name, c.city, c.county, c.zip, c.ee, c.vertical,
  c.has_5500, c.has_health_carriers, c.multi_state_likely, c.fed_contractor,
  c.growth_signal, c.pitch_signal, c.pitch_angle
FROM companies c
WHERE NOT c.disqualified
  AND (c.state = 'NC' OR c.zip LIKE '27%' OR c.zip LIKE '28%')
  AND c.ee BETWEEN 11 AND 55
  AND c.id NOT IN (SELECT company_id FROM cadences WHERE company_id IS NOT NULL)
ORDER BY
  (c.has_health_carriers::int + c.has_5500::int +
   c.multi_state_likely::int + c.fed_contractor::int) DESC,
  c.ee DESC
LIMIT 100;
"""

DROP_SQL = """
DROP VIEW IF EXISTS v_today_actions   CASCADE;
DROP VIEW IF EXISTS v_active_pipeline CASCADE;
DROP VIEW IF EXISTS v_bench_top100    CASCADE;
DROP TABLE IF EXISTS weekly_metrics   CASCADE;
DROP TABLE IF EXISTS weights_current  CASCADE;
DROP TABLE IF EXISTS brokers          CASCADE;
DROP TABLE IF EXISTS meddpicc         CASCADE;
DROP TABLE IF EXISTS tasks            CASCADE;
DROP TABLE IF EXISTS notes            CASCADE;
DROP TABLE IF EXISTS outcomes_ledger  CASCADE;
DROP TABLE IF EXISTS touches          CASCADE;
DROP TABLE IF EXISTS cadences         CASCADE;
DROP TABLE IF EXISTS incumbent_peo    CASCADE;
DROP TABLE IF EXISTS triggers         CASCADE;
DROP TABLE IF EXISTS carriers         CASCADE;
DROP TABLE IF EXISTS contacts         CASCADE;
DROP TABLE IF EXISTS companies        CASCADE;
"""

# ─── company dedupe registry ────────────────────────────────────────────────────
class CompanyRegistry:
    """Tracks companies seen across sources, keyed on (name_normalized, zip5-or-state).

    Stores a dict of company attributes; later sources can fill in nulls but
    never overwrite a non-null value with null.
    """
    def __init__(self):
        # primary index: (name_norm, zip5) -> row_dict
        self.by_key: dict[tuple[str, str], dict] = {}
        # secondary index: name_norm -> list of (zip5, key) for fallback matching
        self.by_name: dict[str, list[tuple[str, tuple[str, str]]]] = defaultdict(list)
        # synthetic 1-based id assigned at flush time
        self._next_id = 1

    @staticmethod
    def _bucket(name_norm: str, zip5: str | None, state: str | None) -> tuple[str, str]:
        return (name_norm, zip5 or (state or "NC"))

    def upsert(self, row: dict) -> tuple[str, str] | None:
        """Insert or merge a row. Returns the dedup key, or None if rejected."""
        name = (row.get("name") or "").strip()
        if not name:
            return None
        nn = norm_name(name)
        if not nn:
            return None
        zip5 = clean_zip(row.get("zip"))
        state = (row.get("state") or "NC").strip().upper() or "NC"
        key = self._bucket(nn, zip5, state)
        # try fallback: if zip missing here but we've seen this name with a zip, merge into it
        if zip5 is None and nn in self.by_name:
            # take the first one we've seen
            key = self.by_name[nn][0][1]
        existing = self.by_key.get(key)
        if existing is None:
            row["name_normalized"] = nn
            row["state"] = state
            row["zip"] = zip5
            self.by_key[key] = row
            self.by_name[nn].append((zip5 or "", key))
        else:
            # fill nulls + accumulate booleans
            for k, v in row.items():
                if v in (None, "", False) and k not in ("disqualified",):
                    continue
                cur = existing.get(k)
                if k in ("multi_state_likely", "fed_contractor", "has_5500",
                         "has_health_carriers", "disqualified"):
                    existing[k] = bool(cur) or bool(v)
                elif cur in (None, "", False) or k == "last_updated":
                    existing[k] = v
            # merge sources (track every origin)
            existing_src = existing.get("source") or ""
            new_src = row.get("source") or ""
            if new_src and new_src not in existing_src.split("|"):
                existing["source"] = (existing_src + "|" + new_src).strip("|")
        return key

    def assign_ids(self) -> dict[tuple[str, str], int]:
        """Assign 1..N synthetic IDs in insertion order. Returns key -> id."""
        ids: dict[tuple[str, str], int] = {}
        for i, k in enumerate(self.by_key.keys(), start=1):
            ids[k] = i
        return ids

    def rows(self):
        return self.by_key.items()

    def get_by_name(self, name: str, zip5: str | None = None) -> tuple[str, str] | None:
        nn = norm_name(name)
        if not nn:
            return None
        if zip5:
            key = (nn, zip5)
            if key in self.by_key:
                return key
        # fallback: any bucket sharing the name
        for _, key in self.by_name.get(nn, []):
            return key
        return None

    def __len__(self):
        return len(self.by_key)


# ─── stage 1: load companies + enrichments from pipeline.sqlite ─────────────────
# Fields we promote from key/value enrichments into the flat companies row.
SCALAR_FIELDS = {
    "phone", "website", "linkedin_url", "industry", "naics", "naics_primary",
    "naics_inferred", "ee_count", "ee_band", "growth_signal", "pitch_signal_primary",
    "pitch_angle", "schedule_a_carriers", "is_5500_filer",
}
BOOL_FIELDS_TRUTHY = {
    "is_5500_filer", "federal_contractor", "on_competing_peo",
}

def load_sqlite(reg: CompanyRegistry, progress_every: int = 10_000) -> tuple[int, int]:
    """Returns (n_companies_seen, n_companies_kept)."""
    if not SQLITE_PATH.exists():
        print(f"  [warn] sqlite not found: {SQLITE_PATH}")
        return (0, 0)
    print(f"  reading {SQLITE_PATH.name}...")
    conn = sqlite3.connect(str(SQLITE_PATH))
    conn.row_factory = sqlite3.Row

    # 1. Aggregate enrichments per company_id ONCE (one streaming query, faster than per-company joins)
    enrich_idx: dict[int, dict[str, str]] = defaultdict(dict)
    print("  scanning enrichments for relevant fields...")
    field_filter = ",".join(["?"] * len(SCALAR_FIELDS | BOOL_FIELDS_TRUTHY | {
        "primary_contact_name", "primary_contact_title", "contact_email",
        "owner_or_ceo", "ceo", "wc_carrier", "wc_expiration_date",
        "biz_event_acquisition", "warn_filed_date", "sba_ppp_loan_amount",
    }))
    all_fields = list(SCALAR_FIELDS | BOOL_FIELDS_TRUTHY | {
        "primary_contact_name", "primary_contact_title", "contact_email",
        "owner_or_ceo", "ceo", "wc_carrier", "wc_expiration_date",
        "biz_event_acquisition", "warn_filed_date", "sba_ppp_loan_amount",
    })
    q = f"SELECT company_id, field, value FROM enrichments WHERE field IN ({field_filter})"
    cur = conn.execute(q, all_fields)
    n_enr = 0
    while True:
        batch = cur.fetchmany(50_000)
        if not batch:
            break
        for cid, field, value in batch:
            if value is None or value == "":
                continue
            # keep first non-null value per (company,field)
            if field not in enrich_idx[cid]:
                enrich_idx[cid][field] = value
        n_enr += len(batch)
        if n_enr % 200_000 == 0:
            print(f"    enrichments scanned: {n_enr:,}")
    print(f"    enrichments indexed for {len(enrich_idx):,} companies")

    # 2. Stream companies in chunks, build registry entries
    n_companies = conn.execute("SELECT COUNT(*) FROM companies").fetchone()[0]
    print(f"  streaming {n_companies:,} companies in chunks...")
    seen = 0
    kept = 0
    offset = 0
    CHUNK = 10_000
    while True:
        rows = conn.execute(
            "SELECT company_id, name_display, name_normalized, ein, domain, "
            "city, zip, county, state, first_seen_at, last_updated_at "
            "FROM companies ORDER BY company_id LIMIT ? OFFSET ?",
            (CHUNK, offset)
        ).fetchall()
        if not rows:
            break
        for r in rows:
            seen += 1
            name = r["name_display"]
            if not name or not name.strip():
                continue
            cid = r["company_id"]
            e = enrich_idx.get(cid, {})
            ee = clean_int(e.get("ee_count"))
            naics = e.get("naics_primary") or e.get("naics") or e.get("naics_inferred")
            zip5 = clean_zip(r["zip"])
            disq, reason = is_disqualified(name, naics, zip5)
            # outside-NC -> mark disqualified (territory filter), but keep the row
            if not is_nc_territory(r["state"], zip5):
                disq, reason = True, reason or "OUT_OF_TERRITORY"
            # accumulate
            row = {
                "name": name.strip(),
                "city": (r["city"] or "").strip() or None,
                "county": (r["county"] or "").strip() or None,
                "state": (r["state"] or "NC").strip() or "NC",
                "zip": zip5,
                "ein": (r["ein"] or "").strip() or None,
                "domain": (r["domain"] or "").strip() or None,
                "website": e.get("website") or None,
                "linkedin_url": e.get("linkedin_url") or None,
                "naics": (naics or "").strip() or None,
                "vertical": e.get("industry") or None,
                "ee": ee,
                "multi_state_likely": False,
                "fed_contractor": truthy(e.get("federal_contractor")),
                "has_5500": truthy(e.get("is_5500_filer")),
                "has_health_carriers": bool(e.get("schedule_a_carriers")),
                "growth_signal": e.get("growth_signal"),
                "pitch_signal": e.get("pitch_signal_primary"),
                "pitch_angle": e.get("pitch_angle"),
                "disqualified": disq,
                "disqualified_reason": reason,
                "source": "sqlite",
            }
            reg.upsert(row)
            kept += 1
            # stash enrichment dict on the row for later contact/carrier extraction
            key = reg.get_by_name(name, zip5)
            if key:
                # tuck the cid in for later cross-ref (contacts, carriers, drops)
                row_in_reg = reg.by_key[key]
                row_in_reg.setdefault("_sqlite_company_ids", set()).add(cid)
                row_in_reg.setdefault("_enrichments", {}).update(e)
        offset += CHUNK
        if offset % progress_every == 0 or offset >= n_companies:
            print(f"    progress: {offset:,}/{n_companies:,} ({kept:,} kept)")
    conn.close()
    return (seen, kept)


def collect_sqlite_aux(reg: CompanyRegistry) -> tuple[list[tuple], list[tuple], list[tuple], list[tuple], list[tuple]]:
    """After companies are loaded, pull contacts/carriers/drops/triggers/incumbents
    from the sqlite enrichments and carriers table for the company_ids we tracked.

    Returns (contacts_rows, carriers_rows, triggers_rows, incumbent_rows, notes_rows)
    formatted for psycopg executemany. company_id is the *Neon* id (assigned by
    reg.assign_ids), not the sqlite id.
    """
    if not SQLITE_PATH.exists():
        return [], [], [], [], []
    print("  collecting contacts/carriers/triggers from sqlite...")
    conn = sqlite3.connect(str(SQLITE_PATH))
    conn.row_factory = sqlite3.Row
    ids = reg.assign_ids()
    contacts, carriers_rows, triggers, incumbents, notes = [], [], [], [], []
    # build sqlite_cid -> neon_id reverse map
    cid_to_neon: dict[int, int] = {}
    for key, neon_id in ids.items():
        row = reg.by_key[key]
        for sid in row.get("_sqlite_company_ids", set()):
            cid_to_neon[sid] = neon_id

    # carriers — straight pull
    print("    carriers...")
    cur = conn.execute(
        "SELECT company_id, carrier_name, benefit_type, naic, premium, "
        "covered_lives, plan_year, source FROM carriers"
    )
    while True:
        batch = cur.fetchmany(10_000)
        if not batch:
            break
        for c in batch:
            nid = cid_to_neon.get(c["company_id"])
            if not nid:
                continue
            carriers_rows.append((
                nid, c["carrier_name"], c["benefit_type"], c["naic"],
                c["premium"], c["covered_lives"], c["plan_year"], c["source"],
            ))

    # contacts — derive from enrichments (primary_contact_name + title + contact_email)
    # plus owner_or_ceo / ceo as additional contacts.
    print("    contacts...")
    contact_fields = ("primary_contact_name", "primary_contact_title",
                      "contact_email", "owner_or_ceo", "ceo", "phone")
    field_filter = ",".join(["?"] * len(contact_fields))
    cur = conn.execute(
        f"SELECT company_id, field, value FROM enrichments WHERE field IN ({field_filter})",
        contact_fields
    )
    contact_idx: dict[int, dict[str, str]] = defaultdict(dict)
    while True:
        batch = cur.fetchmany(50_000)
        if not batch:
            break
        for cid, field, value in batch:
            if value and field not in contact_idx[cid]:
                contact_idx[cid][field] = value
    for sid, e in contact_idx.items():
        nid = cid_to_neon.get(sid)
        if not nid:
            continue
        pname = e.get("primary_contact_name")
        ptitle = e.get("primary_contact_title")
        pemail = e.get("contact_email")
        pphone = clean_phone(e.get("phone"))
        if pname or pemail:
            contacts.append((
                nid, pname, ptitle, pemail, pphone, None, True, "sqlite_primary",
            ))
        owner = e.get("owner_or_ceo") or e.get("ceo")
        if owner and owner != pname:
            contacts.append((
                nid, owner, "Owner/CEO", None, None, None, False, "sqlite_owner",
            ))

    # triggers — emit one row per company per trigger field if present
    # (pitch_signal_primary, biz_event_acquisition, warn_filed_date, sba_ppp_loan_amount)
    print("    triggers...")
    trig_fields = ("pitch_signal_primary", "biz_event_acquisition",
                   "warn_filed_date", "wc_expiration_date", "growth_signal")
    field_filter = ",".join(["?"] * len(trig_fields))
    cur = conn.execute(
        f"SELECT company_id, field, value FROM enrichments WHERE field IN ({field_filter})",
        trig_fields
    )
    while True:
        batch = cur.fetchmany(50_000)
        if not batch:
            break
        for cid, field, value in batch:
            nid = cid_to_neon.get(cid)
            if not nid or not value:
                continue
            tdate = clean_date(value) if field in ("wc_expiration_date", "warn_filed_date") else None
            triggers.append((nid, field, None, str(value)[:500], tdate, "sqlite"))

    # incumbent_peo — competing_peo_brand / competing_peo_name + peo_registry_brand
    print("    incumbent PEOs...")
    peo_fields = ("competing_peo_brand", "competing_peo_name", "peo_registry_brand",
                  "peo_sponsor_name", "peo_evidence", "peo_confidence_score")
    field_filter = ",".join(["?"] * len(peo_fields))
    cur = conn.execute(
        f"SELECT company_id, field, value FROM enrichments WHERE field IN ({field_filter})",
        peo_fields
    )
    peo_idx: dict[int, dict[str, str]] = defaultdict(dict)
    while True:
        batch = cur.fetchmany(50_000)
        if not batch:
            break
        for cid, field, value in batch:
            if value and field not in peo_idx[cid]:
                peo_idx[cid][field] = value
    for sid, e in peo_idx.items():
        nid = cid_to_neon.get(sid)
        if not nid:
            continue
        brand = e.get("competing_peo_brand") or e.get("competing_peo_name") or e.get("peo_registry_brand") or e.get("peo_sponsor_name")
        if not brand:
            continue
        evidence = e.get("peo_evidence")
        conf = e.get("peo_confidence_score")
        incumbents.append((nid, str(brand)[:200], conf, evidence, None, "sqlite_5500"))

    conn.close()
    return contacts, carriers_rows, triggers, incumbents, notes


# ─── stage 2: load competing PEO CSV ────────────────────────────────────────────
def load_competing_peo_csv(reg: CompanyRegistry) -> tuple[int, list[dict]]:
    """Load 674 displacement targets. Returns (rows_seen, per_row_aux_payload)
    where aux carries contacts/incumbents to append later (after IDs assigned).
    """
    if not COMPETING_PEO_CSV.exists():
        print(f"  [warn] CSV not found: {COMPETING_PEO_CSV}")
        return 0, []
    print(f"  reading {COMPETING_PEO_CSV.name}...")
    aux = []
    n = 0
    with open(COMPETING_PEO_CSV, newline="", encoding="utf-8") as f:
        r = csv.DictReader(f)
        for row in r:
            n += 1
            name = (row.get("company_name") or "").strip()
            if not name:
                continue
            zip5 = clean_zip(row.get("zip"))
            ee = clean_int(row.get("ee_count"))
            disq, reason = is_disqualified(name, None, zip5)
            if (row.get("viable_for_totalsource") or "").strip().lower() in ("false", "0", "no"):
                disq = True
                reason = reason or (row.get("disqualification_reason") or "NOT_VIABLE")
            if (row.get("is_adp_existing_customer") or "").strip().lower() in ("true", "1", "yes"):
                disq = True
                reason = reason or "IS_ADP_CUSTOMER"
            crow = {
                "name": name,
                "city": (row.get("city") or "").strip() or None,
                "county": (row.get("county") or "").strip() or None,
                "state": "NC",
                "zip": zip5,
                "ee": ee,
                "vertical": row.get("recommended_section") or None,
                "disqualified": disq,
                "disqualified_reason": reason,
                "source": "competing_peo_csv",
            }
            reg.upsert(crow)
            aux.append({
                "name": name,
                "zip": zip5,
                "phone": clean_phone(row.get("phone")),
                "dm_name": row.get("named_dm_name") or None,
                "dm_role": row.get("named_dm_role") or None,
                "dm_email": row.get("named_dm_email") or None,
                "current_peo": row.get("current_peo") or None,
                "current_peo_evidence": row.get("current_peo_evidence") or None,
                "fitness_score": clean_float(row.get("combined_fitness_score")),
                "fitness_tier": row.get("fitness_tier") or None,
                "talk_track": row.get("talk_track") or None,
                "signals_seen": row.get("signals_seen") or None,
            })
    return n, aux


# ─── stage 3: load sales_os_state.json (cadences, buyer_cast, weights) ──────────
def load_sales_os(reg: CompanyRegistry) -> dict:
    if not SALES_OS_JSON.exists():
        print(f"  [warn] sales OS state not found: {SALES_OS_JSON}")
        return {}
    print(f"  reading {SALES_OS_JSON.name}...")
    state = json.loads(SALES_OS_JSON.read_text())
    for c in state.get("active_cadences", []):
        name = c.get("company") or ""
        zip5 = clean_zip(c.get("zip"))
        ee = clean_int(c.get("ee"))
        crow = {
            "name": name,
            "city": c.get("city") or None,
            "county": c.get("county") or None,
            "state": "NC",
            "zip": zip5,
            "ee": ee,
            "vertical": c.get("vertical") or None,
            "naics": c.get("naics") or None,
            "multi_state_likely": bool(c.get("multi_state_likely")),
            "has_health_carriers": bool(c.get("has_health_benefits")),
            "growth_signal": c.get("growth_tier") or None,
            "pitch_signal": c.get("primary_trigger") or None,
            "source": "sales_os",
        }
        reg.upsert(crow)
    return state


# ─── stage 4: load trigger engine ───────────────────────────────────────────────
TRIGGER_TAB_TYPE = {
    "WC_Renewals_120d": "wc_renewal",
    "Health_Renewals": "health_renewal",
    "OSHA_Hot": "osha_hot",
    "Hiring_Velocity": "hiring_velocity",
}

def load_trigger_engine(reg: CompanyRegistry) -> list[dict]:
    if not TRIGGER_XLSX.exists():
        print(f"  [warn] trigger xlsx not found: {TRIGGER_XLSX}")
        return []
    print(f"  reading {TRIGGER_XLSX.name}...")
    aux = []
    wb = openpyxl.load_workbook(TRIGGER_XLSX, read_only=True, data_only=True)
    for tab, ttype in TRIGGER_TAB_TYPE.items():
        if tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        rows = list(ws.iter_rows(values_only=True))
        # find header row
        header_row = None
        for i, r in enumerate(rows[:5]):
            if r and r[0] == "Company":
                header_row = i
                break
        if header_row is None:
            continue
        headers = rows[header_row]
        for r in rows[header_row+1:]:
            if not r or not r[0]:
                continue
            d = {h: v for h, v in zip(headers, r) if h}
            name = (d.get("Company") or "").strip()
            if not name:
                continue
            zip5 = clean_zip(d.get("ZIP"))
            ee = clean_int(d.get("EE"))
            crow = {
                "name": name,
                "city": (d.get("City") or "").strip() or None,
                "state": "NC",
                "zip": zip5,
                "ee": ee,
                "vertical": d.get("Vertical") or None,
                "source": "trigger_engine",
            }
            reg.upsert(crow)
            aux.append({
                "name": name,
                "zip": zip5,
                "trigger_type": ttype,
                "score": clean_float(d.get("Score")),
                "evidence": (
                    d.get("Evidence")
                    or d.get("Carriers (Sched A)")
                    or (f"WC carrier: {d.get('WC Carrier')}; renewal: {d.get('Renewal')}"
                        if d.get("WC Carrier") or d.get("Renewal") else None)
                ),
                "trigger_date": clean_date(d.get("Renewal")) if ttype == "wc_renewal" else None,
            })
    wb.close()
    return aux


# ─── stage 5: load verification JSON ────────────────────────────────────────────
def load_verification(reg: CompanyRegistry) -> list[dict]:
    if not VERIFICATION_JSON.exists():
        return []
    print(f"  reading {VERIFICATION_JSON.name}...")
    data = json.loads(VERIFICATION_JSON.read_text())
    notes_aux = []
    if isinstance(data, list):
        for v in data:
            name = (v.get("company") or "").strip()
            if not name:
                continue
            reg.upsert({"name": name, "state": "NC", "source": "verification"})
            verdict = v.get("overall_verdict")
            failed = v.get("failed_fields") or []
            review = v.get("review_notes") or ""
            body = f"VERIFICATION: {verdict}"
            if failed:
                body += f"\nFailed fields: {', '.join(failed)}"
            if review:
                body += f"\nNotes: {review}"
            notes_aux.append({"name": name, "body": body, "source": "verification_v2"})
    return notes_aux


# ─── stage 6: load CRM master xlsx ──────────────────────────────────────────────
def load_crm_master(reg: CompanyRegistry) -> list[dict]:
    if not CRM_XLSX.exists():
        print(f"  [warn] CRM xlsx not found: {CRM_XLSX}")
        return []
    print(f"  reading {CRM_XLSX.name}...")
    aux = []
    wb = openpyxl.load_workbook(CRM_XLSX, read_only=True, data_only=True)
    # All Prospects: row 2 is header, row 3+ is data
    ws = wb["All Prospects"]
    rows = ws.iter_rows(values_only=True)
    section = next(rows, None)
    headers = next(rows, None)
    if not headers:
        return aux
    headers = [str(h).strip() if h else f"col{i}" for i, h in enumerate(headers)]
    idx = {h: i for i, h in enumerate(headers)}
    def g(r, key):
        i = idx.get(key)
        return r[i] if i is not None and i < len(r) else None
    n = 0
    for r in rows:
        if not r or not r[0]:
            continue
        n += 1
        name = str(r[0]).strip()
        if not name:
            continue
        zip5 = clean_zip(g(r, "Zip"))
        ee = clean_int(g(r, "Emp"))
        naics = None
        crow = {
            "name": name,
            "city": (g(r, "City") or "").strip() or None if g(r, "City") else None,
            "county": (g(r, "County") or "").strip() or None if g(r, "County") else None,
            "state": "NC",
            "zip": zip5,
            "ee": ee,
            "vertical": g(r, "Industry") or None,
            "website": g(r, "Website") or None,
            "linkedin_url": g(r, "LinkedIn") or None,
            "pitch_angle": g(r, "Pitch Angle") or None,
            "source": "crm_master",
        }
        disq, reason = is_disqualified(name, naics, zip5)
        if disq:
            crow["disqualified"] = True
            crow["disqualified_reason"] = reason
        reg.upsert(crow)
        # capture decision maker + phone as contacts
        dm = g(r, "Decision Maker")
        email = g(r, "Email")
        phone = clean_phone(g(r, "Phone"))
        if dm or email or phone:
            aux.append({
                "name": name, "zip": zip5,
                "kind": "contact",
                "dm_name": dm, "dm_email": email, "phone": phone,
            })
        # capture PEO Tier note (free-form memo)
        peo_score = g(r, "PEO Score")
        peo_tier = g(r, "PEO Tier")
        peo_reason = g(r, "PEO Reason")
        if peo_tier or peo_reason:
            aux.append({
                "name": name, "zip": zip5,
                "kind": "note",
                "body": f"CRM PEO Tier: {peo_tier} | Score: {peo_score} | Reason: {peo_reason}",
            })
    wb.close()
    return aux


# ─── writer: push everything to Neon ────────────────────────────────────────────
def write_to_neon(reg: CompanyRegistry, dry_run: bool, fresh: bool,
                   competing_aux: list[dict], sales_state: dict,
                   trigger_aux: list[dict], crm_aux: list[dict],
                   notes_aux: list[dict],
                   sqlite_contacts: list[tuple], sqlite_carriers: list[tuple],
                   sqlite_triggers: list[tuple], sqlite_incumbents: list[tuple]):
    if dry_run:
        print("\n[dry-run] skipping all DB writes.")
        print(f"  would insert: {len(reg):,} companies")
        return {}

    counts: dict[str, int] = {}
    with psycopg.connect(DSN) as conn:
        with conn.cursor() as cur:
            if fresh:
                print("  --fresh: dropping existing tables...")
                cur.execute(DROP_SQL)
            print("  creating schema (idempotent)...")
            cur.execute(SCHEMA_SQL)
            conn.commit()

            # ─── companies (bulk insert) ────────────────────────────
            print("  inserting companies...")
            ids = reg.assign_ids()
            comp_rows = []
            for key, row in reg.rows():
                comp_rows.append((
                    row.get("name"), row.get("name_normalized"),
                    row.get("city"), row.get("county"),
                    row.get("state") or "NC", row.get("zip"),
                    row.get("ein"), row.get("domain"),
                    row.get("website"), row.get("linkedin_url"),
                    row.get("naics"), row.get("vertical"),
                    row.get("ee"),
                    bool(row.get("multi_state_likely")),
                    bool(row.get("fed_contractor")),
                    bool(row.get("has_5500")),
                    bool(row.get("has_health_carriers")),
                    row.get("growth_signal"), row.get("pitch_signal"),
                    row.get("pitch_angle"),
                    bool(row.get("disqualified")),
                    row.get("disqualified_reason"),
                    row.get("source"),
                ))
            # truncate first if not fresh (so we don't get UNIQUE conflicts on rerun)
            cur.execute("TRUNCATE companies RESTART IDENTITY CASCADE")
            # use COPY-like fast path via executemany w/ batches
            BATCH = 5_000
            for i in range(0, len(comp_rows), BATCH):
                cur.executemany(
                    "INSERT INTO companies (name, name_normalized, city, county, state, zip, "
                    "ein, domain, website, linkedin_url, naics, vertical, ee, "
                    "multi_state_likely, fed_contractor, has_5500, has_health_carriers, "
                    "growth_signal, pitch_signal, pitch_angle, disqualified, disqualified_reason, source) "
                    "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                    comp_rows[i:i+BATCH]
                )
                if (i + BATCH) % 20_000 < BATCH:
                    print(f"    inserted: {min(i+BATCH, len(comp_rows)):,}/{len(comp_rows):,}")
            counts["companies"] = len(comp_rows)
            conn.commit()

            # build (name_norm, zip-or-state) -> Neon serial id  by reselecting
            print("  fetching Neon company ids for FK linking...")
            cur.execute("SELECT id, name_normalized, COALESCE(zip, state) FROM companies")
            neon_id: dict[tuple[str, str], int] = {}
            for cid, nn, zos in cur.fetchall():
                neon_id[(nn, zos)] = cid

            def fk(name: str, zip5: str | None) -> int | None:
                nn = norm_name(name)
                if not nn:
                    return None
                if zip5:
                    z = clean_zip(zip5)
                    if z and (nn, z) in neon_id:
                        return neon_id[(nn, z)]
                if (nn, "NC") in neon_id:
                    return neon_id[(nn, "NC")]
                # last-ditch: any zip with same name
                for (k_nn, k_z), v in neon_id.items():
                    if k_nn == nn:
                        return v
                return None

            # ─── carriers (sqlite) ────────────────────────────────
            print("  inserting carriers...")
            if sqlite_carriers:
                # remap from "synthetic" id-by-assignment to real Neon id
                ids_by_pos = list(ids.values())
                pos_to_neon: dict[int, int] = {}
                # ids[key] returned 1..N matching insertion order of comp_rows
                # since we TRUNCATE+INSERT in the same loop, the Neon serials should be 1..N
                # (with RESTART IDENTITY this is guaranteed)
                # so the synthetic id == Neon id. no remap needed.
                carriers_to_insert = sqlite_carriers
                cur.executemany(
                    "INSERT INTO carriers (company_id, carrier_name, benefit_type, naic, "
                    "premium, covered_lives, plan_year, source) "
                    "VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
                    carriers_to_insert
                )
                counts["carriers"] = len(carriers_to_insert)
            else:
                counts["carriers"] = 0

            # ─── contacts (sqlite + CRM + competing_peo_csv) ──────
            print("  inserting contacts...")
            contact_rows = list(sqlite_contacts)  # already (company_id, ...) tuples
            for a in competing_aux:
                cid = fk(a["name"], a.get("zip"))
                if not cid:
                    continue
                if a.get("dm_name") or a.get("dm_email") or a.get("phone"):
                    contact_rows.append((
                        cid, a.get("dm_name"), a.get("dm_role"),
                        a.get("dm_email"), a.get("phone"), None, True, "competing_peo_csv",
                    ))
            for a in crm_aux:
                if a.get("kind") != "contact":
                    continue
                cid = fk(a["name"], a.get("zip"))
                if not cid:
                    continue
                contact_rows.append((
                    cid, a.get("dm_name"), None,
                    a.get("dm_email"), a.get("phone"), None, True, "crm_master",
                ))
            for i in range(0, len(contact_rows), BATCH):
                cur.executemany(
                    "INSERT INTO contacts (company_id, name, title, email, phone, linkedin_url, is_primary, source) "
                    "VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
                    contact_rows[i:i+BATCH]
                )
            counts["contacts"] = len(contact_rows)
            conn.commit()

            # ─── triggers (sqlite + trigger_engine) ───────────────
            print("  inserting triggers...")
            trigger_rows = list(sqlite_triggers)
            for a in trigger_aux:
                cid = fk(a["name"], a.get("zip"))
                if not cid:
                    continue
                trigger_rows.append((
                    cid, a["trigger_type"], a.get("score"),
                    a.get("evidence"), a.get("trigger_date"), "trigger_engine_xlsx",
                ))
            for i in range(0, len(trigger_rows), BATCH):
                cur.executemany(
                    "INSERT INTO triggers (company_id, trigger_type, score, evidence, trigger_date, source) "
                    "VALUES (%s,%s,%s,%s,%s,%s)",
                    trigger_rows[i:i+BATCH]
                )
            counts["triggers"] = len(trigger_rows)
            conn.commit()

            # ─── incumbent_peo (sqlite + competing_peo_csv + sales_os) ──
            print("  inserting incumbent_peo...")
            incumbent_rows = list(sqlite_incumbents)
            for a in competing_aux:
                cid = fk(a["name"], a.get("zip"))
                if not cid or not a.get("current_peo"):
                    continue
                incumbent_rows.append((
                    cid, a["current_peo"], a.get("fitness_tier"),
                    a.get("current_peo_evidence"), None, "competing_peo_csv",
                ))
            for c in sales_state.get("active_cadences", []):
                if not c.get("incumbent_peo"):
                    continue
                cid = fk(c.get("company") or "", c.get("zip"))
                if not cid:
                    continue
                incumbent_rows.append((
                    cid, c["incumbent_peo"], None,
                    None, None, "sales_os",
                ))
            for i in range(0, len(incumbent_rows), BATCH):
                cur.executemany(
                    "INSERT INTO incumbent_peo (company_id, peo_brand, confidence, evidence, filing_year, source) "
                    "VALUES (%s,%s,%s,%s,%s,%s)",
                    incumbent_rows[i:i+BATCH]
                )
            counts["incumbent_peo"] = len(incumbent_rows)
            conn.commit()

            # ─── cadences + touches (sales_os) ────────────────────
            print("  inserting cadences + touches...")
            cad_rows = []
            for c in sales_state.get("active_cadences", []):
                cid = fk(c.get("company") or "", c.get("zip"))
                cad_rows.append((
                    cid, c["key"], c.get("status") or "active",
                    c.get("start_date"), c.get("route_day"),
                    c.get("score"), c.get("score_raw"), c.get("score_overlay"),
                    c.get("weight_mult"), c.get("primary_trigger"),
                    c.get("tier"), c.get("fitness_tier"),
                    None, c.get("evidence"), c.get("talk_track"),
                    c.get("vertical"),
                ))
            cur.executemany(
                "INSERT INTO cadences (company_id, company_key, status, start_date, route_day, "
                "score, score_raw, score_overlay, weight_mult, primary_trigger, tier, fitness_tier, "
                "enrichment_notes, evidence, talk_track, vertical) "
                "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                cad_rows
            )
            counts["cadences"] = len(cad_rows)
            # touches need cadence ids — fetch them by company_key
            cur.execute("SELECT id, company_key FROM cadences")
            cad_id_by_key: dict[str, int] = {ck: cid for cid, ck in cur.fetchall()}
            touch_rows = []
            for c in sales_state.get("active_cadences", []):
                cad_id = cad_id_by_key.get(c["key"])
                if not cad_id:
                    continue
                for t in c.get("touches", []):
                    touch_rows.append((
                        cad_id, t.get("day"), t.get("channel"),
                        t.get("scheduled_for"), bool(t.get("completed")),
                        t.get("outcome") or None, t.get("notes") or None,
                        t.get("broker_captured") or None,
                    ))
            cur.executemany(
                "INSERT INTO touches (cadence_id, day_offset, channel, scheduled_for, "
                "completed, outcome, notes, broker_captured) "
                "VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
                touch_rows
            )
            counts["touches"] = len(touch_rows)
            conn.commit()

            # ─── outcomes_ledger ──────────────────────────────────
            print("  inserting outcomes_ledger...")
            ol_rows = []
            for o in sales_state.get("weekly_outcomes", []):
                ol_rows.append((
                    None, None, o.get("company_key"),
                    o.get("week_start"), o.get("logged_at"),
                    o.get("trigger"), o.get("vertical"),
                    o.get("channel"), o.get("day_offset"), o.get("route_day"),
                    o.get("scheduled_for"), o.get("outcome"),
                    o.get("broker_captured"), o.get("notes"),
                ))
            if ol_rows:
                cur.executemany(
                    "INSERT INTO outcomes_ledger (touch_id, cadence_id, company_key, "
                    "week_start, logged_at, trigger_type, vertical, channel, "
                    "day_offset, route_day, scheduled_for, outcome, broker_captured, notes) "
                    "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                    ol_rows
                )
            counts["outcomes_ledger"] = len(ol_rows)

            # ─── notes ────────────────────────────────────────────
            print("  inserting notes...")
            note_rows = []
            for n in notes_aux:
                cid = fk(n["name"], None)
                if not cid:
                    continue
                note_rows.append((cid, n["body"], n.get("source") or "verification"))
            for a in crm_aux:
                if a.get("kind") != "note":
                    continue
                cid = fk(a["name"], a.get("zip"))
                if not cid:
                    continue
                note_rows.append((cid, a["body"], "crm_master"))
            # competing PEO talk tracks → notes
            for a in competing_aux:
                if not a.get("talk_track"):
                    continue
                cid = fk(a["name"], a.get("zip"))
                if not cid:
                    continue
                body = f"Talk track: {a['talk_track']}"
                if a.get("signals_seen"):
                    body += f"\nSignals: {a['signals_seen']}"
                if a.get("fitness_score") is not None:
                    body += f"\nFitness score: {a['fitness_score']}"
                note_rows.append((cid, body, "competing_peo_csv"))
            for i in range(0, len(note_rows), BATCH):
                cur.executemany(
                    "INSERT INTO notes (company_id, body, source) VALUES (%s,%s,%s)",
                    note_rows[i:i+BATCH]
                )
            counts["notes"] = len(note_rows)
            counts["tasks"] = 0  # no source today; table provisioned for future use
            conn.commit()

            # ─── meddpicc ─────────────────────────────────────────
            print("  inserting meddpicc...")
            mp_rows = []
            for ck, m in (sales_state.get("meddpicc") or {}).items():
                cid = fk(m.get("company") or "", None) if isinstance(m, dict) else None
                if isinstance(m, dict):
                    mp_rows.append((
                        cid, ck, m.get("first_meeting_date"),
                        m.get("stage"), m.get("M_metrics"), m.get("E_econ_buyer"),
                        m.get("D1_decision_criteria"), m.get("D2_decision_process"),
                        m.get("P_paper_process"), m.get("I_pain"),
                        m.get("C_champion"), m.get("Cmp_competition"),
                        m.get("next_action"),
                    ))
            if mp_rows:
                cur.executemany(
                    "INSERT INTO meddpicc (company_id, company_key, first_meeting_date, stage, "
                    "m_metrics, e_econ_buyer, d1_decision_criteria, d2_decision_process, "
                    "p_paper_process, i_pain, c_champion, cmp_competition, next_action) "
                    "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                    mp_rows
                )
            counts["meddpicc"] = len(mp_rows)

            # ─── brokers ──────────────────────────────────────────
            print("  inserting brokers...")
            br_rows = []
            for b in (sales_state.get("channel_build") or {}).get("brokers", []):
                if not b.get("name"):
                    continue
                br_rows.append((
                    b["name"], b.get("county"), b.get("phone"), b.get("email"),
                    b.get("first_seen"), b.get("first_seen_via"),
                    b.get("last_brief_sent") or None,
                    b.get("lunch_status"),
                    b.get("n_clients_observed"), b.get("notes"),
                ))
            if br_rows:
                cur.executemany(
                    "INSERT INTO brokers (name, county, phone, email, first_seen, first_seen_via, "
                    "last_brief_sent, lunch_status, n_clients_observed, notes) "
                    "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) ON CONFLICT (name) DO NOTHING",
                    br_rows
                )
            counts["brokers"] = len(br_rows)

            # ─── weights_current ──────────────────────────────────
            print("  inserting weights_current...")
            w = sales_state.get("weights") or {}
            last = w.get("last_recomputed")
            nout = w.get("n_outcomes_at_recompute", 0) or 0
            wr_rows = []
            for dim in ("trigger", "vertical", "channel", "route_day"):
                for k, mult in (w.get(dim) or {}).items():
                    wr_rows.append((dim, str(k), mult, last, nout))
            if wr_rows:
                cur.executemany(
                    "INSERT INTO weights_current (dim, key, mult, last_recomputed, n_outcomes) "
                    "VALUES (%s,%s,%s,%s,%s) ON CONFLICT (dim, key) DO UPDATE SET "
                    "mult = EXCLUDED.mult, last_recomputed = EXCLUDED.last_recomputed",
                    wr_rows
                )
            counts["weights_current"] = len(wr_rows)

            # ─── weekly_metrics ───────────────────────────────────
            print("  inserting weekly_metrics...")
            wm_rows = []
            for ws_key, m in (sales_state.get("weekly_metrics") or {}).items():
                if not isinstance(m, dict):
                    continue
                wm_rows.append((
                    ws_key,
                    m.get("touches") or 0,
                    m.get("progressions") or 0,
                    m.get("meetings_booked") or 0,
                    m.get("meetings_held") or 0,
                    m.get("killed") or 0,
                ))
            if wm_rows:
                cur.executemany(
                    "INSERT INTO weekly_metrics (week_start, touches, progressions, "
                    "meetings_booked, meetings_held, killed) "
                    "VALUES (%s,%s,%s,%s,%s,%s) ON CONFLICT (week_start) DO UPDATE SET "
                    "touches = EXCLUDED.touches",
                    wm_rows
                )
            counts["weekly_metrics"] = len(wm_rows)

            conn.commit()

    return counts


# ─── main ───────────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--fresh", action="store_true", help="DROP existing tables before recreating")
    ap.add_argument("--dry-run", action="store_true", help="Parse + dedupe, but skip all DB writes")
    args = ap.parse_args()

    t0 = time.time()
    print("ADP PEO → Neon migration")
    print("=" * 60)

    reg = CompanyRegistry()
    per_source: dict[str, int] = {}

    # 1. pipeline.sqlite
    print("[1/6] pipeline.sqlite")
    n_seen, n_kept = load_sqlite(reg)
    per_source["sqlite_companies_seen"] = n_seen
    per_source["sqlite_companies_kept_in_reg"] = n_kept

    # collect contacts/carriers/triggers/incumbents from sqlite (FK to companies)
    sq_contacts, sq_carriers, sq_triggers, sq_incumbents, _ = collect_sqlite_aux(reg)
    per_source["sqlite_contacts"] = len(sq_contacts)
    per_source["sqlite_carriers"] = len(sq_carriers)
    per_source["sqlite_triggers"] = len(sq_triggers)
    per_source["sqlite_incumbents"] = len(sq_incumbents)

    # 2. competing PEO CSV
    print("[2/6] competing PEO CSV")
    n_csv, comp_aux = load_competing_peo_csv(reg)
    per_source["competing_peo_csv_rows"] = n_csv

    # 3. sales_os state.json
    print("[3/6] _sales_os_state.json")
    sales_state = load_sales_os(reg)
    per_source["sales_os_active_cadences"] = len(sales_state.get("active_cadences", []))

    # 4. trigger engine
    print("[4/6] Trigger_Engine_Output.xlsx")
    trig_aux = load_trigger_engine(reg)
    per_source["trigger_engine_rows"] = len(trig_aux)

    # 5. verification
    print("[5/6] verification_V2.json")
    notes_aux = load_verification(reg)
    per_source["verification_v2_rows"] = len(notes_aux)

    # 6. CRM master
    print("[6/6] ADP_Territory_CRM_Master.xlsx")
    crm_aux = load_crm_master(reg)
    per_source["crm_master_aux"] = len(crm_aux)

    print()
    print(f"Total deduped companies in registry: {len(reg):,}")

    # write to Neon
    print()
    print("Writing to Neon...")
    counts = write_to_neon(
        reg, args.dry_run, args.fresh,
        comp_aux, sales_state, trig_aux, crm_aux, notes_aux,
        sq_contacts, sq_carriers, sq_triggers, sq_incumbents,
    )

    # final report
    print()
    print("─" * 60)
    print("Per-source rows scanned:")
    for k, v in per_source.items():
        print(f"  {k:40} {v:>10,}")
    print()
    if counts:
        print("Final Neon row counts:")
        for k, v in counts.items():
            print(f"  {k:25} {v:>10,}")
    print()
    print(f"Total elapsed: {time.time()-t0:.1f}s")


if __name__ == "__main__":
    main()
